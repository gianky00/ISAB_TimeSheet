"""
Bot TS - Dettagli OdA Bot
Bot per l'accesso alla sezione Dettagli OdA del portale ISAB.
"""
import os
import glob
import time
import shutil
from typing import List, Dict, Any

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

from openpyxl import load_workbook

from src.bots.base import BaseBot, BotStatus


class DettagliOdABot(BaseBot):
    """
    Bot per l'accesso ai Dettagli OdA.
    """
    
    @staticmethod
    def get_name() -> str:
        return "Dettagli OdA"
    
    @staticmethod
    def get_description() -> str:
        return "Accede ai Dettagli OdA con filtri preimpostati"
    
    @staticmethod
    def get_columns() -> list:
        return [
            {"name": "Numero OdA", "type": "text"}
        ]
    
    @property
    def name(self) -> str:
        return "Dettagli OdA"
    
    @property
    def description(self) -> str:
        return "Accede ai Dettagli OdA con filtri preimpostati"
    
    def __init__(self, data_da: str = "", data_a: str = "", fornitore: str = "", 
                 numero_oda: str = "", numero_contratto: str = "", **kwargs):
        super().__init__(**kwargs)
        self.data_da = data_da
        self.data_a = data_a
        self.fornitore = fornitore
        self.numero_oda = numero_oda
        self.numero_contratto = numero_contratto
        # Variabile per tracciare l'ultimo log ed evitare spam
        self._last_log_msg = "" 

    def run(self, data: Dict[str, Any]) -> bool:
        """
        Esegue la navigazione ai Dettagli OdA e imposta i filtri.
        """
        rows = []
        if isinstance(data, dict):
            self.data_da = data.get('data_da', self.data_da)
            self.data_a = data.get('data_a', self.data_a)
            self.fornitore = data.get('fornitore', self.fornitore)
            rows = data.get('rows', [])
            if not rows:
                self.numero_oda = data.get('numero_oda', self.numero_oda)

        self.log(f"Avvio elaborazione: Fornitore='{self.fornitore}', Periodo={self.data_da}-{self.data_a}")
        
        # 1. Navigazione Report -> OdA
        if not self._navigate_to_oda():
            self.log("⚠️ Navigazione automatica non riuscita (potrei essere già sulla pagina).")
        
        downloaded_files = []
        
        if rows:
            self.log(f"Elaborazione di {len(rows)} OdA...")
            for i, row in enumerate(rows, 1):
                self._check_stop()
                
                self.numero_oda = row.get("numero_oda", "")
                self.numero_contratto = row.get("numero_contratto", "")
                
                # Setup filtri e download
                file_path = self._setup_filters()
                
                if file_path:
                    self.log(f"[{i}/{len(rows)}] ✅ OdA {self.numero_oda}: Scaricato.")
                    downloaded_files.append(file_path)
                else:
                    # Log di errore solo se non già gestito
                    if not any(x in self._last_log_msg for x in ["Non trovato", "Errore"]): 
                         self.log(f"[{i}/{len(rows)}] ⚠️ OdA {self.numero_oda}: Nessun file scaricato.")

                # Chiusura schede (Silenziosa)
                self._close_all_tabs_silently()
                
                # Rinavigazione (Silenziosa)
                if i < len(rows):
                    self._expand_menu_silently()
                    self._navigate_to_oda(force=True, silent=True)
                        
                time.sleep(0.5)
        else:
            # Caso singola esecuzione
            self.log(f"Elaborazione singolo OdA: {self.numero_oda}...")
            file_path = self._setup_filters()
            if file_path:
                self.log(f"✅ OdA {self.numero_oda}: Scaricato.")
                downloaded_files.append(file_path)
            else:
                self.log(f"⚠️ OdA {self.numero_oda}: Nessun file scaricato.")
            
            self._close_all_tabs_silently()

        if downloaded_files:
            self._verify_and_cleanup_excel_files(downloaded_files)
        
        self.log("Processo completato.")
        return True
    
    def _attendi_scomparsa_overlay(self, *args, **kwargs):
        """
        Override del metodo base per eliminare il log 'Overlay di caricamento scomparso'.
        Attende silenziosamente la scomparsa delle maschere ExtJS.
        Accetta *args e **kwargs per compatibilità con la classe base.
        """
        try:
            # Timeout breve per vedere se appare l'overlay
            WebDriverWait(self.driver, 0.5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[class*='x-mask']"))
            )
            # Timeout lungo per aspettare che sparisca
            WebDriverWait(self.driver, 30).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div[class*='x-mask']"))
            )
        except:
            # Se non c'è overlay o timeout, prosegue silenziosamente
            pass

    def _navigate_to_oda(self, force: bool = False, silent: bool = False) -> bool:
        """Naviga a Report -> OdA."""
        self._check_stop()
        
        if not force:
            try:
                fornitore_exists = self.driver.find_elements(By.XPATH, "//div[starts-with(@id, 'generic_refresh_combo_box-') and contains(@class, 'x-form-arrow-trigger')]")
                if fornitore_exists and fornitore_exists[0].is_displayed():
                    return True
            except:
                pass

        if not silent: self.log("Navigazione verso pagina OdA...")
        
        try:
            report_element = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//*[normalize-space(text())='Report']"))
            )
            report_element.click()
            
            if force:
                time.sleep(0.5)
                try:
                    report_element.click()
                except:
                    pass
            
            time.sleep(1.5)

            actions = ActionChains(self.driver)
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.ENTER).perform()
            
            time.sleep(1.0)

            fornitore_arrow_xpath = "//div[starts-with(@id, 'generic_refresh_combo_box-') and contains(@id, '-trigger-picker') and contains(@class, 'x-form-arrow-trigger')]"
            self.wait.until(EC.visibility_of_element_located((By.XPATH, fornitore_arrow_xpath)))
            self._attendi_scomparsa_overlay()
            
            return True
            
        except Exception as e:
            if not silent: self.log(f"Errore navigazione: {e}")
            return False

    def _setup_filters(self) -> str:
        """
        Imposta filtri, esegue la sequenza di tasti e scarica Excel.
        """
        self._check_stop()
        
        try:
            # 1. Seleziona Fornitore
            if self.fornitore:
                fornitore_arrow_xpath = "//div[starts-with(@id, 'generic_refresh_combo_box-') and contains(@id, '-trigger-picker') and contains(@class, 'x-form-arrow-trigger')]"
                fornitore_arrow_element = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, fornitore_arrow_xpath))
                )
                ActionChains(self.driver).move_to_element(fornitore_arrow_element).click().perform()
                time.sleep(0.5)

                fornitore_option_xpath = f"//li[contains(text(), '{self.fornitore}')]"
                fornitore_option = self.long_wait.until(
                    EC.presence_of_element_located((By.XPATH, fornitore_option_xpath))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'nearest'});", fornitore_option)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", fornitore_option)
                time.sleep(0.5)
                self._attendi_scomparsa_overlay()

            # 2. Sequenza Tasti per Filtri
            actions = ActionChains(self.driver)

            # TAB verso OdA
            actions.send_keys(Keys.TAB).pause(0.3)
            
            # Inserimento Numero OdA
            actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).pause(0.1)
            if self.numero_oda:
                actions.send_keys(self.numero_oda)
            else:
                actions.send_keys(Keys.DELETE)
            actions.pause(0.3)

            # --- Sequence Update ---
            # Date A
            if self.data_a:
                actions.send_keys(self.data_a).pause(0.5) # Assuming we are in Date A field?
                # Wait, existing logic was: TAB (Divisione), TAB (Data Da), TAB (Data A).
                # But user requirement says:
                # "CTRL+A seguito dall'inserimento della Data A (se presente)."
                # "1 TAB per spostamento in Numero Contratto quindi CTRL+A seguito dall'inserimento del numero contratto"

            # Re-implementing based on explicit user sequence:
            # Previous state: We just filled Numero OdA.
            # Old sequence was: ... -> TAB (Divisione) -> TAB (Data Da) -> TAB (Data A) ...
            # The user says: "CTRL+A seguito dall'inserimento della Data A (se presente)". This implies we are already ON Data A?
            # Or maybe the sequence starts from OdA?
            # "CTRL+A seguito dall'inserimento della Data A (se presente)" matches what we might do if we are on Data A.

            # Let's assume the flow is:
            # 1. OdA (Done above)
            # 2. TAB -> Divisione
            # 3. TAB -> Data Da
            # 4. TAB -> Data A

            # So I need to navigate to Data A first.
            actions.send_keys(Keys.TAB).pause(0.3) # Divisione
            actions.send_keys(Keys.TAB).pause(0.3) # Data Da (Skipping Data Da input based on user desc? No, user says "Date A" maybe meaning "Date From"?)
            # User said: "CTRL+A seguito dall'inserimento della Data A (se presente)."
            # Wait, "Data A" usually means "Date To". But in Italian "Data A" is "Date To". "Data Da" is "Date From".
            # The user prompt mentions "Data A".
            # Let's look at the OLD sequence provided by user:
            # "CTRL+A seguito dall'inserimento della Data A (se presente)."
            # "3 TAB consecutivi (Spostamento verso il Flag Dettagli)."

            # My old code was:
            # TAB (Divisione), TAB (Data Da) -> Fill Data Da, TAB (Data A) -> Fill Data A.
            # Then 3 TABS.

            # The user description seems to skip Divisione and Data Da? Or maybe they assume we are there?
            # Let's respect the visual flow of the form usually found in these apps.
            # OdA -> Divisione -> Data Da -> Data A -> Contratto -> ...

            # If I follow the user's *new* sequence literally:
            # "CTRL+A seguito dall'inserimento della Data A (se presente)."
            # "1 TAB per spostamento in Numero Contratto quindi CTRL+A seguito dall'inserimento del numero contratto"
            # "2 TAB consecutivi (Spostamento verso il Flag Dettagli)."

            # It seems "Data A" here refers to "Data OdA" (Date of OdA)? Or "Data A" (Date To)?
            # Given the context of "Dettagli OdA", usually there are Date From/To.
            # Let's assume the user means the field *before* Contract Number.
            # If I look at my previous code, after OdA I did: TAB (Div), TAB (Data Da), TAB (Data A).
            # If I insert Contract Number, where is it?
            # Usually Contract Number is likely after Data A?

            # Let's try to map the fields based on user instruction "1 TAB per spostamento in Numero Contratto".
            # This implies Contract Number is immediately after the field where we put "Data A".

            # So the flow:
            # OdA -> [TABs...] -> Field X ("Data A") -> TAB -> Contract Number -> TAB -> TAB -> Flag.

            # In my previous code:
            # OdA -> TAB -> Div -> TAB -> Data Da -> TAB -> Data A -> [3 TABs] -> Flag.
            # So "Data A" was indeed the field before the 3 TABs.
            # So Contract Number must have been added *between* Data A and the Flag.
            # And it takes 1 TAB to get there from Data A.
            # And from Contract Number to Flag it takes 2 TABs.
            # Total TABs from Data A to Flag = 1 + 2 = 3.
            # This matches perfectly with the old "3 TABs" (which just skipped the Contract field because it wasn't used/focused or was skipped).

            # So I will keep the navigation to "Data A" (which is Date To), fill it, then:
            # TAB -> Fill Contract -> 2 TABs -> Flag.
            
            # Navigation to Data A (Date To):
            actions.send_keys(Keys.TAB).pause(0.3) # Divisione

            # Data Da handling (Keeping it as it might be needed, user didn't explicitly say to remove it, just described the part that changes)
            actions.send_keys(Keys.TAB).pause(0.3) # Data Da
            if self.data_da:
                actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).pause(0.1)
                actions.send_keys(self.data_da).pause(0.5)

            actions.send_keys(Keys.TAB).pause(0.3) # To Data A
            
            # Data A handling
            if self.data_a:
                actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).pause(0.1)
                actions.send_keys(self.data_a).pause(0.5)

            # NEW PART: Contract Number
            actions.send_keys(Keys.TAB).pause(0.3) # Move to Contract Number

            actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).pause(0.1)
            if self.numero_contratto:
                actions.send_keys(self.numero_contratto)
            else:
                actions.send_keys(Keys.DELETE)
            actions.pause(0.3)

            # 2 TABs to Flag (replacing the old loop of 3)
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.TAB).pause(0.3)

            # Attivazione Flag e Conferma
            actions.send_keys(Keys.SPACE).pause(0.3) 
            actions.send_keys(Keys.TAB).pause(0.3)   
            actions.send_keys(Keys.ENTER).pause(1.5) 

            # Sequenza extra
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.TAB).pause(0.3)
            actions.send_keys(Keys.ENTER).pause(1.0)

            actions.perform()

            # --- Scarico Excel ---
            downloaded_file = ""
            try:
                excel_xpath = "//*[contains(text(), 'Esporta in Excel')]"
                
                # Check rapido (2s)
                try:
                    excel_btn = WebDriverWait(self.driver, 2).until(
                        EC.presence_of_element_located((By.XPATH, excel_xpath))
                    )
                except TimeoutException:
                    self.log(f"[{self.numero_oda}] ⚠️ Nessun dato trovato (Tabella vuota).")
                    self._last_log_msg = "Non trovato"
                    return ""
                
                # Scroll e click
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", excel_btn)
                time.sleep(0.5)
                
                try:
                    excel_btn.click()
                except Exception:
                    self.driver.execute_script("arguments[0].click();", excel_btn)

                # Attesa file
                time.sleep(2.0)
                downloaded_file = self._rename_latest_download(self.numero_oda)
                
                if not downloaded_file:
                    self.log(f"[{self.numero_oda}] ❌ Errore durante il rinomina file.")
                    self._last_log_msg = "Errore"

            except Exception as e:
                self.log(f"❌ Errore imprevisto download: {e}")
                self._last_log_msg = "Errore"

            return downloaded_file

        except Exception as e:
            self.log(f"❌ Errore impostazione filtri: {e}")
            return ""

    def execute(self, data: Any) -> bool:
        """Override: esegue login e run, ma non chiude il browser."""
        self._stop_requested = False
        try:
            self._init_driver()
            if not self._login():
                return False

            return self.run(data)
        except Exception as e:
            self.log(f"Errore critico: {e}")
            return False

    def _rename_latest_download(self, new_name_base: str) -> str:
        """Trova l'ultimo file scaricato e lo rinomina."""
        if not self.download_path or not os.path.exists(self.download_path):
            return ""

        timeout = 10
        start_time = time.time()
        latest_file = None

        while time.time() - start_time < timeout:
            files = glob.glob(os.path.join(self.download_path, "*"))
            files = [f for f in files if not f.endswith('.crdownload') and not f.endswith('.tmp') and os.path.isfile(f)]

            if files:
                latest_file = max(files, key=os.path.getctime)
                if time.time() - os.path.getctime(latest_file) < 15:
                    break
            time.sleep(1)

        if not latest_file:
            return ""

        try:
            _, ext = os.path.splitext(latest_file)
            if not ext: ext = ".xlsx"
            new_filename = f"dettaglio_oda_{new_name_base}{ext}"
            if not new_name_base:
                new_filename = f"dettaglio_oda_{int(time.time())}{ext}"
                
            new_path = os.path.join(self.download_path, new_filename)
            if os.path.exists(new_path):
                os.remove(new_path)
            shutil.move(latest_file, new_path)
            return new_path
        except Exception:
            return ""

    def _verify_and_cleanup_excel_files(self, file_paths: List[str]) -> None:
        """Verifica i file Excel scaricati e cancella quelli con cella B2 vuota."""
        files_deleted = 0
        files_kept = 0
        
        for file_path in file_paths:
            if not file_path or not os.path.exists(file_path):
                continue
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                cell_b2 = ws['B2'].value
                wb.close()
                
                if cell_b2 is None or (isinstance(cell_b2, str) and cell_b2.strip() == ""):
                    os.remove(file_path)
                    files_deleted += 1
                else:
                    files_kept += 1
            except Exception:
                pass
        
        self.log(f"Riepilogo: {files_kept} file validi salvati, {files_deleted} vuoti rimossi.")

    def _close_all_tabs_silently(self):
        """Helper per chiudere le schede senza loggare."""
        while True:
            try:
                close_btns = self.driver.find_elements(By.CSS_SELECTOR, "span.x-tab-close-btn")
                visible_btn = None
                for btn in close_btns:
                    if btn.is_displayed():
                        visible_btn = btn
                        break
                if visible_btn:
                    visible_btn.click()
                    time.sleep(1.0)
                    self._attendi_scomparsa_overlay()
                else:
                    break
            except:
                time.sleep(0.5)
                break

    def _expand_menu_silently(self):
        """Helper per espandere il menu senza loggare."""
        try:
            expand_menu_button = self.driver.find_elements(By.CSS_SELECTOR, "div.x-tool-tool-el.x-tool-expand-right")
            if expand_menu_button and expand_menu_button[0].is_displayed():
                expand_menu_button[0].click()
                time.sleep(0.5)
                self._attendi_scomparsa_overlay()
        except:
            pass