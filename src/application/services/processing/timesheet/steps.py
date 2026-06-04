"""Passaggi di elaborazione per la trasformazione strutturale dei file Timesheet."""

from __future__ import annotations

import time
from contextlib import suppress
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from src.application.services.processing.base import ProcessingStep
from src.domain.timesheet import TimesheetMetadata
from src.infrastructure.utils.secure_logger import get_secure_logger

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = get_secure_logger("TimesheetSteps")


class LoadWorkbookStep(ProcessingStep):
    """Passaggio per caricare il file Excel."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il caricamento del workbook."""
        file_path = Path(context["file_path"])
        if not file_path.exists():
            raise FileNotFoundError(f"File non trovato: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        context["workbook"] = wb
        if "Timesheet" not in wb.sheetnames:
            raise ValueError("Foglio 'Timesheet' non trovato.")
        context["worksheet"] = wb["Timesheet"]


class ExtractMetadataStep(ProcessingStep):
    """Passaggio per estrarre ODC e POS."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue l'estrazione dei metadati."""
        ws: Worksheet = context["worksheet"]

        # ODC
        odc = str(ws["A2"].value).strip() if ws["A2"].value else ""
        if not odc:
            odc = self._deduce_odc_from_filename(context["file_path"])
            if not odc:
                raise ValueError("Valore ODC (cella A2) mancante.")

        # Controlla se il foglio ha righe di dati valide
        if self._is_sheet_empty(ws):
            context["is_empty"] = True
            context["metadata"] = TimesheetMetadata(odc=odc, pos_values=set(), first_pos_cleaned="")
            logger.info("File Timesheet privo di record di dati.")
            return

        context["is_empty"] = False

        # POS
        pos_values: set[str] = set()
        first_pos_cleaned = ""

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            val = str(row[0].value).strip() if row[0].value is not None else ""
            if val:
                pos_values.add(val)
                if not first_pos_cleaned:
                    first_pos_cleaned = self._clean_pos_value(val)

        context["metadata"] = TimesheetMetadata(
            odc=odc, pos_values=pos_values, first_pos_cleaned=first_pos_cleaned
        )

    def _clean_pos_value(self, val: str | None) -> str:
        """Converte un valore POS in stringa intera pulita."""
        if val is None:
            return ""
        s = val.strip()
        if s.replace(".", "", 1).isdigit():
            with suppress(Exception):
                return str(int(float(s)))
        return val

    def _deduce_odc_from_filename(self, file_path_str: str | Path) -> str:
        """Deduce l'ODC dal nome del file sorgente.

        Args:
          file_path_str: Percorso del file Excel.

        Returns:
          str: Il codice ODC estratto.
        """
        file_path = Path(file_path_str)
        name_parts = file_path.stem.split("_")
        min_parts_len = 2
        if len(name_parts) >= min_parts_len:
            return name_parts[1].split("-")[0].strip()
        return ""

    def _is_sheet_empty(self, ws: Worksheet) -> bool:
        """Verifica se il foglio Excel non contiene righe di dati valide.

        Args:
          ws: Il foglio di lavoro da controllare.

        Returns:
          bool: True se il foglio è considerato vuoto.
        """
        min_rows_with_data = 2
        if ws.max_row < min_rows_with_data:
            return True
        # Controlla se c'è almeno un valore non-None nelle prime 15 colonne della riga 2
        max_cols_to_check = 15
        return all(ws.cell(row=2, column=col).value is None for col in range(1, max_cols_to_check + 1))


class TransformSheetStep(ProcessingStep):
    """Passaggio per applicare le trasformazioni strutturali."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue la trasformazione del foglio."""
        if context.get("is_empty", False):
            logger.info("Salto trasformazione foglio vuoto.")
            return

        ws: Worksheet = context["worksheet"]

        # 1. Rinomina Intestazioni
        headers = {
            "B1": "POS",
            "C1": "Data",
            "N1": "Ing",
            "O1": "Usc",
            "P1": "Tot",
            "Q1": "Pre",
            "R1": "ORE_C",
            "S1": "ORE_M",
            "T1": "ORE_ST_NOT",
            "U1": "ORE_ST_DIU",
            "V1": "ORE_FEST_NOT",
            "W1": "ORE_FEST_DIU",
        }
        for cell_coord, val in headers.items():
            ws[cell_coord].value = val

        # 2. Pulizia numerica colonna B
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            cell = cast("Any", row[0])
            if cell.value is not None:
                s_val = str(cell.value).strip()
                if s_val.replace(".", "", 1).isdigit():
                    with suppress(Exception):
                        cell.value = int(float(s_val))
                        cell.number_format = "0"

        # 3. Eliminazione Colonne (ordine inverso)
        cols = ["AC", "Z", "X", "L", "I", "H", "G", "F", "E", "D", "A"]
        indices = sorted([column_index_from_string(c) for c in cols], reverse=True)
        for idx in indices:
            ws.delete_cols(idx)

        # 4. Autofit stimato
        self._autofit_columns(ws)

    def _autofit_columns(self, ws: Worksheet) -> None:
        """Regola la larghezza delle colonne in base al contenuto.

        Args:
          ws: Il foglio di lavoro da elaborare.
        """
        for col in ws.columns:
            max_len = 0
            first_cell = cast("Any", col[0])
            column_idx = first_cell.column
            if isinstance(column_idx, int):
                col_letter = get_column_letter(column_idx)
                for cell in col:
                    with suppress(Exception):
                        val_len = len(str(cell.value))
                        max_len = max(max_len, val_len)
                ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2


class SaveWorkbookStep(ProcessingStep):
    """Passaggio per salvare il file finale."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il salvataggio del workbook."""
        wb: Workbook = context["workbook"]
        dest_dir = Path(context["dest_dir"])
        metadata: TimesheetMetadata = context["metadata"]

        if context.get("is_empty", False):
            # Rimuove il file temporaneo sorgente per non lasciare residui
            file_path = Path(context["file_path"])
            with suppress(Exception):
                file_path.unlink()
            context["dest_path"] = None
            logger.info("Rimosso file temporaneo Timesheet vuoto.")
            return

        base = (
            f"{metadata.odc}_TS"
            if len(metadata.pos_values) > 1
            else f"{metadata.odc}_{metadata.first_pos_cleaned}_TS"
        )

        dest_path = dest_dir / f"{base}.xlsx"
        if dest_path.exists():
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            dest_path = dest_dir / f"{base}_{timestamp}.xlsx"

        try:
            dest_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            raise PermissionError("Impossibile creare dest_dir") from None

        wb.save(dest_path)
        context["dest_path"] = dest_path
