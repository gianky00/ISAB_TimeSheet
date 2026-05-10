# mypy: disable-error-code="no-untyped-def, no-untyped-call, unused-ignore, arg-type"
"""
SyncroJob - Consuntivo Modifica Esistente Tab
Tab intelligente per la scansione, auto-fill e manipolazione di file esistenti.
"""

import os
import threading
import time
from contextlib import suppress
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtWidgets import (
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QScrollArea,
    QVBoxLayout,
    QWidget,
)

from src.core import config_manager
from src.core.preventivi_manager import MacroWorker
from src.gui.dialogs.confirmation_dialog import ConfirmationDialog
from src.gui.styles import COLORS
from src.gui.widgets.contabilita.consuntivo.log_widget import OperationLogWidget
from src.gui.widgets.contabilita.consuntivo.workflow_widgets import WorkflowMapWidget, WorkflowStepButton
from src.gui.widgets.core_widgets import (
    FilterComboBox,
    PrimaryButton,
    StandardInput,
    StandardTextEdit,
)
from src.gui.widgets.modern_card import ModernContentCard


class ModificaEsistenteTab(QWidget):
    """Tab intelligente per la gestione di consuntivi esistenti con caching delle scansioni di rete."""

    step_clicked = Signal(str)
    _scan_finished = Signal()
    _scan_error = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.macro_worker: MacroWorker | None = None
        self.loaded_file: str | None = None
        self._last_scan_time = 0.0
        self._cached_files: list[tuple[str, str]] = []
        self._is_scanning = False
        self._scan_finished.connect(self._update_combo_from_cache)
        self._scan_error.connect(self._on_scan_error)
        self._setup_ui()
        QTimer.singleShot(500, self._scan_directory)

    def _on_scan_error(self, msg: str) -> None:
        self.log_widget.append_log(msg, "error")

    def _setup_ui(self) -> None:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 12, 20, 16)
        layout.setSpacing(14)

        self._setup_file_selection_card(layout)
        self._setup_data_extraction_card(layout)
        self._setup_footer(layout)

        layout.addStretch()
        scroll.setWidget(content)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _setup_file_selection_card(self, layout: QVBoxLayout) -> None:
        """Card: SELEZIONE FILE."""
        file_card = ModernContentCard()
        file_lay = file_card.content_layout
        file_lay.setContentsMargins(20, 15, 20, 20)
        file_lay.setSpacing(12)

        file_title = QLabel("SELEZIONE CONSUNTIVO")
        file_title.setStyleSheet(
            f"font-weight: 800; font-size: 13px; letter-spacing: 0.5px; "
            f"color: {COLORS['primary_dark']}; border: none;"
        )
        file_lay.addWidget(file_title)

        top_row = QHBoxLayout()
        top_row.setSpacing(15)
        anno_lbl = QLabel("ANNO")
        anno_lbl.setStyleSheet(
            f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;"
        )
        top_row.addWidget(anno_lbl)
        self.anno_combo = FilterComboBox()
        self.anno_combo.addItems([str(y) for y in range(datetime.now(UTC).year, 2024, -1)])
        self.anno_combo.setFixedWidth(100)
        self.anno_combo.currentIndexChanged.connect(lambda: self._scan_directory(force=True))
        top_row.addWidget(self.anno_combo)
        self._dir_label = QLabel("")
        self._dir_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px; border: none;")
        top_row.addWidget(self._dir_label, 1)
        top_row.addStretch()
        file_lay.addLayout(top_row)

        file_sel_row = QHBoxLayout()
        file_sel_row.setSpacing(10)
        file_lbl = QLabel("FILE")
        file_lbl.setStyleSheet(
            f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;"
        )
        file_sel_row.addWidget(file_lbl)
        self.file_combo = FilterComboBox()
        self.file_combo.setMinimumHeight(38)
        self.file_combo.currentIndexChanged.connect(self._on_file_selected)
        file_sel_row.addWidget(self.file_combo, 1)
        self._file_count_label = QLabel("Ricerca in corso...")
        self._file_count_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px; border: none;")
        file_sel_row.addWidget(self._file_count_label)
        file_lay.addLayout(file_sel_row)
        layout.addWidget(file_card)

    def _setup_data_extraction_card(self, layout: QVBoxLayout) -> None:
        """Card: DATI ESTRATTI."""
        data_card = ModernContentCard()
        data_lay = data_card.content_layout
        data_lay.setContentsMargins(20, 15, 20, 20)
        data_lay.setSpacing(12)
        data_title = QLabel("DATI ESTRATTI DAL FILE")
        data_title.setStyleSheet(
            f"font-weight: 800; font-size: 13px; letter-spacing: 0.5px; "
            f"color: {COLORS['primary_dark']}; border: none;"
        )
        data_lay.addWidget(data_title)
        self._status_label = QLabel("Seleziona un file per visualizzare i dati.")
        self._status_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px; border: none;")
        data_lay.addWidget(self._status_label)

        grid = QGridLayout()
        grid.setSpacing(10)
        self._fields: dict[str, StandardInput] = {}
        field_defs = [
            ("Data (A5)", "data", 0, 0),
            ("TCL (A7)", "tcl", 0, 1),
            ("ODC (B5)", "odc", 0, 2),
            ("Avviso (C7)", "avviso", 0, 3),
            ("Ordine (C5)", "ordine", 1, 0),
            ("Stato (D11)", "stato", 1, 1),
            ("Tipo Prev. (D13)", "tipo_prev", 1, 2),
            ("Tipo Econ. (E13)", "tipo_econ", 1, 3),
            ("Progressivo", "progressivo", 2, 0),
        ]
        for label, key, row, col in field_defs:
            self._add_field_to_grid(grid, label, key, row, col)
        data_lay.addLayout(grid)

        desc_lbl = QLabel("Descrizione Lavoro (A11:A21)")
        desc_lbl.setStyleSheet(
            f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;"
        )
        data_lay.addWidget(desc_lbl)
        self._desc_lavoro_display = StandardTextEdit()
        self._desc_lavoro_display.setMaximumHeight(80)
        data_lay.addWidget(self._desc_lavoro_display)

        self._save_btn = PrimaryButton("Salva Modifiche nel File Excel")
        self._save_btn.setMinimumHeight(42)
        self._save_btn.clicked.connect(self._save_to_file)
        data_lay.addWidget(self._save_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(data_card)

    def _add_field_to_grid(self, grid: QGridLayout, label: str, key: str, row: int, col: int) -> None:
        """Helper per aggiungere un campo alla griglia dati."""
        lbl = QLabel(label)
        lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
        inp = StandardInput()
        inp.setMinimumHeight(32)
        self._fields[key] = inp
        v = QVBoxLayout()
        v.setSpacing(2)
        v.addWidget(lbl)
        v.addWidget(inp)
        grid.addLayout(v, row, col)

    def _setup_footer(self, layout: QVBoxLayout) -> None:
        """Sezione workflow e log."""
        self.workflow_map = WorkflowMapWidget()
        self.workflow_map.step_clicked.connect(self._on_workflow_step)
        layout.addWidget(self.workflow_map)

        self.log_widget = OperationLogWidget()
        self.log_widget.setMinimumHeight(160)
        layout.addWidget(self.log_widget)


    def _get_dynamic_dir(self) -> str:
        year = self.anno_combo.currentText()
        base = config_manager.get_config_value(
            "base_network_path_preventivi", r"\\192.168.11.251\Database_Tecnico_SMI\Contabilità strumentale"
        )
        return os.path.join(base, year, "CONSUNTIVI", year)

    def _scan_directory(self, force: bool = False) -> None:
        """Esegue la scansione dei file con caching per evitare lag su rete."""
        now = time.time()
        directory = self._get_dynamic_dir()
        self._dir_label.setText(directory)

        # Se meno di 30 secondi dall'ultima scansione e non forzato, usa la cache
        if not force and (now - self._last_scan_time < 30) and self._cached_files:
            self._update_combo_from_cache()
            return

        if self._is_scanning:
            return

        self._is_scanning = True
        self._last_scan_time = now

        def run_scan() -> None:
            temp_files = []
            try:
                dir_path = Path(directory)
                if dir_path.is_dir():
                    files = sorted(
                        [f for f in os.listdir(directory) if f.lower().endswith(".xlsm")],
                        reverse=True,
                    )
                    for f in files:
                        full = dir_path / f
                        try:
                            size_kb = full.stat().st_size / 1024
                            temp_files.append((f"{f} ({size_kb:.0f} KB)", str(full)))
                        except OSError:
                            temp_files.append((f, str(full)))

                self._cached_files = temp_files
                self._scan_finished.emit()
            except Exception as e:
                err_msg = str(e)
                self._scan_error.emit(f"Errore scansione: {err_msg}")
            finally:
                self._is_scanning = False

        threading.Thread(target=run_scan, daemon=True).start()

    def _update_combo_from_cache(self) -> None:
        self.file_combo.blockSignals(True)
        current_text = self.file_combo.currentText()
        self.file_combo.clear()

        for label, data in self._cached_files:
            self.file_combo.addItem(label, data)

        # Prova a ripristinare la selezione precedente
        idx = self.file_combo.findText(current_text)
        if idx >= 0:
            self.file_combo.setCurrentIndex(idx)
        elif self.file_combo.count() > 0:
            self.file_combo.setCurrentIndex(0)

        count = len(self._cached_files)
        self._file_count_label.setText(f"{count} file trovati")
        color = "#2E7D32" if count > 0 else COLORS["text_muted"]
        self._file_count_label.setStyleSheet(
            f"color: {color}; font-size: 11px; font-weight: 600; border: none;"
        )
        self.file_combo.blockSignals(False)

    def _on_file_selected(self, index: int) -> None:
        if index < 0 or self.file_combo.count() == 0:
            return
        file_path = self.file_combo.itemData(index)
        if not file_path or not Path(file_path).exists():
            self._status_label.setText("File non trovato.")
            return
        self.loaded_file = file_path
        self.log_widget.append_log(f"Lettura: {os.path.basename(file_path)}", "step")
        self._auto_fill_from_file(file_path)

    def _auto_fill_from_file(self, file_path: str) -> None:
        """Estrae i dati dal file Excel e popola i campi."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = self._find_data_sheet(wb)
            if sheet is None:
                self._status_label.setText("Nessun foglio dati trovato.")
                wb.close()
                return

            self._fill_main_fields(sheet)
            self._fill_vba_progressivo(wb)
            self._fill_description_lines(sheet)

            wb.close()
            n = sum(1 for f in self._fields.values() if f.text())
            self._update_status_success(n)
        except ImportError:
            self._status_label.setText("openpyxl non disponibile")
            self.log_widget.append_log("openpyxl non installato", "warning")
        except Exception as e:
            self._status_label.setText(f"Errore lettura: {e}")
            self.log_widget.append_log(f"Errore: {e}", "error")

    def _find_data_sheet(self, wb: object) -> Any:
        """Cerca il foglio di inserimento dati nel workbook."""
        # @obj: openpyxl.workbook.Workbook
        sheet_names = getattr(wb, "sheetnames", [])
        for name in ("inserimento dati", "Inserimento Dati", "inserimento_dati"):
            if name in sheet_names:
                return wb[name]
        return wb[sheet_names[0]] if sheet_names else None

    def _get_cell_value(self, sheet: Any, addr: str) -> str:
        """Legge e formatta il valore di una cella."""
        try:
            v = sheet[addr].value
            if v is None:
                return ""
            if hasattr(v, "strftime"):
                return cast("str", v.strftime("%d/%m/%Y"))
            return str(v).strip()
        except Exception:
            return ""

    def _fill_main_fields(self, sheet: Any) -> None:
        """Popola i campi principali dalla mappa celle."""
        cell_map = {
            "data": "A5",
            "tcl": "A7",
            "odc": "B5",
            "avviso": "C7",
            "ordine": "C5",
            "stato": "D11",
            "tipo_prev": "D13",
            "tipo_econ": "E13",
        }
        for key, addr in cell_map.items():
            self._fields[key].setText(self._get_cell_value(sheet, addr))

    def _fill_vba_progressivo(self, wb: Any) -> None:
        """Tenta di leggere il progressivo dal foglio rif.VBA."""
        prog = ""
        if "rif.VBA" in wb.sheetnames:
            with suppress(Exception):
                prog_val = wb["rif.VBA"]["A4"].value
                if prog_val:
                    prog = str(prog_val).strip()
        self._fields["progressivo"].setText(prog)

    def _fill_description_lines(self, sheet: Any) -> None:
        """Legge le righe della descrizione lavoro (A11:A21)."""
        lines = [self._get_cell_value(sheet, f"A{r}") for r in range(11, 22) if self._get_cell_value(sheet, f"A{r}")]
        self._desc_lavoro_display.setPlainText("\n".join(lines))

    def _update_status_success(self, count: int) -> None:
        """Aggiorna la UI dopo un auto-fill riuscito."""
        self._status_label.setText(f"Campi compilati automaticamente: {count}")
        self._status_label.setStyleSheet("color: #2E7D32; font-size: 12px; font-weight: 600; border: none;")
        self.log_widget.append_log(f"Auto-fill completato: {count} campi", "success")

    def _save_to_file(self) -> None:
        """Salva i valori dei campiùeditati nel file Excel."""
        if not self.loaded_file or not Path(self.loaded_file).exists():
            ConfirmationDialog.show_error(self, "Errore", "Nessun file selezionato.")
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.loaded_file, keep_vba=True)
            sheet = self._find_data_sheet(wb)
            if sheet is None:
                ConfirmationDialog.show_error(self, "Errore", "Foglio non trovato.")
                wb.close()
                return

            self._update_excel_cells(sheet)
            self._update_excel_descriptions(sheet)

            wb.save(self.loaded_file)
            wb.close()
            self._on_save_success()
        except ImportError:
            ConfirmationDialog.show_error(self, "Errore", "openpyxl non disponibile.")
        except Exception as e:
            self.log_widget.append_log(f"Errore salvataggio: {e}", "error")
            ConfirmationDialog.show_error(self, "Errore Salvataggio", str(e))

    def _update_excel_cells(self, sheet: Any) -> None:
        """Aggiorna le celle principali nel foglio Excel."""
        cell_map = {
            "data": "A5",
            "tcl": "A7",
            "odc": "B5",
            "avviso": "C7",
            "ordine": "C5",
            "stato": "D11",
            "tipo_prev": "D13",
            "tipo_econ": "E13",
        }
        for key, addr in cell_map.items():
            val = self._fields[key].text().strip()
            if val:
                sheet[addr] = val

    def _update_excel_descriptions(self, sheet: Any) -> None:
        """Aggiorna la descrizione lavoro (A11:A21)."""
        desc_lines = self._desc_lavoro_display.toPlainText().split("\n")
        for i in range(11, 22):
            idx = i - 11
            sheet[f"A{i}"] = desc_lines[idx] if idx < len(desc_lines) else ""

    def _on_save_success(self) -> None:
        """Gestisce il completamento del salvataggio."""
        self.log_widget.append_log("Modifiche salvate con successo!", "success")
        self._status_label.setText("Modifiche salvate nel file Excel")
        self._status_label.setStyleSheet("color: #2E7D32; font-size: 12px; font-weight: 600; border: none;")
        ConfirmationDialog.show_info(
            self, "Salvato", f"Le modifiche sono state salvate in:\n{os.path.basename(self.loaded_file)}"
        )
        self._last_scan_time = 0  # Forza refresh scansione

    def _on_workflow_step(self, step_id: str) -> None:
        macros = self.workflow_map.get_macros_for_step(step_id)
        if not macros:
            return
        if not self.loaded_file or not Path(self.loaded_file).exists():
            ConfirmationDialog.show_error(self, "Errore", "Nessun file selezionato.")
            return
        self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ACTIVE)
        self.log_widget.append_log(
            f"Esecuzione: {', '.join(macros)} su {os.path.basename(self.loaded_file)}",
            "step",
        )
        self.setEnabled(False)
        self.macro_worker = MacroWorker(self.loaded_file, macros)
        self.macro_worker.finished_signal.connect(lambda ok, msg: self._on_macro_finished(ok, msg, step_id))
        self.macro_worker.start()

    def _on_macro_finished(self, success: bool, result: str, step_id: str) -> None:
        self.setEnabled(True)
        if success:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.COMPLETED)
            self.log_widget.append_log(f"✅ {result}", "success")
            ConfirmationDialog.show_info(self, "Macro Eseguite", result)
        else:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ERROR)
            self.log_widget.append_log(f"❌ {result}", "error")
            ConfirmationDialog.show_error(self, "Errore Macro", result)
