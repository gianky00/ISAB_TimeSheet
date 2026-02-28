"""
SyncroJob - Consuntivo Panel
Pannello premium per la generazione e manipolazione dei consuntivi automatizzati.
Integra la logica migrata dal GeneratorePreventiviTab con:
- Generazione file Excel da Master template via PreventiviGeneratorManager
- Esecuzione macro VBA via COM/win32com (MacroWorker)
- Mappa workflow interattiva con 7 step
- Due tab: Crea Nuovo e Modifica Esistente
"""

import logging
import os
from datetime import datetime
from typing import Any

from PyQt6.QtCore import (
    QEasingCurve,
    QPropertyAnimation,
    Qt,
    QThread,
    QTimer,
    pyqtProperty,
    pyqtSignal,
)
from PyQt6.QtGui import (
    QColor,
    QFont,
)
from PyQt6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from src.core import config_manager
from src.core.preventivi_manager import PreventiviGeneratorManager
from src.gui.dialogs.confirmation_dialog import ConfirmationDialog
from src.gui.styles import COLORS
from src.gui.widgets.core_widgets import PrimaryButton, StandardInput
from src.gui.widgets.modern_card import ModernContentCard
from src.gui.components.animated_tab_widget import AnimatedTabWidget

logger = logging.getLogger(__name__)


# =============================================================================
# WORKERS (migrati da generatore_preventivi_tab.py)
# =============================================================================
class GeneratoreWorker(QThread):
    """Esegue la generazione del file Excel in background."""

    finished_signal = pyqtSignal(bool, str)

    def __init__(self, master_path: str, data: dict, dest_path: str) -> None:
        super().__init__()
        self.master_path = master_path
        self.data = data
        self.dest_path = dest_path

    def run(self) -> None:
        try:
            manager = PreventiviGeneratorManager(self.master_path)
            success, result = manager.generate_preventivo(self.data, self.dest_path)
            self.finished_signal.emit(success, result)
        except Exception as e:
            self.finished_signal.emit(False, f"Errore critico thread: {e}")


class MacroWorker(QThread):
    """Esegue una o più Macro VBA sul file generato in un thread separato."""

    finished_signal = pyqtSignal(bool, str)

    def __init__(self, file_path: str, macros: list[str]) -> None:
        super().__init__()
        self.file_path = file_path
        self.macros = macros

    def run(self) -> None:
        try:
            import pythoncom

            pythoncom.CoInitialize()
            import win32com.client

            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = True

            wb = excel_app.Workbooks.Open(self.file_path, UpdateLinks=0)

            for macro in self.macros:
                excel_app.Run(f"'{wb.Name}'!{macro}")

            wb.Save()
            self.finished_signal.emit(True, "Macro completate con successo.")
        except Exception as e:
            self.finished_signal.emit(False, f"Errore durante l'esecuzione della macro:\n{e}")
        finally:
            try:
                import pythoncom

                pythoncom.CoUninitialize()
            except Exception:
                pass


# =============================================================================
# WORKFLOW STEP BUTTON
# =============================================================================
class WorkflowStepButton(QFrame):
    """Card pulsante premium per uno step del workflow consuntivo."""

    clicked = pyqtSignal(str)

    class State:
        IDLE = "idle"
        ACTIVE = "active"
        COMPLETED = "completed"
        ERROR = "error"

    def __init__(
        self,
        step_id: str,
        step_number: int,
        title: str,
        description: str,
        is_action: bool = False,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._step_id = step_id
        self._step_number = step_number
        self._state = self.State.IDLE
        self._glow_opacity = 0.0
        self._is_action = is_action

        if is_action:
            self.setFixedSize(200, 110)
        else:
            self.setFixedSize(165, 130)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setToolTip(f"Step {step_number}: {description}")

        self._setup_ui(step_number, title, description)
        self._setup_glow_animation()
        self._apply_style()

    def _setup_ui(self, number: int, title: str, description: str) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(4)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        top_row = QHBoxLayout()
        top_row.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._number_badge = QLabel(str(number))
        self._number_badge.setFixedSize(28, 28)
        self._number_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._number_badge.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        top_row.addWidget(self._number_badge)
        layout.addLayout(top_row)

        self._title_label = QLabel(title)
        self._title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._title_label.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        self._title_label.setWordWrap(True)
        layout.addWidget(self._title_label)

        desc_label = QLabel(description)
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_label.setFont(QFont("Segoe UI", 9))
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet(f"color: {COLORS['text_muted']}; background: transparent;")
        layout.addWidget(desc_label)

    def _setup_glow_animation(self) -> None:
        self._glow_anim = QPropertyAnimation(self, b"glowOpacity")
        self._glow_anim.setDuration(1200)
        self._glow_anim.setEasingCurve(QEasingCurve.Type.InOutSine)
        self._glow_anim.setStartValue(0.3)
        self._glow_anim.setEndValue(1.0)
        self._glow_anim.setLoopCount(-1)

    def get_glow_opacity(self) -> float:
        return self._glow_opacity

    def set_glow_opacity(self, value: float) -> None:
        self._glow_opacity = value
        self.update()

    glowOpacity = pyqtProperty(float, fget=get_glow_opacity, fset=set_glow_opacity)

    def set_state(self, state: str) -> None:
        self._state = state
        self._apply_style()
        if state == self.State.ACTIVE:
            self._glow_anim.start()
        else:
            self._glow_anim.stop()
            self._glow_opacity = 0.0

    def _apply_style(self) -> None:
        state_styles = {
            self.State.IDLE: {
                "bg": "#ffffff", "border": COLORS["border_light"],
                "badge_bg": "#f1f3f5", "badge_color": COLORS["text_muted"],
                "title_color": COLORS["text_dark"],
                "shadow_color": QColor(0, 0, 0, 25),
            },
            self.State.ACTIVE: {
                "bg": "#e8f5e9", "border": "#4CAF50",
                "badge_bg": "#4CAF50", "badge_color": "#ffffff",
                "title_color": "#2E7D32",
                "shadow_color": QColor(76, 175, 80, 80),
            },
            self.State.COMPLETED: {
                "bg": "#f0fdf4", "border": "#2E7D32",
                "badge_bg": "#2E7D32", "badge_color": "#ffffff",
                "title_color": "#1b5e20",
                "shadow_color": QColor(46, 125, 50, 60),
            },
            self.State.ERROR: {
                "bg": "#fef2f2", "border": COLORS["error_red"],
                "badge_bg": COLORS["error_red"], "badge_color": "#ffffff",
                "title_color": COLORS["error_red"],
                "shadow_color": QColor(220, 53, 69, 60),
            },
        }
        s = state_styles.get(self._state, state_styles[self.State.IDLE])

        if self._is_action:
            self.setStyleSheet("""
                WorkflowStepButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #1a237e, stop:0.5 #283593, stop:1 #1565c0);
                    border: 2px solid #1a237e; border-radius: 14px;
                }
                WorkflowStepButton:hover {
                    border: 2px solid #42a5f5;
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #283593, stop:0.5 #1565c0, stop:1 #1976d2);
                }
            """)
            self._number_badge.setStyleSheet(
                "background: rgba(255,255,255,0.2); color: #ffffff; "
                "border-radius: 14px; border: 1px solid rgba(255,255,255,0.3);"
            )
            self._title_label.setStyleSheet("color: #ffffff; background: transparent;")
        else:
            self.setStyleSheet(f"""
                WorkflowStepButton {{
                    background-color: {s["bg"]}; border: 2px solid {s["border"]};
                    border-radius: 14px;
                }}
                WorkflowStepButton:hover {{
                    border: 2px solid #009688; background-color: #e0f2f1;
                }}
            """)
            self._number_badge.setStyleSheet(
                f"background-color: {s['badge_bg']}; color: {s['badge_color']}; "
                f"border-radius: 14px; border: none;"
            )
            self._title_label.setStyleSheet(
                f"color: {s['title_color']}; background: transparent;"
            )

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(16)
        shadow.setOffset(0, 4)
        shadow.setColor(s["shadow_color"])
        self.setGraphicsEffect(shadow)

    def mousePressEvent(self, event: Any) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self._step_id)
        super().mousePressEvent(event)

    def enterEvent(self, event: Any) -> None:
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(24)
        shadow.setOffset(0, 6)
        shadow.setColor(QColor(0, 150, 136, 100))
        self.setGraphicsEffect(shadow)
        super().enterEvent(event)

    def leaveEvent(self, event: Any) -> None:
        self._apply_style()
        super().leaveEvent(event)


# =============================================================================
# WORKFLOW MAP WIDGET
# =============================================================================
class WorkflowMapWidget(QWidget):
    """Widget mappa workflow con step connessi da frecce."""

    step_clicked = pyqtSignal(str)

    STEPS = [
        ("carica_dati", 1, "CARICA\nDATI", "CaricaDatiMultiplo"),
        ("elabora_dati", 2, "ELABORA\nDATI", "elaboraDati"),
        ("compila_consuntivo", 3, "COMPILA\nCONSUNTIVO", "EseguiTuttiSmista"),
        ("verifica", 4, "VERIFICA", "VerificaConsuntivo"),
        ("stampa", 5, "STAMPA", "verificaEstampaFogli"),
    ]

    ACTIONS = [
        ("esegui_1_4", 6, "🚀 ESEGUI\n1→4", "Carica → Elabora → Compila → Verifica"),
        ("esegui_1_5", 7, "🚀 ESEGUI\n1→5", "Intero workflow completo"),
    ]

    # Mapping step_id → lista macro VBA da eseguire
    MACRO_MAP: dict[str, list[str]] = {
        "carica_dati": ["CaricaDatiMultiplo"],
        "elabora_dati": ["elaboraDati"],
        "compila_consuntivo": ["EseguiTuttiSmista"],
        "verifica": ["VerificaConsuntivo"],
        "stampa": ["verificaEstampaFogli"],
        "esegui_1_4": ["CaricaDatiMultiplo", "elaboraDati", "EseguiTuttiSmista", "VerificaConsuntivo"],
        "esegui_1_5": ["CaricaDatiMultiplo", "elaboraDati", "EseguiTuttiSmista", "VerificaConsuntivo", "verificaEstampaFogli"],
    }

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setMinimumHeight(300)
        self._step_buttons: dict[str, WorkflowStepButton] = {}
        self._setup_ui()

    def _setup_ui(self) -> None:
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        container = QFrame()
        container.setObjectName("workflowContainer")
        container.setStyleSheet("""
            QFrame#workflowContainer {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fffe, stop:0.3 #e8f5e9, stop:0.7 #e0f7fa, stop:1 #f3e5f5);
                border: 1px solid rgba(0, 150, 136, 0.15);
                border-radius: 20px;
            }
        """)
        shadow = QGraphicsDropShadowEffect(container)
        shadow.setBlurRadius(30)
        shadow.setOffset(0, 8)
        shadow.setColor(QColor(0, 0, 0, 20))
        container.setGraphicsEffect(shadow)

        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(24, 20, 24, 20)
        container_layout.setSpacing(16)

        title = QLabel("⚡ WORKFLOW PIPELINE")
        title.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_dark']}; background: transparent; letter-spacing: 2px;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(title)

        # Riga principale: 5 step
        main_row = QHBoxLayout()
        main_row.setSpacing(0)
        main_row.setAlignment(Qt.AlignmentFlag.AlignCenter)

        for i, (sid, num, label, desc) in enumerate(self.STEPS):
            btn = WorkflowStepButton(sid, num, label, desc)
            btn.clicked.connect(self.step_clicked.emit)
            self._step_buttons[sid] = btn
            main_row.addWidget(btn)
            if i < len(self.STEPS) - 1:
                arrow = self._create_arrow_label()
                main_row.addWidget(arrow)

        container_layout.addLayout(main_row)

        # Riga azioni composte
        action_row = QHBoxLayout()
        action_row.setSpacing(16)
        action_row.setAlignment(Qt.AlignmentFlag.AlignCenter)

        for sid, num, label, desc in self.ACTIONS:
            btn = WorkflowStepButton(sid, num, label, desc, is_action=True)
            btn.setFixedSize(190, 100)
            btn.clicked.connect(self.step_clicked.emit)
            self._step_buttons[sid] = btn
            action_row.addWidget(btn)

        # Separatore verticale
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setFixedHeight(70)
        sep.setStyleSheet("background: rgba(0,0,0,0.12); max-width: 2px; min-width: 2px;")
        action_row.addWidget(sep)

        # Extra macro buttons con label complete
        extras = [
            ("email_gen", 8, "📧 Email\nGenerica", "InviaEmailGenerico"),
            ("email_chiamata", 9, "📞 Email\nChiamata", "InviaEmailConsuntivoChiamata"),
            ("relazione", 10, "📝 Relazione\nTecnica", "CreaEConvertiRelazioneTecnica"),
        ]
        for sid, num, label, desc in extras:
            btn = WorkflowStepButton(sid, num, label, desc, is_action=True)
            btn.setFixedSize(140, 80)
            btn.clicked.connect(self.step_clicked.emit)
            self._step_buttons[sid] = btn
            action_row.addWidget(btn)

        container_layout.addLayout(action_row)
        main_layout.addWidget(container)

    def _create_arrow_label(self, color: str = "#009688") -> QLabel:
        arrow = QLabel("→")
        arrow.setFont(QFont("Segoe UI", 22, QFont.Weight.Bold))
        arrow.setStyleSheet(f"color: {color}; background: transparent; padding: 0 6px;")
        arrow.setAlignment(Qt.AlignmentFlag.AlignCenter)
        arrow.setFixedWidth(40)
        return arrow

    def set_step_state(self, step_id: str, state: str) -> None:
        if btn := self._step_buttons.get(step_id):
            btn.set_state(state)

    def reset_all(self) -> None:
        for btn in self._step_buttons.values():
            btn.set_state(WorkflowStepButton.State.IDLE)

    def get_macros_for_step(self, step_id: str) -> list[str]:
        """Restituisce la lista di macro VBA per lo step indicato."""
        extra_map: dict[str, list[str]] = {
            "email_gen": ["InviaEmailGenerico"],
            "email_chiamata": ["InviaEmailConsuntivoChiamata"],
            "relazione": ["CreaEConvertiRelazioneTecnica"],
        }
        return self.MACRO_MAP.get(step_id, extra_map.get(step_id, []))


# =============================================================================
# LOG WIDGET - Console operazioni
# =============================================================================
class OperationLogWidget(QFrame):
    """Console dark per i log delle operazioni."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("logWidget")
        self.setStyleSheet("""
            QFrame#logWidget {
                background-color: #1e1e2e;
                border: 1px solid #313244;
                border-radius: 16px;
            }
        """)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(12)
        shadow.setOffset(0, 3)
        shadow.setColor(QColor(0, 0, 0, 20))
        self.setGraphicsEffect(shadow)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        header_row = QHBoxLayout()
        header_label = QLabel("🖥️ Console Operazioni")
        header_label.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        header_label.setStyleSheet("color: #cdd6f4;")
        header_row.addWidget(header_label)
        header_row.addStretch()

        clear_btn = QPushButton("Pulisci")
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255,255,255,0.08); color: #bac2de;
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 6px; padding: 4px 12px; font-size: 11px;
            }
            QPushButton:hover { background: rgba(255,255,255,0.15); }
        """)
        clear_btn.clicked.connect(self.clear)
        header_row.addWidget(clear_btn)
        layout.addLayout(header_row)

        self._log_text = QTextEdit()
        self._log_text.setReadOnly(True)
        self._log_text.setFont(QFont("Cascadia Code", 10))
        self._log_text.setStyleSheet("""
            QTextEdit {
                background: transparent; color: #a6e3a1; border: none;
                selection-background-color: rgba(137, 180, 250, 0.3);
            }
            QScrollBar:vertical { border: none; background: transparent; width: 6px; }
            QScrollBar::handle:vertical { background: rgba(255,255,255,0.15); border-radius: 3px; }
        """)
        layout.addWidget(self._log_text)

    def append_log(self, message: str, level: str = "info") -> None:
        colors = {
            "info": "#89b4fa", "success": "#a6e3a1", "warning": "#f9e2af",
            "error": "#f38ba8", "step": "#cba6f7",
        }
        color = colors.get(level, colors["info"])
        self._log_text.append(f'<span style="color:{color};">{message}</span>')

    def clear(self) -> None:
        self._log_text.clear()


# =============================================================================
# TAB CREA NUOVO (migrato da GeneratorePreventiviTab)
# =============================================================================
class CreaNuovoTab(QWidget):
    """Tab per la generazione di un nuovo consuntivo con tutti i campi del vecchio Generatore Preventivi."""

    step_clicked = pyqtSignal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.worker: GeneratoreWorker | None = None
        self.macro_worker: MacroWorker | None = None
        self.last_generated_file: str | None = None
        self._macro_buttons: list[QPushButton] = []
        self._setup_ui()

    def _setup_ui(self) -> None:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 12, 20, 16)
        layout.setSpacing(14)

        # --- CARD 1: SETUP E PERCORSI ---
        card1, card1_lay = self._create_card("IMPOSTAZIONI E DESTINAZIONE")
        config_row = QHBoxLayout()
        config_row.setSpacing(20)

        self.anno_combo = QComboBox()
        self.anno_combo.addItems([str(y) for y in range(datetime.now().year, 2024, -1)])
        self.anno_combo.currentIndexChanged.connect(self._update_dynamic_path)
        config_row.addLayout(self._create_input_group("ANNO", self.anno_combo, width=100))

        self.progressivo_edit = StandardInput()
        self.progressivo_edit.setStyleSheet(f"color: {COLORS['teal_accent']}; font-weight: bold;")
        config_row.addLayout(self._create_input_group("PROGRESSIVO", self.progressivo_edit, width=120))

        self.dest_path_edit = StandardInput()
        self.dest_path_edit.setReadOnly(True)
        self.dest_path_edit.setStyleSheet(f"background-color: {COLORS['bg_light']}; color: {COLORS['text_muted']};")
        config_row.addLayout(self._create_input_group("PERCORSO DI RETE (AUTOMATICO)", self.dest_path_edit))

        card1_lay.addLayout(config_row)
        layout.addWidget(card1)

        # --- CARD 2: DATI IDENTIFICATIVI ---
        card2, id_layout = self._create_card("DETTAGLI INTERVENTO E CLASSIFICAZIONE")

        row1 = QHBoxLayout()
        row1.setSpacing(15)
        self.data_edit = StandardInput()
        self.data_edit.setText(datetime.now().strftime("%d/%m/%Y"))
        row1.addLayout(self._create_input_group("DATA (A5)", self.data_edit, width=120))

        self.tcl_combo = QComboBox()
        self.tcl_combo.addItems([
            "MESSINA I.", "AGUSTA D.", "CALDARELLA F.",
            "PREZZAVENTO M.", "BOSCO F.", "RUGGIERI F.", "BARBAGALLO G.",
        ])
        row1.addLayout(self._create_input_group("TCL (A7)", self.tcl_combo, width=180))

        self.odc_edit = StandardInput()
        row1.addLayout(self._create_input_group("ODC (B5)", self.odc_edit, width=140))

        self.avviso_edit = StandardInput()
        row1.addLayout(self._create_input_group("AVVISO (C7)", self.avviso_edit, width=140))

        self.ordine_edit = StandardInput()
        row1.addLayout(self._create_input_group("ORDINE (C5)", self.ordine_edit, width=140))
        row1.addStretch()
        id_layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.setSpacing(15)

        self.stato_combo = QComboBox()
        self.stato_combo.addItems([
            "ATTIVITA' DA COMPLETARE", "IN ATTESA TCL",
            "RICHIESTA ODC MIDOLO", "CONTABILIZZATA",
        ])
        row2.addLayout(self._create_input_group("STATO ATTIVITÀ (D11)", self.stato_combo, width=220))

        self.tipo_prev_combo = QComboBox()
        self.tipo_prev_combo.addItems(["MISURA", "SQUADRA", "CHIAMATA", "FORNITURA", "PREVENTIVO"])
        row2.addLayout(self._create_input_group("TIPOLOGIA PREVENTIVO (D13)", self.tipo_prev_combo, width=220))

        self.tipo_econ_combo = QComboBox()
        self.tipo_econ_combo.addItems(["SQUADRA GIORNALIERA", "SQUADRA SETTIMANALE", "CONSTATAZIONE PURA"])
        row2.addLayout(self._create_input_group("TIPOLOGIA ECONOMIA (E13)", self.tipo_econ_combo, width=220))
        row2.addStretch()
        id_layout.addLayout(row2)
        layout.addWidget(card2)

        # --- CARD 3: DESCRIZIONI ---
        card3, desc_layout = self._create_card("DESCRIZIONE DELLE ATTIVITÀ")

        desc_row = QHBoxLayout()
        desc_row.setSpacing(20)

        self.desc_lavoro_edit = QTextEdit()
        self.desc_lavoro_edit.setPlaceholderText("Es. Smontaggio valvola...")
        self.desc_lavoro_edit.setMinimumHeight(80)
        self.desc_lavoro_edit.setMaximumHeight(110)
        self.desc_lavoro_edit.setStyleSheet(f"""
            QTextEdit {{
                border: 1px solid {COLORS['border_light']}; border-radius: 6px;
                padding: 10px; background-color: {COLORS['bg_white']};
                color: {COLORS['text_dark']}; font-size: 13px;
            }}
            QTextEdit:focus {{ border: 2px solid {COLORS['teal_accent']}; }}
        """)
        desc_row.addLayout(self._create_input_group("DESCRIZIONE LAVORO (A11:A21)", self.desc_lavoro_edit))

        self.desc_relazione_edit = QTextEdit()
        self.desc_relazione_edit.setPlaceholderText("Inserisci eventuali note (A32)...")
        self.desc_relazione_edit.setMinimumHeight(80)
        self.desc_relazione_edit.setMaximumHeight(110)
        self.desc_relazione_edit.setStyleSheet(self.desc_lavoro_edit.styleSheet())
        desc_row.addLayout(self._create_input_group("DESCRIZIONE RELAZIONE (A32)", self.desc_relazione_edit))

        desc_layout.addLayout(desc_row)
        layout.addWidget(card3)

        # --- MAPPA WORKFLOW ---
        self.workflow_map = WorkflowMapWidget()
        self.workflow_map.step_clicked.connect(self._on_workflow_step)
        layout.addWidget(self.workflow_map)

        # --- BOTTONE GENERA ---
        self.btn_generate = PrimaryButton("⚡ GENERA CONSUNTIVO EXCEL")
        self.btn_generate.setMinimumHeight(55)
        self.btn_generate.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['teal_accent']}; color: white;
                font-size: 16px; font-weight: bold; border-radius: 8px;
                letter-spacing: 1px;
            }}
            QPushButton:hover {{ background-color: #2b9e95; }}
            QPushButton:disabled {{
                background-color: {COLORS['border_light']};
                color: {COLORS['text_muted']};
            }}
        """)
        self.btn_generate.clicked.connect(self._on_generate)
        layout.addWidget(self.btn_generate)

        # --- LOG ---
        self.log_widget = OperationLogWidget()
        self.log_widget.setMinimumHeight(160)
        layout.addWidget(self.log_widget)

        layout.addStretch()
        scroll.setWidget(content)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

        # Init
        self._update_dynamic_path()

    # --- Helper UI ---
    def _create_card(self, title_text: str) -> tuple[ModernContentCard, QVBoxLayout]:
        card = ModernContentCard()
        lay = card.content_layout
        lay.setContentsMargins(20, 15, 20, 20)
        lay.setSpacing(15)
        title_lbl = QLabel(title_text)
        title_lbl.setStyleSheet(
            f"font-weight: 800; font-size: 13px; letter-spacing: 0.5px; "
            f"color: {COLORS['primary_dark']}; border: none;"
        )
        lay.addWidget(title_lbl)
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"background-color: {COLORS['bg_alt']}; border: none; min-height: 1px; max-height: 1px;")
        lay.addWidget(line)
        return card, lay

    def _create_input_group(self, label_text: str, widget: QWidget, width: int = 0) -> QVBoxLayout:
        lay = QVBoxLayout()
        lay.setSpacing(4)
        lbl = QLabel(label_text)
        lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
        lay.addWidget(lbl)
        if width > 0:
            widget.setFixedWidth(width)
        if isinstance(widget, StandardInput):
            widget.setMinimumHeight(36)
            widget.setMaximumHeight(36)
        lay.addWidget(widget)
        return lay

    # --- Logica percorso automatico ---
    def _update_dynamic_path(self) -> None:
        year = self.anno_combo.currentText()
        base_network = r"\\192.168.11.251\Database_Tecnico_SMI\Contabilita' strumentale"
        dynamic_path = os.path.join(base_network, year, "CONSUNTIVI", year)
        self.dest_path_edit.setText(dynamic_path)
        self.dest_path_edit.setToolTip(dynamic_path)

        try:
            manager = PreventiviGeneratorManager("")
            next_prog = manager.get_next_progressive(dynamic_path)
            self.progressivo_edit.setText(next_prog)
        except Exception:
            self.progressivo_edit.setText("001")

    # --- Generazione file ---
    def _on_generate(self) -> None:
        config = config_manager.load_config()
        master_path = config.get("master_preventivi_path", "")

        if not master_path or not os.path.exists(master_path):
            ConfirmationDialog.show_error(
                self, "Configurazione Errata",
                "Il file Master non è stato configurato nelle Impostazioni.",
            )
            return

        data = {
            "progressivo": self.progressivo_edit.text(),
            "anno_short": self.anno_combo.currentText()[-2:],
            "data": self.data_edit.text(),
            "tcl": self.tcl_combo.currentText(),
            "odc": self.odc_edit.text(),
            "avviso": self.avviso_edit.text(),
            "ordine": self.ordine_edit.text(),
            "stato_attivita": self.stato_combo.currentText(),
            "tipologia_preventivo": self.tipo_prev_combo.currentText(),
            "tipologia_economia": self.tipo_econ_combo.currentText(),
            "descrizione_lavoro": self.desc_lavoro_edit.toPlainText(),
            "descrizione_relazione": self.desc_relazione_edit.toPlainText(),
        }

        self.setEnabled(False)
        self.btn_generate.setText("GENERAZIONE IN CORSO...")
        self.log_widget.append_log("🔄 Generazione file in corso...", "step")

        self.worker = GeneratoreWorker(master_path, data, self.dest_path_edit.text())
        self.worker.finished_signal.connect(self._on_generate_finished)
        self.worker.start()

    def _on_generate_finished(self, success: bool, result: str) -> None:
        self.setEnabled(True)
        self.btn_generate.setText("⚡ GENERA CONSUNTIVO EXCEL")

        if success:
            self.last_generated_file = result
            self.log_widget.append_log(f"✅ File generato: {result}", "success")
            self.log_widget.append_log("ℹ️ I pulsanti della mappa workflow sono ora attivi.", "info")
            # I pulsanti della mappa workflow ora possono eseguire macro
            ConfirmationDialog.show_info(
                self, "File Generato",
                f"Il file Excel è pronto:\n\n{result}\n\nPuoi ora lanciare le Macro dalla mappa workflow.",
            )
        else:
            self.log_widget.append_log(f"❌ Errore: {result}", "error")
            ConfirmationDialog.show_error(self, "Errore Generazione", result)

    # --- Macro VBA workflow ---
    def _on_workflow_step(self, step_id: str) -> None:
        """Gestisce il click su uno step: esegue la macro VBA corrispondente."""
        macros = self.workflow_map.get_macros_for_step(step_id)
        if not macros:
            return

        if not self.last_generated_file or not os.path.exists(self.last_generated_file):
            ConfirmationDialog.show_error(
                self, "Errore",
                "Nessun file generato. Genera prima il consuntivo Excel.",
            )
            return

        self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ACTIVE)
        self.log_widget.append_log(
            f"🔄 Esecuzione: {', '.join(macros)} su {os.path.basename(self.last_generated_file)}",
            "step",
        )

        self.setEnabled(False)
        self.macro_worker = MacroWorker(self.last_generated_file, macros)
        self.macro_worker.finished_signal.connect(
            lambda ok, msg: self._on_macro_finished(ok, msg, step_id)
        )
        self.macro_worker.start()

    def _on_macro_finished(self, success: bool, result: str, step_id: str) -> None:
        self.setEnabled(True)
        if success:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.COMPLETED)
            self.log_widget.append_log(f"✅ {result}", "success")
            ConfirmationDialog.show_info(self, "Macro Eseguite", result)
        else:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ERROR)
            self.log_widget.append_log(f"❌ {result}", "error")
            ConfirmationDialog.show_error(self, "Errore Macro", result)


# =============================================================================
# TAB MODIFICA ESISTENTE (con auto-scan directory + auto-fill)
# =============================================================================
class ModificaEsistenteTab(QWidget):
    """Tab intelligente: scansiona la directory preventivi, elenca i file .xlsm,
    e auto-compila i campi leggendo il file selezionato senza aprire Excel."""

    step_clicked = pyqtSignal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.macro_worker: MacroWorker | None = None
        self.loaded_file: str | None = None
        self._setup_ui()
        QTimer.singleShot(500, self._scan_directory)

    def _setup_ui(self) -> None:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 12, 20, 16)
        layout.setSpacing(14)

        # --- CARD: SELEZIONE FILE ---
        file_card = ModernContentCard()
        file_lay = file_card.content_layout
        file_lay.setContentsMargins(20, 15, 20, 20)
        file_lay.setSpacing(12)

        file_title = QLabel("📂 SELEZIONE CONSUNTIVO")
        file_title.setStyleSheet(
            f"font-weight: 800; font-size: 13px; letter-spacing: 0.5px; "
            f"color: {COLORS['primary_dark']}; border: none;"
        )
        file_lay.addWidget(file_title)

        top_row = QHBoxLayout()
        top_row.setSpacing(15)
        anno_lbl = QLabel("ANNO")
        anno_lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
        top_row.addWidget(anno_lbl)
        self.anno_combo = QComboBox()
        self.anno_combo.addItems([str(y) for y in range(datetime.now().year, 2024, -1)])
        self.anno_combo.setFixedWidth(100)
        self.anno_combo.currentIndexChanged.connect(self._scan_directory)
        top_row.addWidget(self.anno_combo)
        self._dir_label = QLabel("")
        self._dir_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px; border: none;")
        top_row.addWidget(self._dir_label, 1)
        scan_btn = QPushButton("🔄 Aggiorna")
        scan_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        scan_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['teal_accent']}; color: white; border: none;
                border-radius: 6px; padding: 6px 16px; font-weight: 600; font-size: 12px;
            }}
            QPushButton:hover {{ background: #2b9e95; }}
        """)
        scan_btn.clicked.connect(self._scan_directory)
        top_row.addWidget(scan_btn)
        file_lay.addLayout(top_row)

        file_sel_row = QHBoxLayout()
        file_sel_row.setSpacing(10)
        file_lbl = QLabel("FILE")
        file_lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
        file_sel_row.addWidget(file_lbl)
        self.file_combo = QComboBox()
        self.file_combo.setMinimumHeight(38)
        self.file_combo.setStyleSheet(f"""
            QComboBox {{
                border: 2px solid {COLORS['border_light']}; border-radius: 8px;
                padding: 6px 12px; background: {COLORS['bg_white']};
                color: {COLORS['text_dark']}; font-size: 13px; font-weight: 600;
            }}
            QComboBox:focus {{ border: 2px solid {COLORS['teal_accent']}; }}
            QComboBox::drop-down {{ width: 30px; border-left: 1px solid {COLORS['border_light']}; }}
        """)
        self.file_combo.currentIndexChanged.connect(self._on_file_selected)
        file_sel_row.addWidget(self.file_combo, 1)
        self._file_count_label = QLabel("0 file trovati")
        self._file_count_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px; border: none;")
        file_sel_row.addWidget(self._file_count_label)
        file_lay.addLayout(file_sel_row)
        layout.addWidget(file_card)

        # --- CARD: DATI ESTRATTI ---
        data_card = ModernContentCard()
        data_lay = data_card.content_layout
        data_lay.setContentsMargins(20, 15, 20, 20)
        data_lay.setSpacing(12)
        data_title = QLabel("📋 DATI ESTRATTI DAL FILE")
        data_title.setStyleSheet(
            f"font-weight: 800; font-size: 13px; letter-spacing: 0.5px; "
            f"color: {COLORS['primary_dark']}; border: none;"
        )
        data_lay.addWidget(data_title)
        self._status_label = QLabel("Seleziona un file per visualizzare i dati.")
        self._status_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px; border: none;")
        data_lay.addWidget(self._status_label)

        ro_style = (
            f"background-color: #f0f9f8; color: {COLORS['text_dark']}; "
            f"border: 1px solid {COLORS['border_light']}; border-radius: 6px; "
            f"padding: 6px 10px; font-size: 13px;"
        )
        grid = QGridLayout()
        grid.setSpacing(10)
        self._fields: dict[str, QLineEdit] = {}
        field_defs = [
            ("Data (A5)", "data", 0, 0), ("TCL (A7)", "tcl", 0, 1),
            ("ODC (B5)", "odc", 0, 2), ("Avviso (C7)", "avviso", 0, 3),
            ("Ordine (C5)", "ordine", 1, 0), ("Stato (D11)", "stato", 1, 1),
            ("Tipo Prev. (D13)", "tipo_prev", 1, 2), ("Tipo Econ. (E13)", "tipo_econ", 1, 3),
            ("Progressivo", "progressivo", 2, 0),
        ]
        for label, key, row, col in field_defs:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
            inp = QLineEdit()
            inp.setStyleSheet(
                f"background-color: {COLORS['bg_white']}; color: {COLORS['text_dark']}; "
                f"border: 1px solid {COLORS['border_light']}; border-radius: 6px; "
                f"padding: 6px 10px; font-size: 13px;"
            )
            inp.setMinimumHeight(32)
            self._fields[key] = inp
            v = QVBoxLayout()
            v.setSpacing(2)
            v.addWidget(lbl)
            v.addWidget(inp)
            grid.addLayout(v, row, col)
        data_lay.addLayout(grid)

        desc_lbl = QLabel("Descrizione Lavoro (A11:A21)")
        desc_lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {COLORS['text_muted']}; border: none;")
        data_lay.addWidget(desc_lbl)
        self._desc_lavoro_display = QTextEdit()
        self._desc_lavoro_display.setMaximumHeight(80)
        self._desc_lavoro_display.setStyleSheet(
            f"background-color: #f0f9f8; color: {COLORS['text_dark']}; "
            f"border: 1px solid {COLORS['border_light']}; border-radius: 6px; "
            f"padding: 6px; font-size: 12px;"
        )
        data_lay.addWidget(self._desc_lavoro_display)

        # Bottone salva modifiche
        self._save_btn = QPushButton("💾 Salva Modifiche nel File Excel")
        self._save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._save_btn.setMinimumHeight(42)
        self._save_btn.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        self._save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['teal_accent']}; color: white;
                border: none; border-radius: 8px; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: #2b9e95; }}
            QPushButton:disabled {{ background-color: {COLORS['border_light']}; color: {COLORS['text_muted']}; }}
        """)
        self._save_btn.clicked.connect(self._save_to_file)
        data_lay.addWidget(self._save_btn)

        layout.addWidget(data_card)

        # Workflow
        self.workflow_map = WorkflowMapWidget()
        self.workflow_map.step_clicked.connect(self._on_workflow_step)
        layout.addWidget(self.workflow_map)

        # Log
        self.log_widget = OperationLogWidget()
        self.log_widget.setMinimumHeight(160)
        layout.addWidget(self.log_widget)

        layout.addStretch()
        scroll.setWidget(content)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _get_dynamic_dir(self) -> str:
        year = self.anno_combo.currentText()
        base = r"\\192.168.11.251\Database_Tecnico_SMI\Contabilita' strumentale"
        return os.path.join(base, year, "CONSUNTIVI", year)

    def _scan_directory(self) -> None:
        directory = self._get_dynamic_dir()
        self._dir_label.setText(directory)
        self._dir_label.setToolTip(directory)
        self.file_combo.blockSignals(True)
        self.file_combo.clear()
        try:
            if os.path.isdir(directory):
                files = sorted(
                    [f for f in os.listdir(directory) if f.lower().endswith(".xlsm")],
                    reverse=True,
                )
                for f in files:
                    full = os.path.join(directory, f)
                    try:
                        size_kb = os.path.getsize(full) / 1024
                        self.file_combo.addItem(f"{f}  ({size_kb:.0f} KB)", full)
                    except OSError:
                        self.file_combo.addItem(f, full)
                count = len(files)
                self._file_count_label.setText(f"{count} file trovati")
                color = "#2E7D32" if count > 0 else COLORS["text_muted"]
                self._file_count_label.setStyleSheet(
                    f"color: {color}; font-size: 11px; font-weight: 600; border: none;"
                )
                self.log_widget.append_log(f"📂 Scansione: {count} file in {directory}", "info")
            else:
                self._file_count_label.setText("Directory non raggiungibile")
                self._file_count_label.setStyleSheet(
                    f"color: {COLORS['error_red']}; font-size: 11px; border: none;"
                )
                self.log_widget.append_log(f"⚠️ Directory non trovata: {directory}", "warning")
        except Exception as e:
            self._file_count_label.setText("Errore scansione")
            self.log_widget.append_log(f"❌ Errore scansione: {e}", "error")
        self.file_combo.blockSignals(False)
        if self.file_combo.count() > 0:
            self._on_file_selected(0)

    def _on_file_selected(self, index: int) -> None:
        if index < 0 or self.file_combo.count() == 0:
            return
        file_path = self.file_combo.itemData(index)
        if not file_path or not os.path.exists(file_path):
            self._status_label.setText("⚠️ File non trovato.")
            return
        self.loaded_file = file_path
        self.log_widget.append_log(f"📄 Lettura: {os.path.basename(file_path)}", "step")
        self._auto_fill_from_file(file_path)

    def _auto_fill_from_file(self, file_path: str) -> None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = None
            for name in ["inserimento dati", "Inserimento Dati", "inserimento_dati"]:
                if name in wb.sheetnames:
                    sheet = wb[name]
                    break
            if sheet is None and wb.sheetnames:
                sheet = wb[wb.sheetnames[0]]
            if sheet is None:
                self._status_label.setText("⚠️ Nessun foglio trovato.")
                wb.close()
                return

            def cv(addr: str) -> str:
                try:
                    v = sheet[addr].value
                    if v is None:
                        return ""
                    # Converte datetime in formato dd/mm/yyyy
                    if hasattr(v, 'strftime'):
                        return v.strftime("%d/%m/%Y")
                    return str(v).strip()
                except Exception:
                    return ""

            self._fields["data"].setText(cv("A5"))
            self._fields["tcl"].setText(cv("A7"))
            self._fields["odc"].setText(cv("B5"))
            self._fields["avviso"].setText(cv("C7"))
            self._fields["ordine"].setText(cv("C5"))
            self._fields["stato"].setText(cv("D11"))
            self._fields["tipo_prev"].setText(cv("D13"))
            self._fields["tipo_econ"].setText(cv("E13"))

            prog = ""
            if "rif.VBA" in wb.sheetnames:
                try:
                    prog_val = wb["rif.VBA"]["A4"].value
                    if prog_val:
                        prog = str(prog_val).strip()
                except Exception:
                    pass
            self._fields["progressivo"].setText(prog)

            lines = [cv(f"A{r}") for r in range(11, 22) if cv(f"A{r}")]
            self._desc_lavoro_display.setPlainText("\n".join(lines))
            wb.close()

            n = sum(1 for f in self._fields.values() if f.text())
            self._status_label.setText(f"✅ {n} campi compilati automaticamente")
            self._status_label.setStyleSheet("color: #2E7D32; font-size: 12px; font-weight: 600; border: none;")
            self.log_widget.append_log(f"✅ Auto-fill completato: {n} campi", "success")
        except ImportError:
            self._status_label.setText("⚠️ openpyxl non disponibile")
            self.log_widget.append_log("⚠️ openpyxl non installato", "warning")
        except Exception as e:
            self._status_label.setText(f"❌ Errore lettura: {e}")
            self.log_widget.append_log(f"❌ Errore: {e}", "error")

    def _save_to_file(self) -> None:
        """Salva i valori dei campi editati nel file Excel."""
        if not self.loaded_file or not os.path.exists(self.loaded_file):
            ConfirmationDialog.show_error(self, "Errore", "Nessun file selezionato.")
            return

        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.loaded_file, keep_vba=True)

            # Foglio 'inserimento dati'
            sheet = None
            for name in ["inserimento dati", "Inserimento Dati", "inserimento_dati"]:
                if name in wb.sheetnames:
                    sheet = wb[name]
                    break
            if sheet is None and wb.sheetnames:
                sheet = wb[wb.sheetnames[0]]
            if sheet is None:
                ConfirmationDialog.show_error(self, "Errore", "Foglio non trovato.")
                wb.close()
                return

            # Mappa campi → celle
            cell_map = {
                "data": "A5", "tcl": "A7", "odc": "B5", "avviso": "C7",
                "ordine": "C5", "stato": "D11", "tipo_prev": "D13", "tipo_econ": "E13",
            }

            for key, addr in cell_map.items():
                val = self._fields[key].text().strip()
                if val:
                    sheet[addr] = val

            # Descrizione lavoro (A11:A21)
            desc_lines = self._desc_lavoro_display.toPlainText().split("\n")
            for i in range(11, 22):
                idx = i - 11
                if idx < len(desc_lines):
                    sheet[f"A{i}"] = desc_lines[idx]
                else:
                    sheet[f"A{i}"] = ""

            wb.save(self.loaded_file)
            wb.close()

            self.log_widget.append_log("💾 Modifiche salvate con successo!", "success")
            self._status_label.setText("✅ Modifiche salvate nel file Excel")
            self._status_label.setStyleSheet("color: #2E7D32; font-size: 12px; font-weight: 600; border: none;")
            ConfirmationDialog.show_info(
                self, "Salvato", f"Le modifiche sono state salvate in:\n{os.path.basename(self.loaded_file)}"
            )
        except ImportError:
            ConfirmationDialog.show_error(self, "Errore", "openpyxl non disponibile.")
        except Exception as e:
            self.log_widget.append_log(f"❌ Errore salvataggio: {e}", "error")
            ConfirmationDialog.show_error(self, "Errore Salvataggio", str(e))


    def _on_workflow_step(self, step_id: str) -> None:
        macros = self.workflow_map.get_macros_for_step(step_id)
        if not macros:
            return
        if not self.loaded_file or not os.path.exists(self.loaded_file):
            ConfirmationDialog.show_error(self, "Errore", "Nessun file selezionato.")
            return
        self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ACTIVE)
        self.log_widget.append_log(
            f"🔄 {', '.join(macros)} su {os.path.basename(self.loaded_file)}", "step",
        )
        self.setEnabled(False)
        self.macro_worker = MacroWorker(self.loaded_file, macros)
        self.macro_worker.finished_signal.connect(
            lambda ok, msg: self._on_macro_finished(ok, msg, step_id)
        )
        self.macro_worker.start()

    def _on_macro_finished(self, success: bool, result: str, step_id: str) -> None:
        self.setEnabled(True)
        if success:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.COMPLETED)
            self.log_widget.append_log(f"✅ {result}", "success")
            ConfirmationDialog.show_info(self, "Macro Eseguite", result)
        else:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ERROR)
            self.log_widget.append_log(f"❌ {result}", "error")
            ConfirmationDialog.show_error(self, "Errore Macro", result)







        icon_label = QLabel("📂")
        icon_label.setFont(QFont("Segoe UI", 32))
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_label.setStyleSheet("background: transparent;")
        file_layout.addWidget(icon_label)

        self._file_info_label = QLabel("Nessun file caricato")
        self._file_info_label.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
        self._file_info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._file_info_label.setStyleSheet(f"color: {COLORS['text_muted']}; background: transparent;")
        file_layout.addWidget(self._file_info_label)

        load_btn = QPushButton("📥 Carica File Consuntivo (.xlsm)")
        load_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        load_btn.setFixedHeight(44)
        load_btn.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        load_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1565c0, stop:1 #0288d1);
                color: white; border: none; border-radius: 12px;
                padding: 10px 32px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0d47a1, stop:1 #01579b);
            }
        """)
        load_btn.clicked.connect(self._load_file)
        file_layout.addWidget(load_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(file_section)

        # Mappa Workflow
        self.workflow_map = WorkflowMapWidget()
        self.workflow_map.step_clicked.connect(self._on_workflow_step)
        layout.addWidget(self.workflow_map)

        # Log
        self.log_widget = OperationLogWidget()
        self.log_widget.setMinimumHeight(160)
        layout.addWidget(self.log_widget)

        layout.addStretch()
        scroll.setWidget(content)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _load_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleziona Consuntivo",
            "", "Excel Macro Files (*.xlsm);;All Files (*)",
        )
        if not path:
            return

        self.loaded_file = path
        filename = os.path.basename(path)
        self._file_info_label.setText(f"✅ {filename}")
        self._file_info_label.setStyleSheet(
            "color: #2E7D32; background: transparent; font-weight: 600;"
        )
        self.log_widget.append_log(f"📂 File caricato: {filename}", "info")
        self.log_widget.append_log("ℹ️ I pulsanti della mappa workflow sono ora attivi.", "info")

    def _on_workflow_step(self, step_id: str) -> None:
        macros = self.workflow_map.get_macros_for_step(step_id)
        if not macros:
            return

        if not self.loaded_file or not os.path.exists(self.loaded_file):
            ConfirmationDialog.show_error(
                self, "Errore", "Nessun file caricato o file non trovato.",
            )
            return

        self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ACTIVE)
        self.log_widget.append_log(
            f"🔄 Esecuzione: {', '.join(macros)} su {os.path.basename(self.loaded_file)}",
            "step",
        )

        self.setEnabled(False)
        self.macro_worker = MacroWorker(self.loaded_file, macros)
        self.macro_worker.finished_signal.connect(
            lambda ok, msg: self._on_macro_finished(ok, msg, step_id)
        )
        self.macro_worker.start()

    def _on_macro_finished(self, success: bool, result: str, step_id: str) -> None:
        self.setEnabled(True)
        if success:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.COMPLETED)
            self.log_widget.append_log(f"✅ {result}", "success")
            ConfirmationDialog.show_info(self, "Macro Eseguite", result)
        else:
            self.workflow_map.set_step_state(step_id, WorkflowStepButton.State.ERROR)
            self.log_widget.append_log(f"❌ {result}", "error")
            ConfirmationDialog.show_error(self, "Errore Macro", result)


# =============================================================================
# CONSUNTIVO PANEL
# =============================================================================
class ConsuntivoPanel(QWidget):
    """Pannello per la gestione dei consuntivi strumentale."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._setup_ui()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # AnimatedTabWidget — stile premium con indicatore glow
        self.tabs = AnimatedTabWidget()
        self.tabs.currentChanged.connect(self._on_tab_changed)

        self._tab_new = CreaNuovoTab()
        self._tab_modify = ModificaEsistenteTab()

        self.tabs.addTab(self._tab_new, "Crea Nuovo")
        self.tabs.addTab(self._tab_modify, "Modifica Esistente")
        layout.addWidget(self.tabs)

    def _on_tab_changed(self, index: int) -> None:
        """Aggiorna il progressivo quando si cambia tab."""
        widget = self.tabs.widget(index)
        if isinstance(widget, CreaNuovoTab):
            widget._update_dynamic_path()
        elif isinstance(widget, ModificaEsistenteTab):
            widget._scan_directory()

