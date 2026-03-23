"""
SyncroJob - Timesheet Processing Logic (VBA Replacement)
Replica fedelmente la logica VBA "ProcessTimesheetFiles" per elaborazione, pulizia e rinomina.
"""

import time
from contextlib import suppress
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from src.utils.secure_logger import get_secure_logger

logger = get_secure_logger("TimesheetProcessor")


class TimesheetProcessor:
    """Classe per elaborare i file timesheet sostituendo la macro VBA."""

    @staticmethod
    def process_and_move(file_path: Path, dest_dir: Path) -> tuple[bool, str]:
        """
        Elabora il file Excel secondo la logica VBA e lo salva nella cartella di destinazione.
        """
        if not file_path.exists():
            return False, f"File sorgente non trovato: {file_path}"

        try:
            dest_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return False, f"Impossibile creare dest_dir: {e}"

        try:
            wb = openpyxl.load_workbook(file_path)
            if "Timesheet" not in wb.sheetnames:
                wb.close()
                return False, "Foglio 'Timesheet' non trovato."

            ws: Any = wb["Timesheet"]

            # 1. Estrazione Metadata (ODC, POS)
            odc = str(ws["A2"].value).strip() if ws["A2"].value else ""
            if not odc:
                wb.close()
                return False, "Valore ODC (cella A2) mancante."

            pos_values, first_pos_cleaned = TimesheetProcessor._analyze_pos_column(ws)

            # 2. Generazione Percorso Destinazione
            dest_path = TimesheetProcessor._get_destination_path(dest_dir, odc, pos_values, first_pos_cleaned)

            # 3. Trasformazione Foglio (Headers, Pulizia, Eliminazione, Autofit)
            TimesheetProcessor._apply_transformations(ws)

            # 4. Salvataggio e Pulizia
            wb.save(dest_path)
            wb.close()

            TimesheetProcessor._cleanup_source(file_path, dest_path)

            return True, f"Salvato in: {dest_path.name}"  # noqa: TRY300

        except Exception as e:
            logger.error(f"Errore elaborazione {file_path.name}: {e}")  # noqa: TRY400
            return False, str(e)

    @staticmethod
    def _analyze_pos_column(ws: Any) -> tuple[set[str], str]:  # noqa: ANN401
        """Analizza la colonna B per contare i POS univoci e pulire il primo valore."""
        pos_values: set[str] = set()
        first_pos_cleaned = ""

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            val = str(row[0].value).strip() if row[0].value is not None else ""
            if val:
                pos_values.add(val)
                if not first_pos_cleaned:
                    first_pos_cleaned = TimesheetProcessor._clean_pos_value(val)

        return pos_values, first_pos_cleaned

    @staticmethod
    def _clean_pos_value(val: str) -> str:
        """Converte un valore POS in stringa intera pulita (es. '10.0' -> '10')."""
        if val.replace(".", "", 1).isdigit():
            try:
                return str(int(float(val)))
            except Exception:
                return val
        return val

    @staticmethod
    def _get_destination_path(dest_dir: Path, odc: str, pos_values: set[str], pos_cleaned: str) -> Path:
        """Determina il nome file finale gestendo conflitti e conteggio POS."""
        base = f"{odc}_TS" if len(pos_values) > 1 else f"{odc}_{pos_cleaned}_TS"

        dest_path = dest_dir / f"{base}.xlsx"
        if dest_path.exists():
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            dest_path = dest_dir / f"{base}_{timestamp}.xlsx"

        return dest_path

    @staticmethod
    def _apply_transformations(ws: Any) -> None:  # noqa: ANN401
        """Applica tutte le modifiche strutturali al foglio di lavoro."""
        # 1. Rinomina Intestazioni
        headers = {
            "B1": "POS",
            "C1": "Data",
            "N1": "Ing",
            "O1": "Usc",
            "P1": "Tot",
            "Q1": "Pre",
            "R1": "ORE_C",
            "S1": "ORE_M",
            "T1": "ORE_ST_NOT",
            "U1": "ORE_ST_DIU",
            "V1": "ORE_FEST_NOT",
            "W1": "ORE_FEST_DIU",
        }
        for cell, val in headers.items():
            ws[cell].value = val

        # 2. Pulizia numerica colonna B
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            cell = row[0]
            if cell.value is not None:
                s_val = str(cell.value).strip()
                if s_val.replace(".", "", 1).isdigit():
                    with suppress(Exception):
                        cell.value = int(float(s_val))
                        cell.number_format = "0"

        # 3. Eliminazione Colonne (ordine inverso)
        cols = ["AC", "Z", "X", "L", "I", "H", "G", "F", "E", "D", "A"]
        indices = sorted([column_index_from_string(c) for c in cols], reverse=True)
        for idx in indices:
            ws.delete_cols(idx)

        # 4. Autofit stimato
        TimesheetProcessor._autofit_columns(ws)

    @staticmethod
    def _autofit_columns(ws: Any) -> None:  # noqa: ANN401
        """Regola la larghezza delle colonne in base al contenuto."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                with suppress(Exception):
                    val_len = len(str(cell.value))
                    max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2

    @staticmethod
    def _cleanup_source(src: Path, dest: Path) -> None:
        """Rimuove il file sorgente se diverso dalla destinazione."""
        with suppress(Exception):
            if src.resolve() != dest.resolve():
                src.unlink()
