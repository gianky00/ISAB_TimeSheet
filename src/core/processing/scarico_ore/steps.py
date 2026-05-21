"""Passaggi di elaborazione per l'importazione dello Scarico Ore."""

import io
import json
import re
import warnings
from contextlib import suppress
from pathlib import Path
from typing import Any, ClassVar

from src.core.constants import Business
from src.core.logging import get_logger
from src.core.processing.base import ProcessingStep

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = get_logger(__name__)


class LoadScaricoOreStep(ProcessingStep):
    """Carica e decritta il file Excel dello Scarico Ore."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il caricamento del file Excel scarico ore."""
        file_path = Path(context["file_path"])

        if not HAS_OPENPYXL:
            context["success"] = False
            context["message"] = "Modulo 'openpyxl' mancante."
            return

        try:
            wb_data = self._load_scarico_workbook(file_path)
            if "SCARICO ORE" not in wb_data.sheetnames:
                context["success"] = False
                context["message"] = "Foglio 'SCARICO ORÈ non trovato."
                return

            context["ws"] = wb_data["SCARICO ORE"]
            context["success"] = True

        except Exception as e:
            context["success"] = False
            context["message"] = f"Errore caricamento workbook: {e}"

    def _load_scarico_workbook(self, path: Path) -> Any:
        """Carica il workbook Excel, decrittandolo se necessario."""
        wb_file = io.BytesIO()
        is_encrypted = False

        if msoffcrypto:
            with suppress(Exception), path.open("rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=Business.DEFAULT_EXCEL_PASSWORD)  # nosec: B106
                office_file.decrypt(wb_file)
                is_encrypted = True

        if not is_encrypted:
            with path.open("rb") as f:
                wb_file.write(f.read())

        wb_file.seek(0)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            return openpyxl.load_workbook(
                wb_file,
                data_only=True,
                read_only=True,
                keep_vba=False,
                keep_links=False,
            )


class ProcessScaricoOreRowsStep(ProcessingStep):
    """Processa le righe dello Scarico Ore estraendo stili e dati."""

    COL_KEYS: ClassVar[list[str]] = [
        "data",
        "pers1",
        "pers2",
        "odc",
        "pos",
        "dalle",
        "alle",
        "totale_ore",
        "descrizione",
        "finito",
        "commessa",
    ]

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il processamento delle righe."""
        if not context.get("success") or "ws" not in context:
            return

        ws = context["ws"]
        progress_callback = context.get("progress_callback")

        start_row = 6
        total_rows: int = ws.max_row
        rows_to_insert: list[tuple[Any, ...]] = []
        rows_to_insert_append = rows_to_insert.append
        progress_interval = 5000

        try:
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=start_row, min_col=2, max_col=12, values_only=False),
                start=start_row,
            ):
                if progress_callback and row_idx % progress_interval == 0:
                    progress_callback(row_idx, total_rows)

                db_row = self._process_scarico_ore_row(row)
                if db_row:
                    rows_to_insert_append(db_row)

            context["rows"] = rows_to_insert
            context["message"] = f"Importate {len(rows_to_insert)} righe da Scarico Ore."
        except Exception as e:
            context["success"] = False
            context["message"] = f"Errore processamento righe: {e}"

    def _process_scarico_ore_row(self, row: Any) -> tuple[Any, ...] | None:
        """Processa una singola riga estraendo valori e stili."""
        vals = self._extract_row_values(row)
        if not vals:
            return None

        if not self._validate_scarico_row(vals):
            return None

        styles_json = self._extract_row_styles(row, vals)

        return (
            vals[0],
            vals[1],
            vals[2],
            vals[3],
            vals[4],
            vals[5],
            vals[6],
            vals[7],
            vals[8],
            vals[9],
            vals[10],
            styles_json,
        )

    def _extract_row_values(self, row: Any) -> list[str] | None:
        """Estrae i valori testuali dalle celle della riga."""
        c_data, c_p1, c_p2, c_odc, c_pos, c_dalle, c_alle, c_tot, c_desc, c_fin, c_comm = row[0:11]

        if c_odc.value is None and c_pos.value is None:
            return None

        s_data = self._fmt_excel_date(c_data.value)
        s_p1 = self._fmt_excel_val(c_p1.value)
        s_p2 = self._fmt_excel_val(c_p2.value)

        def _clean_zero(v: Any) -> str:
            s = self._fmt_excel_val(v)
            return "" if s in ("0", "0.0") else s

        s_odc = _clean_zero(c_odc.value)
        s_pos = _clean_zero(c_pos.value)
        s_dalle = self._fmt_excel_val(c_dalle.value)
        s_alle = self._fmt_excel_val(c_alle.value)
        s_tot = self._fmt_excel_val(c_tot.value)
        s_desc = self._fmt_excel_val(c_desc.value)
        s_fin = self._fmt_excel_val(c_fin.value)
        s_comm = _clean_zero(c_comm.value)

        return [
            s_data,
            s_p1,
            s_p2,
            s_odc,
            s_pos,
            s_dalle,
            s_alle,
            s_tot,
            s_desc,
            s_fin,
            s_comm,
        ]

    def _extract_row_styles(self, row: Any, vals: list[str]) -> str:
        """Estrae i colori (foreground/background) dalle celle e li serializza in JSON."""
        row_styles: dict[str, dict[str, str]] = {}
        for i, key in enumerate(self.COL_KEYS):
            if vals[i] == "":
                continue

            cell = row[i]
            with suppress(AttributeError, TypeError):
                font = cell.font
                if font and font.color and font.color.type == "rgb":
                    rgb = str(font.color.rgb)
                    hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"  # noqa: PLR2004
                    if hex_code != "#000000":
                        row_styles.setdefault(key, {})["fg"] = hex_code

            with suppress(AttributeError, TypeError):
                fill = cell.fill
                if fill and fill.patternType == "solid":
                    start_color = fill.start_color
                    if start_color and start_color.type == "rgb":
                        rgb = str(start_color.rgb)
                        hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"  # noqa: PLR2004
                        if hex_code not in {"#000000", "#FFFFFF"}:
                            row_styles.setdefault(key, {})["bg"] = hex_code

        return json.dumps(row_styles) if row_styles else ""

    @staticmethod
    def _validate_scarico_row(vals: list[str]) -> bool:
        if not vals[3] or not vals[4] or not vals[7]:
            return False
        return bool(vals[1] or vals[2])

    @staticmethod
    def _fmt_excel_val(val: Any) -> str:
        if val is None:
            return ""
        if isinstance(val, (float, int)) and float(val).is_integer():
            val = int(val)
        s = str(val)
        return re.sub(r"\s+", " ", s).strip()

    @staticmethod
    def _fmt_excel_date(val: Any) -> str:
        if not val:
            return ""
        if hasattr(val, "strftime"):
            return str(val.strftime("%Y-%m-%d"))
        s = str(val).strip()
        return s.split(" ")[0] if " " in s else s


class SyncScaricoOreStep(ProcessingStep):
    """Passaggio per la sincronizzazione dello scarico ore con il database."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue la sincronizzazione dello scarico ore nel database."""
        if not context.get("success"):
            return

        rows = context.get("rows", [])
        if not rows:
            return

        from src.core.data_synchronizer import DataSynchronizer  # noqa: PLC0415
        from src.core.database import db_manager  # noqa: PLC0415

        total_added, total_removed = DataSynchronizer.sync_scarico_ore(db_manager.DB_SCARICO_ORE, rows)

        context["total_added"] = total_added
        context["total_removed"] = total_removed
        context["success"] = True
