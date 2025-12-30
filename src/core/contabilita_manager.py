"""
Bot TS - Contabilita Manager
Gestione dell'importazione e archiviazione dati della Contabilità Strumentale.
"""
import sqlite3
import pandas as pd
from pathlib import Path
import re
import logging
import warnings
import io
import json
import zipfile
from typing import List, Dict, Tuple, Optional, Callable
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from src.utils.parsing import parse_currency
from src.core.config_manager import CONFIG_DIR
from src.core.database import db_manager

# Tentativo di importare msoffcrypto
try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

# Tentativo di importare openpyxl
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

class ContabilitaManager:
    """Manager per la gestione del database e dell'importazione Excel."""

    DB_PATH = CONFIG_DIR / "data" / "contabilita.db"

    # Mapping colonne Excel -> DB (Contabilità / Dati)
    COLUMNS_MAPPING = {
        'DATA PREV.': 'data_prev',
        'MESE': 'mese',
        'N°PREV.': 'n_prev',
        'TOTALE PREV.': 'totale_prev',
        "ATTIVITA'": 'attivita',
        'TCL': 'tcl',
        'ODC': 'odc',
        "STATO ATTIVITA'": 'stato_attivita',
        'TIPOLOGIA': 'tipologia',
        'ORE SP': 'ore_sp',
        'RESA': 'resa',
        'ANNOTAZIONI': 'annotazioni',
        'INDIRIZZO CONSUNTIVO': 'indirizzo_consuntivo',
        'NOME FILE': 'nome_file'
    }

    # Mapping colonne Excel -> DB (Giornaliere)
    GIORNALIERE_MAPPING = {
        'DATA': 'data',
        'PERSONALE': 'personale',
        "DESCRIZIONE ATTIVITA'": 'descrizione',
        'TCL': 'tcl',
        'ODC': 'odc',
        'N° PDL': 'pdl',
        'INIZIO': 'inizio',
        'FINE': 'fine',
        'ORE': 'ore',
        'consuntivo': 'n_prev' # Rinominato come richiesto
    }

    # Mapping Scarico Ore Cantiere
    SCARICO_ORE_COLS = [
        'data', 'pers1', 'pers2', 'odc', 'pos', 'dalle', 'alle',
        'totale_ore', 'descrizione', 'finito', 'commessa', 'styles'
    ]

    # Mapping Attività Programmate
    # Header: PS, AREA, PdL, IMP., DESCRIZIONE ATTIVITA', LUN, MAR, MER, GIO, VEN, STATO PdL, STATO ATTIVITA', DATA CONTROLLO, PERSONALE IMPIEGATO, PO, AVVISO
    ATTIVITA_PROGRAMMATE_MAPPING = {
        'PS': 'ps',
        'AREA': 'area',
        'PdL': 'pdl',
        'IMP.': 'imp',
        "DESCRIZIONE\nATTIVITA'": 'descrizione',
        'LUN': 'lun',
        'MAR': 'mar',
        'MER': 'mer',
        'GIO': 'gio',
        'VEN': 'ven',
        "STATO\nPdL": 'stato_pdl',
        "STATO\nATTIVITA'": 'stato_attivita',
        "DATA\nCONTROLLO": 'data_controllo',
        "PERSONALE\nIMPIEGATO": 'personale',
        'PO': 'po',
        'AVVISO': 'avviso'
    }

    ATTIVITA_PROGRAMMATE_COLS = list(ATTIVITA_PROGRAMMATE_MAPPING.values()) + ['styles'] # Added styles

    # Mapping Certificati Campione
    CERTIFICATI_CAMPIONE_MAPPING = {
        'Modello / Tipo': 'modello',
        'Costruttore': 'costruttore',
        'Matricola': 'matricola',
        'Range Strumento': 'range_strumento',
        'Errore max %': 'errore_max',
        'Certificato Taratura': 'certificato',
        'Scadenza Certificato': 'scadenza',
        'Emissione Certificato': 'emissione',
        'ID-COEMI': 'id_coemi',
        'Stato Certificato': 'stato'
    }

    CERTIFICATI_CAMPIONE_COLS = list(CERTIFICATI_CAMPIONE_MAPPING.values())

    @classmethod
    def scan_scarico_ore_rows(cls, file_path: str) -> int:
        """Stima rapida delle righe per Scarico Ore (DataEase) per calcolo ETA."""
        path = Path(file_path)
        if not path.exists(): return 0

        # Use zipfile to read dimension from xml without full load
        try:
            with zipfile.ZipFile(path, 'r') as z:
                # Try to find the sheet. Usually sheet1 if it's the only one, or lookup in workbook.xml
                # For speed, assume "SCARICO ORE" is likely one of the first few sheets or search largest xml.

                # Check worksheet rels or just check dimensions in all worksheets and take the largest?
                # Faster: Parse xl/worksheets/sheetX.xml and look for <dimension ref="A1:L130000"/>

                # Let's try to find the sheet "SCARICO ORE" via workbook.xml if possible, but
                # iterating all sheet xmls is fast enough.
                max_rows = 0
                for name in z.namelist():
                    if name.startswith("xl/worksheets/sheet"):
                        with z.open(name) as f:
                            # Read first 1024 bytes which usually contain <dimension>
                            head = f.read(1024).decode('utf-8', errors='ignore')
                            match = re.search(r'<dimension ref="[A-Z]+[0-9]+:[A-Z]+(\d+)"', head)
                            if match:
                                r = int(match.group(1))
                                if r > max_rows: max_rows = r
                return max_rows
        except Exception:
            return 0

    @classmethod
    def scan_workload(cls, file_path: str, giornaliere_path: str) -> Tuple[int, int]:
        """Scansiona rapidamente il carico di lavoro (fogli e file) per stima ETA."""
        sheets = 0
        files = 0

        # 1. Scan Excel Sheets (Fast via ZipFile)
        p_file = Path(file_path)
        if file_path and p_file.exists():
            try:
                with zipfile.ZipFile(p_file, 'r') as z:
                    if 'xl/workbook.xml' in z.namelist():
                        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
                        # Estrai nomi fogli e filtra per anno
                        sheet_names = re.findall(r'name="([^"]+)"', wb_xml)
                        sheets = len([s for s in sheet_names if re.search(r'(\d{4})', s)])
            except Exception:
                sheets = 1

        # 2. Scan Giornaliere (Files)
        p_giorn = Path(giornaliere_path)
        if giornaliere_path and p_giorn.exists():
             current_year = datetime.now().year
             for folder in p_giorn.iterdir():
                 if folder.is_dir():
                     match = re.match(r'Giornaliere\s+(\d{4})', folder.name, re.IGNORECASE)
                     if match:
                         year = int(match.group(1))
                         if year >= current_year:
                             files += len([f for f in folder.glob("*.xls*") if not f.name.startswith("~$")])

        return sheets, files

    @classmethod
    def init_db(cls):
        """Inizializza il database tramite DatabaseManager."""
        db_manager.init_db()

    @classmethod
    def import_data_from_excel(cls, file_path: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> Tuple[bool, str, int, int]:
        """Importa i dati dal file Excel specificato (Tabella Dati)."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File non trovato: {file_path}", 0, 0

        total_added = 0
        total_removed = 0

        try:
            # Gestione file cifrati (Tentativo Decryption)
            file_obj = path
            temp_decrypted = None

            if msoffcrypto:
                try:
                    with open(path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password="coemi")
                        temp_decrypted = io.BytesIO()
                        office_file.decrypt(temp_decrypted)
                        temp_decrypted.seek(0)
                        file_obj = temp_decrypted
                except Exception:
                    # Non cifrato o errore msoffcrypto, procediamo col file originale
                    file_obj = path

            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                # Use standard engine (openpyxl) since calamine is missing
                try:
                    xls = pd.ExcelFile(file_obj)
                except Exception:
                    # Fallback if standard fails (e.g. strict format issues)
                    xls = pd.ExcelFile(file_obj, engine='openpyxl')
                
                imported_years = []

                # Use Manager Connection
                with db_manager.get_connection(cls.DB_PATH) as conn:
                    cursor = conn.cursor()

                    # Count valid sheets first for progress
                    valid_sheets = [s for s in xls.sheet_names if re.search(r'(\d{4})', s)]
                    total_sheets = len(valid_sheets)
                    # If no explicit year sheets found, we treat common sheets as current year
                    if total_sheets == 0:
                        fallback_sheets = [s for s in xls.sheet_names if s.lower() in ["dati", "preventivi", "riepilogo"]]
                        if fallback_sheets:
                            total_sheets = len(fallback_sheets)

                    processed_sheets = 0

                    for sheet_name in xls.sheet_names:
                        # Logic: Year Detection
                        year = None
                        match = re.search(r'(\d{4})', sheet_name)

                        if match:
                            year = int(match.group(1))
                            if not (2000 <= year <= 2100):
                                continue
                        elif sheet_name.lower() in ["dati", "preventivi", "riepilogo"]:
                            # Fallback to current year if name matches common defaults
                            year = datetime.now().year
                        else:
                            continue

                        try:
                            # Rilevamento dinamico dell'header (con normalizzazione)
                            preview_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=10)

                            header_row_idx = 1 # Default legacy (riga 2)

                            # Colonne chiave normalizzate (senza spazi/punti)
                            key_cols_norm = ["DATAPREV", "MESE", "NPREV", "TOTALEPREV", "ATTIVITA", "ODC"]

                            for i, row in preview_df.iterrows():
                                # Converte riga in stringa upper e normalizza
                                row_norm = []
                                for val in row.values:
                                    s = str(val).strip().upper()
                                    s = s.replace(" ", "").replace(".", "").replace("°", "")
                                    row_norm.append(s)

                                # Se troviamo almeno 2 match
                                matches = sum(1 for k in key_cols_norm if k in row_norm)
                                if matches >= 2:
                                    header_row_idx = i
                                    break

                            # Leggi con header corretto
                            df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_idx)

                            # Normalizza colonne DF subito
                            df.columns = [str(c).strip().upper() for c in df.columns]

                            if not df.empty: df = df.iloc[:-1] # Drop last row (Total) if exists

                            # Drop full empty rows
                            df.dropna(how='all', inplace=True)

                            if df.empty: continue

                            df['year'] = year

                            # Normalizzazione più robusta per il mapping delle colonne
                            # Prepara mappa normalizzata: "N° PREV." -> "N°PREV." -> "n_prev"
                            normalized_map = {}
                            for k, v in cls.COLUMNS_MAPPING.items():
                                norm_k = k.upper().replace(" ", "").replace(".", "").replace("°", "")
                                normalized_map[norm_k] = v

                            # Mappa effettiva trovata nel DF
                            rename_map = {}
                            for col in df.columns:
                                col_str = str(col).strip().upper()
                                norm_col = col_str.replace(" ", "").replace(".", "").replace("°", "")

                                # Cerca match
                                if norm_col in normalized_map:
                                    rename_map[col] = normalized_map[norm_col]
                                else:
                                    # Fallback per match parziale o specifico
                                    if "PREV" in norm_col and "DATA" in norm_col:
                                        rename_map[col] = "data_prev"
                                    elif "PREV" in norm_col and ("N" in norm_col or "NUM" in norm_col):
                                        rename_map[col] = "n_prev"

                            df.rename(columns=rename_map, inplace=True)

                        except Exception as e:
                            # Log warning but continue with other sheets
                            # print(f"Warning processing sheet {sheet_name}: {e}")
                            pass

                        for db_col in cls.COLUMNS_MAPPING.values():
                            if db_col not in df.columns: df[db_col] = ""

                        target_columns = ['year'] + list(cls.COLUMNS_MAPPING.values())
                        df = df[target_columns]
                        df = df.fillna("")
                        cols_to_str = [c for c in df.columns if c != 'year']
                        df[cols_to_str] = df[cols_to_str].astype(str)
                        # Strip whitespace from all string columns
                        df[cols_to_str] = df[cols_to_str].apply(lambda x: x.str.strip())

                        # --- Diff Logic ---
                        # Fetch existing rows for this year using pandas to ensure identical type/format handling
                        existing_df = pd.read_sql(
                            f"SELECT {', '.join(target_columns)} FROM contabilita WHERE year = ?",
                            conn,
                            params=(year,)
                        )
                        
                        # Apply EXACT SAME cleaning to existing data
                        existing_df = existing_df.fillna("")
                        cols_to_str_ex = [c for c in existing_df.columns if c != 'year']
                        existing_df[cols_to_str_ex] = existing_df[cols_to_str_ex].astype(str)
                        existing_df[cols_to_str_ex] = existing_df[cols_to_str_ex].apply(lambda x: x.str.strip())

                        existing_rows = set(list(existing_df.itertuples(index=False, name=None)))

                        # New rows from DF
                        new_rows_list = list(df.itertuples(index=False, name=None))
                        new_rows_set = set(new_rows_list)

                        added = len(new_rows_set - existing_rows)
                        removed = len(existing_rows - new_rows_set)

                        total_added += added
                        total_removed += removed
                        # ------------------

                        cursor.execute("DELETE FROM contabilita WHERE year = ?", (year,))
                        placeholders = ', '.join(['?'] * len(target_columns))
                        query = f"INSERT INTO contabilita ({', '.join(target_columns)}) VALUES ({placeholders})"
                        cursor.executemany(query, new_rows_list)
                        imported_years.append(year)

                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets)

                    conn.commit()

                if not imported_years: return False, "Nessun anno importato (Controlla nomi fogli: YYYY o 'Dati/Preventivi').", 0, 0
                return True, f"Anni importati: {sorted(list(set(imported_years)))}", total_added, total_removed

        except Exception as e:
            return False, f"Errore: {e}", 0, 0

    @classmethod
    def _process_single_giornaliera(cls, args):
        """Helper per processare un singolo file giornaliera in parallelo."""
        year, file_path, lookup_map = args
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                try:
                    df = pd.read_excel(file_path, sheet_name='RIASSUNTO')
                except ValueError:
                    return (year, [], None) # Sheet not found
                except Exception:
                    # Fallback to openpyxl
                    try:
                        df = pd.read_excel(file_path, sheet_name='RIASSUNTO', engine='openpyxl')
                    except:
                        return (year, [], None)

                df.columns = [str(c).strip() for c in df.columns]


                rename_map = {}
                for excel_col, db_col in cls.GIORNALIERE_MAPPING.items():
                    for c in df.columns:
                        if c.upper() == excel_col.upper():
                            rename_map[c] = db_col
                            break

                if not rename_map: return (year, [], None)

                df.rename(columns=rename_map, inplace=True)
                
                # Exclude 'Totale' rows and always drop the last row (Total)
                if not df.empty:
                    df = df.iloc[:-1] # Always drop last row as requested
                
                if 'personale' in df.columns and not df.empty:
                    df = df[~df['personale'].str.contains("Totale", na=False, case=False)]

                check_cols = [c for c in df.columns if c in cls.GIORNALIERE_MAPPING.values() and c != 'data']
                if check_cols: df.dropna(how='all', subset=check_cols, inplace=True)

                if not df.empty:
                    for db_col in cls.GIORNALIERE_MAPPING.values():
                        if db_col not in df.columns: df[db_col] = ""

                    cols_to_clean = ['odc', 'n_prev', 'data', 'personale', 'descrizione', 'tcl', 'pdl', 'inizio', 'fine', 'ore']
                    df[cols_to_clean] = df[cols_to_clean].astype(str).apply(lambda x: x.str.strip())
                    df[cols_to_clean] = df[cols_to_clean].replace(r'(?i)^nan$', '', regex=True)

                    # Apply Lookup
                    mask_empty_odc = df['odc'] == ""
                    if mask_empty_odc.any() and lookup_map:
                        mapped_values = df.loc[mask_empty_odc, 'n_prev'].map(lookup_map)
                        df.loc[mask_empty_odc, 'odc'] = mapped_values.fillna("")

                    # Se ODC è ancora vuoto, cerca il formato 'commessa' (es. 12/345) nella descrizione
                    mask_still_empty_odc = df['odc'] == ""
                    if mask_still_empty_odc.any():
                        commessa_pattern = r'\b(\d{2}/\d{3})\b'
                        extracted_commessa = df.loc[mask_still_empty_odc, 'descrizione'].str.extract(commessa_pattern, expand=False)
                        df.loc[mask_still_empty_odc, 'odc'] = extracted_commessa.fillna("")
                    
                    # Regex for '5400xxx' format and other cleaning
                    mask_canone = df['odc'].str.contains('canone', case=False, na=False)
                    mask_commessa = df['odc'].str.match(r'^\d{2}/\d{3}$', na=False) # Exclude already valid commessa
                    mask_standard = ~mask_canone & ~mask_commessa
                    extracted = df.loc[mask_standard, 'odc'].str.extract(r'(5400\d+)', expand=False)
                    df.loc[mask_standard, 'odc'] = extracted.fillna("")

                    df['year'] = year
                    df['nome_file'] = file_path.name
                    
                    target_cols = ['year', 'data', 'personale', 'descrizione', 'tcl', 'odc', 'pdl', 'inizio', 'fine', 'ore', 'n_prev', 'nome_file']
                    df_final = df[target_cols]
                    rows = list(df_final.itertuples(index=False, name=None))
                    return (year, rows, None)

                return (year, [], None)
        except Exception as e:
            return (year, [], str(e))

    @classmethod
    def import_giornaliere(cls, root_path: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> Tuple[bool, str, int, int]:
        root = Path(root_path)
        if not root.exists():
            return False, "Directory Giornaliere non trovata.", 0, 0

        current_year = datetime.now().year
        imported_years = set()

        total_added = 0
        total_removed = 0

        try:
            # Cleanup: ensure no data from 2026+ exists (fix for bug)
            with db_manager.get_connection(cls.DB_PATH) as conn:
                conn.execute("DELETE FROM giornaliere WHERE year >= 2026")
                conn.execute("DELETE FROM contabilita WHERE year >= 2026")
                conn.commit()

            # 1. Lookup Map Preparation
            lookup_map = {}
            try:
                with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                    lookup_query = "SELECT n_prev, odc FROM contabilita WHERE odc IS NOT NULL AND odc != ''"
                    lookup_df = pd.read_sql_query(lookup_query, conn)
                    lookup_df = lookup_df.drop_duplicates(subset=['n_prev'])
                    lookup_map = dict(zip(lookup_df['n_prev'], lookup_df['odc']))
            except: pass

            # 2. Collect Tasks
            tasks_args = []
            for folder in root.iterdir():
                if not folder.is_dir(): continue
                match = re.match(r'Giornaliere\s+(\d{4})', folder.name, re.IGNORECASE)
                if not match: continue

                year = int(match.group(1))
                if year > current_year: continue # Skip future years

                for file_path in folder.glob("*.xls*"):
                    if not file_path.name.startswith("~$"):
                        tasks_args.append((year, file_path, lookup_map))

            total_tasks = len(tasks_args)
            processed_count = 0
            
            all_new_rows = [] 
            target_cols = ['year', 'data', 'personale', 'descrizione', 'tcl', 'odc', 'pdl', 'inizio', 'fine', 'ore', 'n_prev', 'nome_file']
            years_encountered = set()

            # 3. Parallel Execution
            if total_tasks > 0:
                with ThreadPoolExecutor(max_workers=4) as executor:
                    for result in executor.map(cls._process_single_giornaliera, tasks_args):
                        processed_count += 1
                        if progress_callback: progress_callback(processed_count, total_tasks)

                        r_year, r_rows, r_err = result
                        if r_rows:
                            all_new_rows.extend(r_rows)
                            years_encountered.add(r_year)
                            imported_years.add(r_year)
                        if r_err:
                            print(f"Errore lettura file (Year {r_year}): {r_err}")

            # 3. Diff and Commit
            with db_manager.get_connection(cls.DB_PATH) as conn:
                conn.row_factory = sqlite3.Row # Temp setting
                cursor = conn.cursor()

                years_to_clear = list(years_encountered) # Only clear years we touched

                # Fetch Existing using pandas for consistent type handling
                existing_rows_set = set()
                if years_to_clear:
                    placeholders = ','.join(['?'] * len(years_to_clear))
                    query = f"SELECT {', '.join(target_cols)} FROM giornaliere WHERE year IN ({placeholders})"
                    
                    existing_df = pd.read_sql(query, conn, params=tuple(years_to_clear))
                    
                    # Clean Existing
                    existing_df = existing_df.fillna("")
                    cols_str_ex = [c for c in existing_df.columns if c != 'year']
                    existing_df[cols_str_ex] = existing_df[cols_str_ex].astype(str).apply(lambda x: x.str.strip())
                    
                    existing_rows_set = set(list(existing_df.itertuples(index=False, name=None)))

                # Prepare New Rows for Diff (Convert back to DF to ensure identical processing)
                if all_new_rows:
                    new_df = pd.DataFrame(all_new_rows, columns=target_cols)
                else:
                    new_df = pd.DataFrame(columns=target_cols)

                new_df = new_df.fillna("")
                cols_str_new = [c for c in new_df.columns if c != 'year']
                new_df[cols_str_new] = new_df[cols_str_new].astype(str).apply(lambda x: x.str.strip())

                new_rows_set = set(list(new_df.itertuples(index=False, name=None)))

                total_added = len(new_rows_set - existing_rows_set)
                total_removed = len(existing_rows_set - new_rows_set)


                # Delete old
                for year in years_to_clear:
                    cursor.execute("DELETE FROM giornaliere WHERE year = ?", (year,))

                # Insert new (Use original all_new_rows or cleaned? Better to use cleaned to match diff)
                # But we must ensure types are correct for SQLite. 
                # all_new_rows was built carefully. new_df is string-ified.
                # Actually, SQLite is flexible. Inserting cleaned strings is safer.
                # Let's use new_df converted back to list of tuples.
                final_rows_to_insert = list(new_df.itertuples(index=False, name=None))

                if final_rows_to_insert:
                    placeholders = ', '.join(['?'] * len(target_cols))
                    query = f"INSERT INTO giornaliere ({', '.join(target_cols)}) VALUES ({placeholders})"
                    cursor.executemany(query, final_rows_to_insert)

                conn.commit()

            if not imported_years and total_tasks == 0:
                return True, "Nessuna nuova giornaliera trovata (check anno >= " + str(current_year) + ").", 0, 0
            return True, f"Importate Giornaliere: {sorted(list(imported_years))}", total_added, total_removed

        except Exception as e:
            return False, f"Errore importazione Giornaliere: {e}", 0, 0

    @classmethod
    def import_attivita_programmate(cls, file_path: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> Tuple[bool, str, int, int]:
        """Importa il file Attività Programmate (veloce, senza colori)."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File Attività Programmate non trovato: {file_path}", 0, 0

        total_added = 0
        total_removed = 0

        try:
            # Use standard engine (openpyxl) since calamine is missing
            # Header is at row 3 (0-based index 2)
            try:
                df = pd.read_excel(path, sheet_name="Riepilogo", header=2)
            except ValueError:
                return False, "Foglio 'Riepilogo' non trovato.", 0, 0
            except Exception as e:
                # Fallback
                try:
                    df = pd.read_excel(path, sheet_name="Riepilogo", header=2, engine='openpyxl')
                except Exception as e2:
                    return False, f"Errore lettura file: {e2}", 0, 0

            # Normalize columns
            df.columns = [str(c).strip() for c in df.columns]

            # Map Columns
            rename_map = {}
            for excel_col, db_col in cls.ATTIVITA_PROGRAMMATE_MAPPING.items():
                # Try exact match or normalized
                if excel_col in df.columns:
                    rename_map[excel_col] = db_col
                else:
                    # Search ignoring newlines or spaces
                    for col in df.columns:
                        if excel_col.replace("\n", " ").strip() == col.replace("\n", " ").strip():
                            rename_map[col] = db_col
                            break
            
            if not rename_map:
                 return False, "Colonne non trovate. Controlla intestazione riga 3.", 0, 0

            df.rename(columns=rename_map, inplace=True)
            
            # Fill missing mapped columns with empty
            for db_col in cls.ATTIVITA_PROGRAMMATE_MAPPING.values():
                if db_col not in df.columns: df[db_col] = ""

            # Filter rows (basic validity check)
            # Check if PS or Area or Desc are present
            check_cols = [c for c in ['ps', 'area', 'descrizione'] if c in df.columns]
            if check_cols:
                df.dropna(how='all', subset=check_cols, inplace=True)

            # Convert to strings and strip
            df = df.fillna("")
            df = df.astype(str)
            df = df.apply(lambda x: x.str.strip())
            
            # Add 'styles' column (Empty for optimization)
            df['styles'] = ""
            
            # Select Final Columns in Order
            db_cols = list(cls.ATTIVITA_PROGRAMMATE_MAPPING.values()) + ['styles']
            
            # Ensure all db_cols exist
            for c in db_cols:
                if c not in df.columns: df[c] = ""
            
            df = df[db_cols]
            
            rows_to_insert = list(df.itertuples(index=False, name=None))

            # DB Update
            with db_manager.get_connection(cls.DB_PATH) as conn:
                cursor = conn.cursor()

                # Diff Logic
                # 1. Fetch Existing
                existing_df = pd.read_sql(f"SELECT {', '.join(db_cols)} FROM attivita_programmate", conn)
                existing_df = existing_df.fillna("")
                existing_df = existing_df.astype(str).apply(lambda x: x.str.strip())
                existing_rows_set = set(list(existing_df.itertuples(index=False, name=None)))

                # 2. Prepare New (Already cleaned via DF)
                new_rows_set = set(rows_to_insert)

                total_added = len(new_rows_set - existing_rows_set)
                total_removed = len(existing_rows_set - new_rows_set)

                cursor.execute("DELETE FROM attivita_programmate")

                if rows_to_insert:
                    placeholders = ', '.join(['?'] * len(db_cols))
                    query = f"INSERT INTO attivita_programmate ({', '.join(db_cols)}) VALUES ({placeholders})"
                    cursor.executemany(query, rows_to_insert)

                conn.commit()

            return True, f"Importate {len(rows_to_insert)} righe in Attività Programmate.", total_added, total_removed

        except Exception as e:
            return False, f"Errore importazione Attività Programmate: {e}", 0, 0

        except Exception as e:
            return False, f"Errore importazione Attività Programmate: {e}", 0, 0

    @classmethod
    def import_scarico_ore(cls, file_path: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> Tuple[bool, str, int, int]:
        """Importa il file Scarico Ore Cantiere (OpenPyXL per colori + Diff Logic)."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File Scarico Ore non trovato: {file_path}", 0, 0

        if not openpyxl:
            return False, "Modulo 'openpyxl' mancante.", 0, 0

        total_added = 0
        total_removed = 0

        try:
            # 1. Decrypt/Load Workbook (OpenPyXL needed for Styles)
            wb_file = io.BytesIO()
            is_encrypted = False

            if msoffcrypto:
                try:
                    with open(path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password="coemi")
                        office_file.decrypt(wb_file)
                        is_encrypted = True
                except Exception:
                    pass

            if not is_encrypted:
                with open(path, "rb") as f:
                    wb_file.write(f.read())

            wb_file.seek(0)

            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
                # read_only=False required for styles
                wb_data = openpyxl.load_workbook(wb_file, data_only=True, read_only=False)

            if "SCARICO ORE" not in wb_data.sheetnames:
                 return False, "Foglio 'SCARICO ORE' non trovato.", 0, 0
            ws_data = wb_data["SCARICO ORE"]

            rows_to_insert = []
            start_row = 6 # Data starts row 6 (1-based)

            # Map Columns (0-based indices for array access if using iter_rows values, but here we use Cells)
            # Excel Cols: B=Data, C=Pers1, D=Pers2, E=ODC, F=POS, G=Dalle, H=Alle, I=TotOre, J=Desc, K=Finito, L=Commessa
            # 1-based indices for openpyxl: B=2 ... L=12
            col_keys = ['data', 'pers1', 'pers2', 'odc', 'pos', 'dalle', 'alle', 'totale_ore', 'descrizione', 'finito', 'commessa']
            col_indices = {k: i + 2 for i, k in enumerate(col_keys)} # data -> 2, pers1 -> 3 ...

            total_rows = ws_data.max_row

            for row_idx, row in enumerate(ws_data.iter_rows(min_row=start_row, min_col=2, max_col=12), start=start_row):
                if progress_callback and row_idx % 200 == 0:
                    progress_callback(row_idx, total_rows)

                # Row is tuple of cells. Index 0 corresponds to min_col (B=2)
                # keys map: 'data' is at index 0 of the tuple
                
                # Check empty row (check vital cols)
                # indices 0(data) to 7(totale_ore)
                subset_vals = [c.value for i, c in enumerate(row) if i <= 7]
                if all(v is None or str(v).strip() == "" for v in subset_vals):
                    continue

                row_vals = {}
                row_styles = {}

                for i, key in enumerate(col_keys):
                    cell = row[i]
                    val = cell.value

                    # Cleaning
                    if key in ['odc', 'pos']:
                        if val == 0 or str(val).strip() in ["0", "0.0"]: val = ""
                    elif key == 'commessa':
                        if val == 0: val = "0"

                    val_str = str(val).strip() if val is not None else ""
                    val_str = val_str.replace('\n', ' ') # Remove newlines in data? Optional.
                    row_vals[key] = val_str

                    # Styles
                    fg_color = None
                    bg_color = None

                    if cell.font and cell.font.color:
                        if cell.font.color.type == 'rgb':
                             c = str(cell.font.color.rgb)
                             if len(c) > 6: c = "#" + c[2:]
                             else: c = "#" + c
                             fg_color = c

                    if cell.fill and cell.fill.patternType == 'solid':
                         if cell.fill.start_color:
                             if cell.fill.start_color.type == 'rgb':
                                 c = str(cell.fill.start_color.rgb)
                                 if len(c) > 6: c = "#" + c[2:]
                                 else: c = "#" + c
                                 bg_color = c

                    if fg_color or bg_color:
                        style_entry = {}
                        if fg_color: style_entry['fg'] = fg_color
                        if bg_color: style_entry['bg'] = bg_color
                        row_styles[key] = style_entry

                # Validation Logic
                
                # 1. Empty Row Check (if all relevant fields are empty)
                check_all_empty = ['pers1', 'pers2', 'odc', 'pos', 'dalle', 'alle', 'totale_ore']
                if all(row_vals.get(k, "") == "" for k in check_all_empty):
                    continue

                # 2. Key Fields Integrity Check
                # Skip if ODC OR POS OR TOTALE_ORE is missing
                if not row_vals.get('odc') or not row_vals.get('pos') or not row_vals.get('totale_ore'):
                    continue
                
                # 3. Personnel Check (At least one person)
                if not row_vals.get('pers1') and not row_vals.get('pers2'):
                    continue

                # Build Row Tuple
                db_row = (
                    row_vals['data'],
                    row_vals['pers1'],
                    row_vals['pers2'],
                    row_vals['odc'],
                    row_vals['pos'],
                    row_vals['dalle'],
                    row_vals['alle'],
                    row_vals['totale_ore'],
                    row_vals['descrizione'],
                    row_vals['finito'],
                    row_vals['commessa'],
                    json.dumps(row_styles) if row_styles else ""
                )
                rows_to_insert.append(db_row)

            # 3. Diff and Update (Standardized)
            with db_manager.get_connection(cls.DB_PATH) as conn:
                db_cols = cls.SCARICO_ORE_COLS
                
                # Fetch Existing
                existing_df = pd.read_sql(f"SELECT {', '.join(db_cols)} FROM scarico_ore", conn)
                existing_df = existing_df.fillna("")
                existing_df = existing_df.astype(str).apply(lambda x: x.str.strip())
                existing_rows_set = set(list(existing_df.itertuples(index=False, name=None)))
                
                # Prepare New (Already mostly clean, but strictly cast to match DF)
                # To be safe, create DF from new rows and clean it exactly same way
                if rows_to_insert:
                    new_df = pd.DataFrame(rows_to_insert, columns=db_cols)
                    new_df = new_df.fillna("")
                    new_df = new_df.astype(str).apply(lambda x: x.str.strip())
                    new_rows_set = set(list(new_df.itertuples(index=False, name=None)))
                else:
                    new_rows_set = set()

                total_added = len(new_rows_set - existing_rows_set)
                total_removed = len(existing_rows_set - new_rows_set)

                cursor = conn.cursor()
                cursor.execute("DELETE FROM scarico_ore")

                if rows_to_insert:
                    placeholders = ', '.join(['?'] * len(db_cols))
                    query = f"INSERT INTO scarico_ore ({', '.join(db_cols)}) VALUES ({placeholders})"
                    cursor.executemany(query, rows_to_insert)

                conn.commit()

            return True, f"Importate {len(rows_to_insert)} righe da Scarico Ore.", total_added, total_removed

        except Exception as e:
            return False, f"Errore importazione Scarico Ore: {e}", 0, 0

        except Exception as e:
            return False, f"Errore importazione Scarico Ore: {e}", 0, 0

    @classmethod
    def get_available_years(cls) -> List[int]:
        """Restituisce la lista degli anni presenti nel DB (unione di Dati e Giornaliere)."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT year FROM contabilita UNION SELECT DISTINCT year FROM giornaliere ORDER BY 1 DESC")
                years = [row[0] for row in cursor.fetchall()]
                return years
        except: return []

    @classmethod
    def get_data_by_year(cls, year: int) -> List[Tuple]:
        """Restituisce i dati tabella Dati per un anno specifico."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cols = [
                    'data_prev', 'mese', 'n_prev', 'totale_prev', 'attivita', 'tcl', 'odc',
                    'stato_attivita', 'tipologia', 'ore_sp', 'resa', 'annotazioni',
                    'indirizzo_consuntivo', 'nome_file'
                ]
                query = f"SELECT {', '.join(cols)} FROM contabilita WHERE year = ? ORDER BY n_prev DESC, id DESC"
                cursor.execute(query, (year,))
                rows = cursor.fetchall()
                return rows
        except: return []

    @classmethod
    def get_giornaliere_by_year(cls, year: int) -> List[Tuple]:
        """Restituisce i dati Giornaliere per un anno specifico."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cols = ['data', 'personale', 'tcl', 'descrizione', 'n_prev', 'odc', 'pdl', 'inizio', 'fine', 'ore', 'nome_file']
                query = f"SELECT {', '.join(cols)} FROM giornaliere WHERE year = ? ORDER BY data DESC, id DESC"
                cursor.execute(query, (year,))
                rows = cursor.fetchall()
                return rows
        except: return []

    @classmethod
    def get_attivita_programmate_data(cls) -> List[Tuple]:
        """Restituisce i dati Attività Programmate (inclusi stili)."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cols = cls.ATTIVITA_PROGRAMMATE_COLS
                query = f"SELECT {', '.join(cols)} FROM attivita_programmate ORDER BY id ASC"
                cursor.execute(query)
                rows = cursor.fetchall()
                return rows
        except: return []

    @classmethod
    def import_certificati_campione(cls, file_path: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> Tuple[bool, str, int, int]:
        """Importa il file Certificati Campione."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File Certificati Campione non trovato: {file_path}", 0, 0

        total_added = 0
        total_removed = 0

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")

                # 1. Identify Sheet
                try:
                    # Remove engine='calamine' to use default (openpyxl)
                    xls = pd.ExcelFile(path)
                    sheet_name = None
                    
                    # Try to find specific sheet (case-insensitive partial match)
                    for name in xls.sheet_names:
                        name_lower = name.lower()
                        if "strumenti campione" in name_lower or "isab sud" in name_lower:
                            sheet_name = name
                            break
                    
                    # Fallback to first sheet if not found
                    if not sheet_name and xls.sheet_names:
                        sheet_name = xls.sheet_names[0]
                        
                    if not sheet_name:
                        return False, "Nessun foglio trovato nel file Excel.", 0, 0

                except Exception as e:
                     return False, f"Errore apertura file Excel: {e}", 0, 0

                # 2. Find Header Row
                try:
                    # Read first 20 rows to find header
                    df_preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=20)
                    
                    header_row_idx = -1
                    max_matches = 0
                    
                    target_columns = set(cls.CERTIFICATI_CAMPIONE_MAPPING.keys())
                    
                    for i, row in df_preview.iterrows():
                        row_values = [str(val).strip() for val in row.values]
                        matches = sum(1 for col in target_columns if col in row_values)
                        
                        # We expect at least a few columns to match (e.g. Matricola, Costruttore, ecc.)
                        if matches > max_matches:
                            max_matches = matches
                            header_row_idx = i
                    
                    if header_row_idx == -1 or max_matches < 3: # Require at least 3 matching columns
                        # Fallback: try hardcoded index 5 if search fails
                        header_row_idx = 5
                        
                    # 3. Read Data with correct header
                    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row_idx)

                except Exception as e:
                     return False, f"Errore lettura file Certificati (sheet: {sheet_name}): {e}", 0, 0

                if df.empty: return False, "Foglio vuoto.", 0, 0

                # Rename columns
                df.columns = [str(c).strip() for c in df.columns]

                # Check mapping
                rename_map = {}
                for excel_col, db_col in cls.CERTIFICATI_CAMPIONE_MAPPING.items():
                    if excel_col in df.columns:
                        rename_map[excel_col] = db_col

                if not rename_map:
                     found_cols = ", ".join(list(df.columns)[:5]) + "..." # Show first 5 cols for context
                     return False, f"Nessuna colonna valida trovata. Sheet: {sheet_name}, Row: {header_row_idx}. Trovate: {found_cols}", 0, 0

                df.rename(columns=rename_map, inplace=True)

                # Filter cols
                target_cols = list(cls.CERTIFICATI_CAMPIONE_MAPPING.values())
                # Add missing
                for c in target_cols:
                    if c not in df.columns: df[c] = ""

                df = df[target_cols]

                # Filter empty rows (mandatory check on matricola or id_coemi?)
                df.dropna(how='all', inplace=True)

                # Format Dates: scadenza, emissione -> DD/MM/YYYY
                def format_date_it(val):
                    if pd.isna(val) or val == "": return ""
                    try:
                        dt = pd.to_datetime(val)
                        return dt.strftime("%d/%m/%Y")
                    except:
                        return str(val)

                df['scadenza'] = df['scadenza'].apply(format_date_it)
                df['emissione'] = df['emissione'].apply(format_date_it)

                # Format Stato: Transform numeric values (days diff) to user-friendly string
                # If Excel formula returns a number like 133 or -985, we format it.
                def format_stato(val):
                    if pd.isna(val) or val == "": return ""
                    try:
                        # Try parsing as float first
                        num = float(val)
                        days = int(round(num))
                        if days > 0:
                            return f"Scade tra {days} giorni"
                        elif days < 0:
                            return f"Scaduto da {abs(days)} giorni"
                        else:
                            return "Scade oggi"
                    except ValueError:
                        # Not a number, maybe already text or invalid
                        return str(val)

                if 'stato' in df.columns:
                    df['stato'] = df['stato'].apply(format_stato)

                # Fill N/A and convert to str
                df = df.fillna("")
                df = df.astype(str)
                # Apply strip to ensure clean comparison
                df = df.apply(lambda x: x.str.strip())

                rows = list(df.itertuples(index=False, name=None))
                new_rows_set = set(rows)

                # DB Ops
                with db_manager.get_connection(cls.DB_PATH) as conn:
                    # 1. Fetch Existing for Diff
                    existing_df = pd.read_sql(f"SELECT {', '.join(target_cols)} FROM certificati_campione", conn)
                    existing_df = existing_df.fillna("")
                    existing_df = existing_df.astype(str).apply(lambda x: x.str.strip())
                    existing_rows_set = set(list(existing_df.itertuples(index=False, name=None)))

                    # 2. Calc Diff
                    added = len(new_rows_set - existing_rows_set)
                    removed = len(existing_rows_set - new_rows_set)

                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM certificati_campione")

                    if rows:
                        placeholders = ', '.join(['?'] * len(target_cols))
                        query = f"INSERT INTO certificati_campione ({', '.join(target_cols)}) VALUES ({placeholders})"
                        cursor.executemany(query, rows)

                    conn.commit()

                return True, f"Importate {len(rows)} righe in Certificati Campione.", added, removed

        except Exception as e:
            return False, f"Errore importazione Certificati Campione: {e}", 0, 0

    @classmethod
    def get_certificati_campione_data(cls) -> List[Tuple]:
        """Restituisce i dati Certificati Campione."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cols = cls.CERTIFICATI_CAMPIONE_COLS
                query = f"SELECT {', '.join(cols)} FROM certificati_campione ORDER BY id ASC"
                cursor.execute(query)
                rows = cursor.fetchall()
                return rows
        except: return []

    @classmethod
    def get_scarico_ore_data(cls) -> List[Tuple]:
        """Restituisce tutti i dati della tabella scarico_ore inclusi gli stili."""
        if not cls.DB_PATH.exists(): return []
        try:
            with db_manager.get_connection(cls.DB_PATH, read_only=True) as conn:
                cursor = conn.cursor()
                cols = cls.SCARICO_ORE_COLS
                query = f"SELECT {', '.join(cols)} FROM scarico_ore ORDER BY id DESC"
                cursor.execute(query)
                rows = cursor.fetchall()
                return rows
        except:
            return []

    @classmethod
    def get_year_stats(cls, year: int) -> Dict:
        """Calcola statistiche avanzate per l'anno specificato (Tabella Dati) + KPI Diretti/Indiretti."""
        data = cls.get_data_by_year(year)

        # Recupera anche Giornaliere per calcolo Dirette/Indirette (basato su richiesta utente)
        # "Se in giornaliera, una riga è associata ad un n°prev oppure ODC, allora è una spesa ore diretta altrimenti è una spesa ore indiretta."
        giornaliere = cls.get_giornaliere_by_year(year)

        stats = {
            "total_prev": 0.0,
            "total_ore": 0.0,
            "count_total": 0,
            "status_counts": {},
            "top_commesse": [],
            # New Metrics
            "ore_dirette": 0.0,
            "ore_indirette": 0.0
        }

        # 1. Stats from Dati (Contabilita)
        commesse = []
        if data:
            for row in data:
                try:
                    # row indices: 2=n_prev, 3=totale_prev, 4=attivita, 7=stato, 9=ore_sp, 10=resa
                    n_prev = str(row[2]).strip()
                    if not n_prev: continue
                    if "totale" in n_prev.lower(): continue

                    val_prev = parse_currency(row[3])
                    val_ore = parse_currency(row[9]) # Ore from 'Dati' table

                    stats["total_prev"] += val_prev
                    stats["total_ore"] += val_ore
                    stats["count_total"] += 1

                    status = str(row[7]).strip().upper()
                    if status:
                        stats["status_counts"][status] = stats["status_counts"].get(status, 0) + 1

                    if val_prev > 0:
                        attivita = str(row[4]).strip() or "N/D"
                        commesse.append((attivita, val_prev))
                except:
                    pass

        stats["top_commesse"] = sorted(commesse, key=lambda x: x[1], reverse=True)[:5]

        # 2. Stats from Giornaliere (Direct vs Indirect)
        # Giornaliere Cols: data, personale, tcl, descrizione, n_prev, odc, pdl, inizio, fine, ore, nome_file
        # Index: 4=n_prev, 5=odc, 9=ore
        if giornaliere:
            for row in giornaliere:
                try:
                    n_prev = str(row[4]).strip()
                    odc = str(row[5]).strip()
                    # Clean placeholders
                    if n_prev.lower() == 'nan': n_prev = ""
                    if odc.lower() == 'nan': odc = ""

                    ore = parse_currency(row[9])

                    # Logic: Associated with N_PREV OR ODC -> Direct
                    if n_prev or odc:
                        stats["ore_dirette"] += ore
                    else:
                        stats["ore_indirette"] += ore
                except:
                    pass

        return stats
