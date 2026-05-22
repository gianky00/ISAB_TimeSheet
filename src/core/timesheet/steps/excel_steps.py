"""Passaggi di elaborazione Excel per i Timesheet."""

import time
from contextlib import suppress
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

from src.core.processing.base import ProcessingStep


class LoadWorkbookStep(ProcessingStep):
    """Carica il file Excel e individua il foglio Timesheet."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il caricamento del file."""
        file_path = context["file_path"]
        wb = openpyxl.load_workbook(file_path)
        if "Timesheet" not in wb.sheetnames:
            raise ValueError("Foglio 'Timesheet' non trovato.")
        context["wb"] = wb
        context["ws"] = wb["Timesheet"]


class TransformSheetStep(ProcessingStep):
    """Esegue la trasformazione strutturale del foglio (rinomina intestazioni, elimina colonne)."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue la trasformazione dei dati nel foglio."""
        ws = context["ws"]

        # Metadata extraction
        odc = str(ws["A2"].value).strip() if ws["A2"].value else ""
        if not odc:
            raise ValueError("Valore ODC mancante.")
        context["odc"] = odc

        # Logica di trasformazione (da vecchio processor)
        headers = {
            "B1": "POS",
            "C1": "Data",
            "N1": "Ing",
            "O1": "Usc",
            "P1": "Tot",
            "Q1": "Pre",
            "R1": "ORE_C",
            "S1": "ORE_M",
            "T1": "ORE_ST_NOT",
            "U1": "ORE_ST_DIU",
            "V1": "ORE_FEST_NOT",
            "W1": "ORE_FEST_DIU",
        }
        for cell_coord, val in headers.items():
            ws[cell_coord].value = val

        cols = ["AC", "Z", "X", "L", "I", "H", "G", "F", "E", "D", "A"]
        indices = sorted([column_index_from_string(c) for c in cols], reverse=True)
        for idx in indices:
            ws.delete_cols(idx)

            ws.delete_cols(idx)


class SaveWorkbookStep(ProcessingStep):
    """Salva il file Excel elaborato nella directory di destinazione."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue il salvataggio del file Excel."""
        wb = context["wb"]
        dest_dir = context["dest_dir"]
        odc = context["odc"]

        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_path = dest_dir / f"{odc}_TS.xlsx"

        if dest_path.exists():
            dest_path = dest_dir / f"{odc}_TS_{time.strftime('%Y%m%d-%H%M%S')}.xlsx"

        wb.save(dest_path)
        context["dest_path"] = dest_path
        wb.close()


class CleanupStep(ProcessingStep):
    """Rimuove il file sorgente temporaneo se diverso dal file di destinazione."""

    def execute(self, context: dict[str, Any]) -> None:
        """Esegue la pulizia dei file temporanei."""
        src = context["file_path"]
        dest = context["dest_path"]
        with suppress(Exception):
            if src.resolve() != dest.resolve():
                src.unlink()
