from __future__ import annotations

import io
import json
import logging
import os
import re
import warnings
import zipfile
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
from src.core.schemas import validate_contabilita, validate_giornaliere

# Lazy import placeholder
_pd = None

# Tentativo di importare msoffcrypto
try:
    import msoffcrypto  # type: ignore
except ImportError:
    msoffcrypto = None

# Tentativo di importare openpyxl
try:
    import openpyxl  # type: ignore

    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore
    HAS_OPENPYXL = False


class ExcelImporter:
    """Gestore per l'importazione di dati da file Excel."""

    @staticmethod
    def _get_pd():
        """Lazy load di pandas"""
        global _pd
        if _pd is None:
            import pandas as _pd
        return _pd

    # Mapping colonne Excel -> DB (Contabilità / Dati)
    COLUMNS_MAPPING = {
        "DATA PREV.": "data_prev",
        "MESE": "mese",
        "N° PREV.": "n_prev",
        "TOTALE PREV.": "totale_prev",
        "ATTIVITA'": "attivita",
        "TCL": "tcl",
        "ODC": "odc",
        "STATO ATTIVITA'": "stato_attivita",
        "TIPOLOGIA": "tipologia",
        "ORE SP": "ore_sp",
        "RESA": "resa",
        "ANNOTAZIONI": "annotazioni",
        "INDIRIZZO CONSUNTIVO": "indirizzo_consuntivo",
        "NOME FILE": "nome_file",
    }

    # Mapping colonne Excel -> DB (Giornaliere)
    GIORNALIERE_MAPPING = {
        "DATA": "data",
        "PERSONALE": "personale",
        "DESCRIZIONE ATTIVITA'": "descrizione",
        "TCL": "tcl",
        "ODC": "odc",
        "N° PDL": "pdl",
        "INIZIO": "inizio",
        "FINE": "fine",
        "ORE": "ore",
        "consuntivo": "n_prev",  # Rinominato come richiesto
    }

    # Mapping Scarico Ore Cantiere
    SCARICO_ORE_COLS = [
        "data",
        "pers1",
        "pers2",
        "odc",
        "pos",
        "dalle",
        "alle",
        "totale_ore",
        "descrizione",
        "finito",
        "commessa",
        "styles",
    ]

    # Mapping Attività Programmate
    ATTIVITA_PROGRAMMATE_MAPPING = {
        "PS": "ps",
        "AREA": "area",
        "PdL": "pdl",
        "IMP.": "imp",
        "DESCRIZIONE\nATTIVITA'": "descrizione",
        "LUN": "lun",
        "MAR": "mar",
        "MER": "mer",
        "GIO": "gio",
        "VEN": "ven",
        "STATO\nPdL": "stato_pdl",
        "STATO\nATTIVITA'": "stato_attivita",
        "DATA\nCONTROLLO": "data_controllo",
        "PERSONALE\nIMPIEGATO": "personale",
        "PO": "po",
        "AVVISO": "avviso",
    }

    ATTIVITA_PROGRAMMATE_COLS = list(ATTIVITA_PROGRAMMATE_MAPPING.values()) + [
        "styles"
    ]  # Added styles

    # Mapping Certificati Campione
    CERTIFICATI_CAMPIONE_MAPPING = {
        "Modello / Tipo": "modello",
        "Costruttore": "costruttore",
        "Matricola": "matricola",
        "Range Strumento": "range_strumento",
        "Errore max %": "errore_max",
        "Certificato Taratura": "certificato",
        "Scadenza Certificato": "scadenza",
        "Emissione Certificato": "emissione",
        "ID-COEMI": "id_coemi",
        "Stato Certificato": "stato",
    }

    CERTIFICATI_CAMPIONE_COLS = list(CERTIFICATI_CAMPIONE_MAPPING.values())

    # Mapping Storico OdA
    STORICO_ODA_MAPPING = {
        "Org. Acq.": "org_acq",
        "Data OdA": "data_oda",
        "OdA": "oda",
        "Pos OdA": "pos_oda",
        "Stato": "stato",
        "Cat. Contab.": "cat_contab",
        "Descrizione": "descrizione",
        "Qta": "qta",
        "UOM": "uom",
        "Data Consegna": "data_consegna",
        "Valore Netto Pos. ODA": "valore_netto_pos",
        "Valore Residuo ODA": "valore_residuo",
        "Valore Netto ODA": "valore_netto_oda",
        "Divisione": "divisione",
        "Destinatario": "destinatario",
        "Nome Destinatario": "nome_destinatario",
        "Codice Fornitore": "codice_fornitore",
        "Descrizione Fornitore": "descrizione_fornitore",
        "Emittente Fattura": "emittente_fattura",
        "Descrizione Emittente Fattura": "desc_emittente_fattura",
        "Contract Card": "contract_card",
        "Contratto": "contratto",
        "Posizione Contratto": "posizione_contratto",
        "Gruppo Acquisti": "gruppo_acquisti",
        "Indicatore Rilascio": "indicatore_rilascio",
        "Stato Rilascio": "stato_rilascio",
        "Attività": "attivita",  # Matches Attivit via robust check
        "Num riga": "num_riga",
        "Quantità": "quantita",  # Matches Quantit
        "Unità di Mis": "unita_mis",  # Matches Unit di Mis
        "Prezzo lordo": "prezzo_lordo",
        "Testo breve": "testo_breve",
    }

    STORICO_ODA_COLS = list(STORICO_ODA_MAPPING.values())

    @staticmethod
    def _decrypt_if_encrypted(file_path: Path) -> Tuple[Any, bool]:
        """Tenta di decifrare un file Excel se protetto da password."""
        if msoffcrypto:
            try:
                from src.core import config_manager

                config = config_manager.load_config()
                # Recupera password da config, default "coemi"
                pwd = config.get("excel_decryption_password", "coemi")

                with Path(file_path).open("rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=pwd)
                    temp_decrypted = io.BytesIO()
                    office_file.decrypt(temp_decrypted)
                    temp_decrypted.seek(0)
                    return temp_decrypted, True
            except Exception:
                # Non cifrato o errore msoffcrypto, procediamo col file originale
                pass
        return file_path, False

    @classmethod
    def _get_excel_file(cls, file_obj) -> pd.ExcelFile:
        """Tenta di aprire il file Excel con motore ottimizzato (calamine > default > openpyxl)."""
        pd = cls._get_pd()
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # 1. Tentativo con Calamine (Rust-based, ultra veloce)
            try:
                return pd.ExcelFile(file_obj, engine="calamine")
            except (ImportError, ValueError, Exception):
                pass

            # 2. Tentativo Standard (Pandas auto-detect)
            try:
                return pd.ExcelFile(file_obj)
            except Exception:
                pass

            # 3. Fallback esplicito OpenPyXL
            return pd.ExcelFile(file_obj, engine="openpyxl")

    @classmethod
    def _identify_sheet_year(cls, sheet_name: str) -> Optional[int]:
        """Estrae l'anno dal nome del foglio o usa l'anno corrente per nomi specifici."""
        match = re.search(r"(\d{4})", sheet_name)
        if match:
            year = int(match.group(1))
            return year if 2000 <= year <= 2100 else None

        if sheet_name.lower() in ("dati", "preventivi", "riepilogo"):
            return datetime.now().year
        return None

    @classmethod
    def _find_header_row(cls, xls, sheet_name) -> int:
        """Cerca l'indice della riga di intestazione basandosi su colonne chiave."""
        pd = cls._get_pd()
        preview_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=15)
        key_cols_norm = ["DATAPREV", "MESE", "NPREV", "TOTALEPREV", "ATTIVITA", "ODC"]

        for i_raw, row in preview_df.iterrows():
            row_norm = []
            for val in row.values:
                s = str(val).strip().upper()
                s = s.replace(" ", "").replace(".", "").replace("°", "")
                row_norm.append(s)

            matches = sum(1 for k in key_cols_norm if k in row_norm)
            if matches >= 2:
                # Debug print removed, return the index
                return int(i_raw)
        return 0  # Fallback to first row

    @classmethod
    def _normalize_columns(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Rinomina le colonne del DataFrame usando la mappa interna."""
        normalized_map = {}
        for k, v in cls.COLUMNS_MAPPING.items():
            norm_k = k.upper().replace(" ", "").replace(".", "").replace("°", "")
            normalized_map[norm_k] = v

        rename_map = {}
        for col in df.columns:
            col_str = str(col).strip().upper()
            norm_col = col_str.replace(" ", "").replace(".", "").replace("°", "")

            if norm_col in normalized_map:
                rename_map[col] = normalized_map[norm_col]
            else:
                # Euristiche extra
                if "PREV" in norm_col and "DATA" in norm_col:
                    rename_map[col] = "data_prev"
        df.rename(columns=rename_map, inplace=True)

        # Validazione Pandera (Contabilità)
        try:
            df = validate_contabilita(df)
        except Exception as e:
            logging.warning(
                f"Validazione Pandera Contabilità fallita (uso fallback): {e})"
            )

        return df

    @classmethod
    def import_contabilita_dati(
        cls,
        file_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, list, list]:
        """
        Importa i dati dal file Excel specificato (Tabella Dati).
        """
        path = Path(file_path)
        if not path.exists():
            return False, f"File non trovato: {file_path}", [], []

        try:
            file_obj, _ = cls._decrypt_if_encrypted(path)
            xls = cls._get_excel_file(file_obj)

            valid_sheets = [
                str(s) for s in xls.sheet_names if cls._identify_sheet_year(str(s))
            ]
            if not valid_sheets:
                return (
                    False,
                    "Nessun anno importato (Controlla nomi fogli: YYYY o 'Dati/Preventivi').",
                    [],
                    [],
                )

            all_rows, imported_years = cls._process_all_sheets(
                xls, valid_sheets, progress_callback
            )

            if not imported_years:
                return (
                    False,
                    "Fogli validi trovati ma nessun dato importato (fogli vuoti?).",
                    [],
                    [],
                )

            return (
                True,
                f"Anni importati: {sorted(set(imported_years))}",
                all_rows,
                list(set(imported_years)),
            )

        except Exception as e:
            logging.error(f"Errore importazione Excel: {e}")
            return False, f"Errore critico importazione: {e}", [], []

    @classmethod
    def scan_scarico_ore_rows(cls, file_path: str) -> int:
        """
        Stima rapida delle righe per Scarico Ore (DataEase) per calcolo ETA.
        Supporta file criptati (password 'coemi').
        """
        path = Path(file_path)
        if not path.exists():
            return 0

        def _scan_zip(zip_file_obj):
            try:
                cnt = 0
                with zipfile.ZipFile(zip_file_obj, "r") as z:
                    for name in z.namelist():
                        if name.startswith("xl/worksheets/sheet"):
                            with z.open(name) as f:
                                # Buffer increased to 32KB for safety
                                head = f.read(32768).decode("utf-8", errors="ignore")
                                match = re.search(
                                    r'<dimension ref="[A-Z]+[0-9]+:[A-Z]+(\d+)"', head
                                )
                                if match:
                                    r = int(match.group(1))
                                    if r > cnt:
                                        cnt = r
                return cnt
            except Exception:
                return 0

        # 1. Tentativo Diretto (File non criptato)
        try:
            res = _scan_zip(path)
            if res > 0:
                return res
        except zipfile.BadZipFile:
            pass
        except Exception:
            pass

        # 3. Tentativo Decrittazione
        if msoffcrypto:
            try:
                decrypted = io.BytesIO()
                with open(path, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password="coemi")
                    office_file.decrypt(decrypted)
                decrypted.seek(0)
                return _scan_zip(decrypted)
            except Exception:
                return 0

        return 0

    @classmethod
    def _process_all_sheets(
        cls, xls, sheet_names: List[str], progress_callback: Optional[Callable]
    ) -> Tuple[List[Tuple], List[int]]:
        """Cicla sui fogli e aggrega i risultati."""
        all_rows = []
        imported_years = []
        total_sheets = len(sheet_names)

        for i, sheet_name in enumerate(sheet_names):
            year = cls._identify_sheet_year(sheet_name)
            if not year:
                continue

            rows = cls._process_single_sheet(xls, sheet_name, year)
            if rows:
                all_rows.extend(rows)
                imported_years.append(year)

            if progress_callback:
                progress_callback(i + 1, total_sheets)

        return all_rows, imported_years

    @classmethod
    def _process_single_sheet(cls, xls, sheet_name: str, year: int) -> List[Tuple]:
        """Processa un singolo foglio del file Excel di contabilità."""
        try:
            pd = cls._get_pd()
            header_row_idx = cls._find_header_row(xls, sheet_name)
            # Optimization: Limit parsing to first 52 columns (A:AZ) to avoid 'ghost columns' slowness
            try:
                df = pd.read_excel(
                    xls, sheet_name=sheet_name, header=header_row_idx, usecols="A:AZ"
                )
            except Exception:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_idx)

            df.columns = [str(c).strip().upper() for c in df.columns]

            if not df.empty:
                df = df.iloc[:-1]  # Rimuovi riga dei totali solitamente presente
            df.dropna(how="all", inplace=True)

            if df.empty:
                return []

            df["year"] = year
            df = cls._normalize_columns(df)
            df = cls._ensure_required_columns(df)

            # Preparazione finale dati
            target_columns = ["year"] + list(cls.COLUMNS_MAPPING.values())
            df = df[target_columns].copy()

            # --- Gestione Tipi Intelligente ---
            for col in df.columns:
                if col == "year":
                    continue

                # 1. Tenta conversione numerica per colonne che dovrebbero essere numeri
                if col in ["totale_prev", "ore_sp"]:
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                    # Arrotonda a 2 decimali per eliminare rumore (es. .00000000001)
                    df[col] = df[col].round(2)

                # 2. Gestione Date
                elif col == "data_prev":
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime(
                        "%Y-%m-%d"
                    )

                # 3. Gestione colonna RESA (mista: numeri o stringhe)
                elif col == "resa":
                    # Converti numeri quando possibile, mantieni stringhe altrimenti
                    df[col] = df[col].apply(
                        lambda x: (
                            str(round(float(x), 2))
                            if pd.notna(x)
                            and str(x)
                            .replace(".", "")
                            .replace(",", "")
                            .replace("-", "")
                            .isdigit()
                            else str(x).strip()
                            if pd.notna(x)
                            else ""
                        )
                    )

                # 4. Pulizia stringhe per il resto
                else:
                    df[col] = (
                        df[col]
                        .astype(str)
                        .str.strip()
                        .replace(r"(?i)^nan$", "", regex=True)
                    )

            # Riempie i NaN rimasti con valori sicuri per il DB
            df = df.fillna("")

            return list(df.itertuples(index=False, name=None))
        except Exception as e:
            logging.warning(f"Errore processamento foglio {sheet_name}: {e}")
            return []

    @classmethod
    def _ensure_required_columns(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Assicura che tutte le colonne mappate esistano nel DataFrame."""
        for db_col in cls.COLUMNS_MAPPING.values():
            if db_col not in df.columns:
                df[db_col] = ""
        return df

    @classmethod
    def _process_single_giornaliera(
        cls, args: Tuple[int, Path, Dict]
    ) -> Tuple[int, List[Tuple], Optional[str]]:
        """Helper per processare un singolo file giornaliera in parallelo."""
        year, file_path, lookup_map = args
        try:
            # Decrittografia se necessaria
            file_obj, _ = cls._decrypt_if_encrypted(file_path)
            df = cls._read_giornaliera_sheet(file_obj)
            if df is None:
                return (year, [], None)

            # Normalizzazione Colonne
            df = cls._normalize_giornaliera_columns(df)
            if df is None:
                return (year, [], None)

            # Pulizia Base
            df = cls._clean_giornaliera_data(df)
            if df.empty:
                return (year, [], None)

            # Logica di Business: Arricchimento ODC
            cls._enrich_giornaliera_odc(df, lookup_map)

            # Preparazione finale
            df["year"] = year
            df["nome_file"] = file_path.name

            target_cols = [
                "year",
                "data",
                "personale",
                "descrizione",
                "tcl",
                "odc",
                "pdl",
                "inizio",
                "fine",
                "ore",
                "n_prev",
                "nome_file",
            ]
            rows = list(df[target_cols].itertuples(index=False, name=None))
            return (year, rows, None)

        except Exception as e:
            return (year, [], str(e))

    @classmethod
    def _read_giornaliera_sheet(cls, file_path: Path) -> Optional[pd.DataFrame]:
        """Legge il foglio RIASSUNTO gestendo i motori pandas."""
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            pd = cls._get_pd()
            try:
                return pd.read_excel(file_path, sheet_name="RIASSUNTO")
            except ValueError:
                return None
            except zipfile.BadZipFile:
                # File non valido o corrotto
                return None
            except Exception:
                try:
                    return pd.read_excel(
                        file_path, sheet_name="RIASSUNTO", engine="openpyxl"
                    )
                except zipfile.BadZipFile:
                    return None
                except Exception as e:
                    raise e

    @classmethod
    def _normalize_giornaliera_columns(cls, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Applica il mapping delle colonne specifico per le giornaliere."""
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {}

        # Costruisci il mapping con euristica per match fuzzy
        for excel_col, db_col in cls.GIORNALIERE_MAPPING.items():
            if excel_col in df.columns:
                rename_map[excel_col] = db_col
            else:
                # Match case-insensitive
                for col in df.columns:
                    if col.upper() == excel_col.upper():
                        rename_map[col] = db_col
                        break

        if not rename_map:
            return None

        df.rename(columns=rename_map, inplace=True)

        # Validazione Pandera (Giornaliere)
        try:
            df = validate_giornaliere(df)
        except Exception as e:
            logging.warning(
                f"Validazione Pandera Giornaliere fallita (uso fallback): {e}"
            )

        return df

    @classmethod
    def _clean_giornaliera_data(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Esegue pulizia righe, totali e nan."""
        if df.empty:
            return df

        # Rimuovi ultima riga (spesso è una riga di totali)
        if len(df) > 0:
            df = df.iloc[:-1]

        if df.empty:
            return df

        # Filtra righe con "Totale" in qualsiasi colonna stringa
        for col in df.columns:
            if df[col].dtype == "object":  # Colonne testuali
                mask = df[col].astype(str).str.contains("Totale", na=False, case=False)
                df = df[~mask]

        if df.empty:
            return df

        # Rimuovi righe con campi critici vuoti (data, personale, ore devono essere presenti)
        critical_cols = []
        if "data" in df.columns:
            critical_cols.append("data")
        if "personale" in df.columns:
            critical_cols.append("personale")
        if "ore" in df.columns:
            critical_cols.append("ore")

        if critical_cols:
            df.dropna(subset=critical_cols, how="any", inplace=True)

        if df.empty:
            return df

        # Assicura colonne esistenti e pulisci stringhe
        for db_col in cls.GIORNALIERE_MAPPING.values():
            if db_col not in df.columns:
                df[db_col] = ""

        cols_to_clean = [
            "odc",
            "n_prev",
            "data",
            "personale",
            "descrizione",
            "tcl",
            "pdl",
            "inizio",
            "fine",
            "ore",
        ]
        df[cols_to_clean] = df[cols_to_clean].astype(str).apply(lambda x: x.str.strip())
        df[cols_to_clean] = df[cols_to_clean].replace(r"(?i)^nan$", "", regex=True)
        return df

    @classmethod
    def _enrich_giornaliera_odc(cls, df: pd.DataFrame, lookup_map: Dict):
        """Estrae o associa l'ODC mancante da varie fonti."""
        # 1. Da Lookup Map (n_prev -> odc)
        mask_empty = df["odc"] == ""
        if mask_empty.any() and lookup_map:
            mapped = df.loc[mask_empty, "n_prev"].map(lookup_map)
            df.loc[mask_empty, "odc"] = mapped.fillna("")

        # 2. Da Descrizione (Pattern XX/XXX)
        mask_empty = df["odc"] == ""
        if mask_empty.any():
            comm_pattern = r"\b(\d{2}/\d{3})\b"
            extracted = df.loc[mask_empty, "descrizione"].str.extract(
                comm_pattern, expand=False
            )
            df.loc[mask_empty, "odc"] = extracted.fillna("")

        # 3. Normalizzazione standard 5400...
        mask_standard = ~df["odc"].str.contains("canone", case=False, na=False) & ~df[
            "odc"
        ].str.match(r"^\d{2}/\d{3}$", na=False)
        if mask_standard.any():
            extracted = df.loc[mask_standard, "odc"].str.extract(
                r"(5400\d+)", expand=False
            )
            df.loc[mask_standard, "odc"] = extracted.fillna("")

    @classmethod
    def import_giornaliere(
        cls,
        root_path: str,
        lookup_map: Dict,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, List[Tuple], List[int]]:
        """
        Importa i dati dalle cartelle Giornaliere, processa i file in parallelo.
        """
        root = Path(root_path)
        if not root.exists():
            return False, "Directory Giornaliere non trovata.", [], []

        tasks_args = cls._collect_giornaliere_tasks(root, lookup_map)
        if not tasks_args:
            return (
                True,
                f"Nessuna nuova giornaliera trovata (check anno >= {datetime.now().year}).",
                [],
                [],
            )

        all_rows, imported_years = cls._run_parallel_import(
            tasks_args, progress_callback
        )

        if not imported_years:
            return True, "Nessuna riga valida importata dai file trovati.", [], []

        return (
            True,
            f"Importate Giornaliere: {sorted(imported_years)}",
            all_rows,
            imported_years,
        )

    @classmethod
    def _collect_giornaliere_tasks(cls, root: Path, lookup_map: Dict) -> List[Tuple]:
        """Scansiona le cartelle e crea la lista dei file da processare."""
        tasks = []
        current_year = datetime.now().year
        for folder in root.iterdir():
            if not folder.is_dir():
                continue
            match = re.match(r"Giornaliere\s+(\d{4})", folder.name, re.IGNORECASE)
            if not match:
                continue

            year = int(match.group(1))
            if year > current_year:
                continue

            for file_path in folder.glob("*.xls*"):
                if not file_path.name.startswith("~$"):
                    tasks.append((year, file_path, lookup_map))
        return tasks

    @classmethod
    def _run_parallel_import(
        cls, tasks: List[Tuple], progress_callback: Optional[Callable]
    ) -> Tuple[List[Tuple], List[int]]:
        """Esegue il processamento parallelo tramite ProcessPoolExecutor."""
        all_rows = []
        years_encountered = set()
        total_tasks = len(tasks)
        processed_count = 0

        # Limit max_workers to avoid memory spikes
        max_workers = min(4, (os.cpu_count() or 1))
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            for result in executor.map(cls._process_single_giornaliera, tasks):
                processed_count += 1
                if progress_callback:
                    progress_callback(processed_count, total_tasks)

                r_year, r_rows, r_err = result
                if r_rows:
                    all_rows.extend(r_rows)
                    years_encountered.add(r_year)
                if r_err:
                    logging.error(f"Errore lettura file (Year {r_year}): {r_err}")

        return all_rows, list(years_encountered)

    @classmethod
    def import_attivita_programmate(
        cls,
        file_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, List[Tuple]]:
        """Importa il file Attività Programmate (veloce, senza colori)."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File Attività Programmate non trovato: {file_path}", []

        try:
            df = cls._read_attivita_programmate_sheet(path)
            if df is None:
                return False, "Foglio 'Riepilogo' non trovato o file illeggibile.", []

            # Normalizzazione Colonne
            df = cls._normalize_attivita_columns(df)
            if df is None:
                return False, "Colonne non trovate. Controlla intestazione riga 3.", []

            # Pulizia e Preparazione
            rows = cls._prepare_attivita_rows(df)
            return True, f"Importate {len(rows)} righe in Attività Programmate.", rows

        except Exception as e:
            return False, f"Errore importazione Attività Programmate: {e}", []

    @classmethod
    def _read_attivita_programmate_sheet(cls, path: Path) -> Optional[pd.DataFrame]:
        """Tenta di leggere il foglio 'Riepilogo'."""
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            pd = cls._get_pd()
            try:
                return pd.read_excel(path, sheet_name="Riepilogo", header=2)
            except (ValueError, Exception):
                try:
                    return pd.read_excel(
                        path, sheet_name="Riepilogo", header=2, engine="openpyxl"
                    )
                except Exception:
                    return None

    @classmethod
    def _normalize_attivita_columns(cls, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Applica il mapping delle colonne e l'euristica per i newline."""
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {}

        for excel_col, db_col in cls.ATTIVITA_PROGRAMMATE_MAPPING.items():
            if excel_col in df.columns:
                rename_map[excel_col] = db_col
            else:
                # Euristiche per newline (es. "STATO\nPdL" -> "STATO PdL")
                for col in df.columns:
                    if (
                        excel_col.replace("\n", " ").strip()
                        == col.replace("\n", " ").strip()
                    ):
                        rename_map[col] = db_col
                        break

        if not rename_map:
            return None

        df.rename(columns=rename_map, inplace=True)
        return df

    @classmethod
    def _prepare_attivita_rows(cls, df: pd.DataFrame) -> List[Tuple]:
        """Pulisce i dati e restituisce le tuple per il database."""
        # 1. Assicura colonne esistenti
        for db_col in cls.ATTIVITA_PROGRAMMATE_MAPPING.values():
            if db_col not in df.columns:
                df[db_col] = ""

        # 2. Drop righe vuote
        check_cols = [c for c in ["ps", "area", "descrizione"] if c in df.columns]
        if check_cols:
            df.dropna(how="all", subset=check_cols, inplace=True)

        # 3. Formattazione stringhe
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        df["styles"] = ""

        # 4. Selezione colonne finale
        db_cols = list(cls.ATTIVITA_PROGRAMMATE_MAPPING.values()) + ["styles"]
        df = df[db_cols]

        return list(df.itertuples(index=False, name=None))

    @classmethod
    def import_scarico_ore(
        cls,
        file_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, List[Tuple]]:
        """Importa il file Scarico Ore Cantiere (OpenPyXL per colori + Diff Logic)."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File Scarico Ore non trovato: {file_path}", []

        if not openpyxl:
            return False, "Modulo 'openpyxl' mancante.", []

        try:
            wb_data = cls._load_scarico_workbook(path)
            if "SCARICO ORE" not in wb_data.sheetnames:
                return False, "Foglio 'SCARICO ORE' non trovato.", []

            ws_data = wb_data["SCARICO ORE"]
            rows = cls._process_all_scarico_rows(ws_data, progress_callback)

            return True, f"Importate {len(rows)} righe da Scarico Ore.", rows

        except Exception as e:
            return False, f"Errore importazione Scarico Ore: {e}", []

    @classmethod
    def _load_scarico_workbook(cls, path: Path) -> Any:
        """Carica il workbook gestendo l'eventuale crittografia."""
        wb_file = io.BytesIO()
        is_encrypted = False

        if msoffcrypto:
            try:
                with open(path, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password="coemi")
                    office_file.decrypt(wb_file)
                    is_encrypted = True
            except Exception:
                pass

        if not is_encrypted:
            with open(path, "rb") as f:
                wb_file.write(f.read())

        wb_file.seek(0)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            # Performance optimizations: read_only=True is CRITICAL for speed on 24MB files
            # Note: Styles ARE available in read_only mode (verified by test).
            return openpyxl.load_workbook(
                wb_file,
                data_only=True,
                read_only=True,
                keep_vba=False,
                keep_links=False,
            )

    @classmethod
    def _process_all_scarico_rows(
        cls, ws, progress_callback: Optional[Callable]
    ) -> List[Tuple]:
        """
        Cicla sulle righe del foglio scarico ore.
        ULTRA-OTTIMIZZATO per performance massime su file grandi (130k+ righe).
        """
        start_row = 6
        col_keys = [
            "data",
            "pers1",
            "pers2",
            "odc",
            "pos",
            "dalle",
            "alle",
            "totale_ore",
            "descrizione",
            "finito",
            "commessa",
        ]
        total_rows = ws.max_row

        # Pre-allocate list for optimal performance
        rows_to_insert: List[Tuple] = []
        rows_to_insert_append = rows_to_insert.append  # Cache method lookup

        # Process all rows with minimal overhead
        progress_interval = 5000  # Update progress even less frequently
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, min_col=2, max_col=12, values_only=False),
            start=start_row,
        ):
            # Update progress very infrequently to minimize overhead
            if progress_callback and row_idx % progress_interval == 0:
                progress_callback(row_idx, total_rows)

            # Process row with optimized logic
            db_row = cls._process_scarico_ore_row(row, col_keys)
            if db_row:
                rows_to_insert_append(db_row)  # Use cached method

        return rows_to_insert

    @classmethod
    def _process_scarico_ore_row(cls, row, col_keys) -> Optional[Tuple]:
        """
        Processa una singola riga estraendo valori e stili.
        OTTIMIZZATO per minimizzare allocazioni e chiamate costose.
        """
        # Indici colonne (fissi per scarico ore)
        # 0: data, 1: pers1, 2: pers2, 3: odc, 4: pos,
        # 5: dalle, 6: alle, 7: tot_ore, 8: desc, 9: finito, 10: comm

        # 1. Fast empty check: se data, pers1, odc, e pos sono vuoti, skip
        # Accesso diretto alle celle raw
        (
            c_data,
            c_p1,
            c_p2,
            c_odc,
            c_pos,
            c_dalle,
            c_alle,
            c_tot,
            c_desc,
            c_fin,
            c_comm,
        ) = row[0:11]

        v_odc = c_odc.value
        v_pos = c_pos.value

        # ODC e POS sono obbligatori
        if v_odc is None and v_pos is None:
            return None

        # 2. Extract values (Inline formatting)
        def _fmt(val):
            if val is None:
                return ""
            s = str(val).strip()
            return s.replace("\n", " ") if s else ""

        # ODC specific handling
        vals = []
        # Data - Strip time if present
        v_data = c_data.value
        s_data = ""
        if v_data:
            if hasattr(v_data, "strftime"):
                s_data = v_data.strftime("%Y-%m-%d")
            else:
                # String cleanup: "2024-01-01 00:00:00" -> "2024-01-01"
                s = str(v_data).strip()
                if " " in s:
                    s = s.split(" ")[0]
                s_data = s
        vals.append(s_data)
        # Pers1
        vals.append(_fmt(c_p1.value))
        # Pers2
        vals.append(_fmt(c_p2.value))

        # ODC (skip 0)
        s_odc = _fmt(v_odc)
        if s_odc == "0" or s_odc == "0.0":
            s_odc = ""
        vals.append(s_odc)

        # POS (skip 0)
        s_pos = _fmt(v_pos)
        if s_pos == "0" or s_pos == "0.0":
            s_pos = ""
        vals.append(s_pos)

        # Dalle
        vals.append(_fmt(c_dalle.value))
        # Alle
        vals.append(_fmt(c_alle.value))

        # Totale Ore
        s_tot = _fmt(c_tot.value)
        vals.append(s_tot)

        # Desc
        vals.append(_fmt(c_desc.value))
        # Finito
        vals.append(_fmt(c_fin.value))

        # Commessa (skip 0)
        v_comm = c_comm.value
        s_comm = _fmt(v_comm)
        if s_comm == "0" or s_comm == "0.0":
            s_comm = ""
        vals.append(s_comm)

        # 3. Validation Logic (Inlined)
        # Check: odc, pos, totale_ore must be present
        if not s_odc or not s_pos or not s_tot:
            return None

        # Check: at least one person
        if not vals[1] and not vals[2]:  # pers1 and pers2
            return None

        # 4. Extract styles (Inline)
        row_styles: Dict[str, Dict[str, str]] = {}
        for i, key in enumerate(col_keys):
            # Skip empty cells
            if vals[i] == "":
                continue

            cell = row[i]
            # Foreground
            try:
                font = cell.font
                if font and font.color and font.color.type == "rgb":
                    rgb = str(font.color.rgb)
                    # OpenPyXL RGB is ARGB usually
                    hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"
                    if hex_code != "#000000":  # Skip default black (optimization)
                        row_styles.setdefault(key, {})["fg"] = hex_code
            except (AttributeError, TypeError):
                pass

            # Background
            try:
                fill = cell.fill
                if fill and fill.patternType == "solid":
                    start_color = fill.start_color
                    if start_color and start_color.type == "rgb":
                        rgb = str(start_color.rgb)
                        hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"
                        if (
                            hex_code != "#000000" and hex_code != "#FFFFFF"
                        ):  # Skip white/black bg
                            row_styles.setdefault(key, {})["bg"] = hex_code
            except (AttributeError, TypeError):
                pass

        # Return tuple
        return (
            vals[0],  # data
            vals[1],  # pers1
            vals[2],  # pers2
            vals[3],  # odc
            vals[4],  # pos
            vals[5],  # dalle
            vals[6],  # alle
            vals[7],  # totale_ore
            vals[8],  # descrizione
            vals[9],  # finito
            vals[10],  # commessa
            json.dumps(row_styles) if row_styles else "",
        )

    @classmethod
    def import_certificati_campione(
        cls,
        file_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, List[Tuple]]:
        """Importa il file Certificati Campione e restituisce le righe."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File non trovato: {file_path}", []

        try:
            pd = cls._get_pd()
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                xls = pd.ExcelFile(path)
                sheet_name = cls._find_certificati_sheet(xls)
                if not sheet_name:
                    return False, "Nessun foglio trovato.", []

                df, header_idx = cls._read_certificati_data(path, sheet_name)
                if df.empty:
                    return False, "Foglio vuoto.", []

                return cls._process_certificati_df(df, sheet_name, header_idx)
        except Exception as e:
            return False, f"Errore importazione Certificati: {e}", []

    @classmethod
    def import_storico_oda(
        cls,
        file_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Tuple[bool, str, List[Tuple]]:
        """Importa il file Storico OdA."""
        path = Path(file_path)
        if not path.exists():
            return False, f"File non trovato: {file_path}", []

        try:
            # 1. Lettura Excel
            df = cls._read_storico_oda_excel(path)
            if df.empty:
                return False, "Foglio vuoto.", []

            # 2. Mappatura Colonne
            rename_map = cls._map_storico_oda_columns(df)
            if not rename_map:
                return False, "Nessuna colonna riconosciuta.", []

            df.rename(columns=rename_map, inplace=True)
            df = df.loc[:, ~df.columns.duplicated()]

            # 3. Normalizzazione e Pulizia
            cls._normalize_storico_oda_df(df)
            cls._clean_storico_oda_data(df)

            # 4. Conversione in tuple
            data = [tuple(x) for x in df.to_numpy()]
            return True, f"Trovate {len(data)} righe.", data

        except Exception as e:
            return False, f"Errore importazione Storico OdA: {e}", []

    @classmethod
    def _read_storico_oda_excel(cls, path: Path) -> pd.DataFrame:
        """Legge il file excel tentando diversi fogli."""
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                return pd.read_excel(path, sheet_name="Formato PF")
            except ValueError:
                return pd.read_excel(path, sheet_name=0)

    @classmethod
    def _map_storico_oda_columns(cls, df: pd.DataFrame) -> dict:
        """Mappa le colonne dell'Excel a quelle del DB."""
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {}
        for excel_col, db_col in cls.STORICO_ODA_MAPPING.items():
            for col in df.columns:
                if excel_col == col:
                    rename_map[col] = db_col
                    break
                # Fallback matching
                if len(excel_col) >= 4 and col.startswith(excel_col[:4]):
                    if excel_col.startswith("Unit") and col.startswith("Unit"):
                        rename_map[col] = db_col
                        break
                    if abs(len(col) - len(excel_col)) <= 2:
                        rename_map[col] = db_col
                        break
        return rename_map

    @classmethod
    def _normalize_storico_oda_df(cls, df: pd.DataFrame):
        """Assicura che tutte le colonne richieste esistano e filtra solo quelle del DB."""
        for db_col in cls.STORICO_ODA_COLS:
            if db_col not in df.columns:
                df[db_col] = ""

        # Filtra solo le colonne del DB (in-place modification via drop)
        cols_to_drop = [c for c in df.columns if c not in cls.STORICO_ODA_COLS]
        if cols_to_drop:
            df.drop(columns=cols_to_drop, inplace=True)

    @classmethod
    def _clean_storico_oda_data(cls, df: pd.DataFrame):
        """Pulisce date, numeri e ID."""
        # Date
        for date_col in ["data_oda", "data_consegna"]:
            df[date_col] = (
                pd.to_datetime(df[date_col], errors="coerce")
                .dt.strftime("%Y-%m-%d")
                .fillna("")
            )

        # Numeri
        num_cols = [
            "qta",
            "valore_netto_pos",
            "valore_residuo",
            "valore_netto_oda",
            "quantita",
            "prezzo_lordo",
        ]
        for num_col in num_cols:
            df[num_col] = df[num_col].apply(cls._clean_euro_num)

        # IDs
        id_cols = [
            "oda",
            "pos_oda",
            "num_riga",
            "divisione",
            "destinatario",
            "contratto",
            "posizione_contratto",
        ]
        for str_col in id_cols:
            df[str_col] = (
                df[str_col]
                .fillna(0)
                .astype(str)
                .str.replace(r"\\.0$", "", regex=True)
                .str.strip()
            )

        # Altre stringhe
        for col in df.columns:
            if col not in num_cols + ["data_oda", "data_consegna"] + id_cols:
                df[col] = df[col].fillna("").astype(str).str.strip()

    @staticmethod
    def _clean_euro_num(x):
        """Helper for European numbers (1.234,56 -> 1234.56)."""
        import pandas as pd

        if pd.isna(x) or str(x).strip() == "":
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if "." in s and "," in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0

    @classmethod
    def _find_certificati_sheet(cls, xls: pd.ExcelFile) -> Optional[str]:
        """Trova il foglio corretto per i certificati."""
        for name in xls.sheet_names:
            n_low = str(name).lower()
            if "strumenti campione" in n_low or "isab sud" in n_low:
                return str(name)
        return str(xls.sheet_names[0]) if xls.sheet_names else None

    @classmethod
    def _read_certificati_data(
        cls, path: Path, sheet_name: str
    ) -> Tuple[pd.DataFrame, int]:
        """Legge i dati individuando l'intestazione."""
        pd = cls._get_pd()
        df_preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=20)
        header_idx = cls._detect_certificati_header(df_preview)
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx)
        return df, header_idx

    @classmethod
    def _detect_certificati_header(cls, df_preview: pd.DataFrame) -> int:
        """Detects the header row index for Certificati Campione."""
        header_row_idx = -1
        max_matches = 0
        target_columns = set(cls.CERTIFICATI_CAMPIONE_MAPPING.keys())

        for i_raw, row in df_preview.iterrows():
            i = int(str(i_raw))
            row_values = [str(val).strip() for val in row.values]
            matches = sum(1 for col in target_columns if col in row_values)

            if matches > max_matches:
                max_matches = matches
                header_row_idx = i

        if header_row_idx == -1 or max_matches < 3:
            header_row_idx = 5

        return header_row_idx

    @classmethod
    def _process_certificati_df(
        cls, df: pd.DataFrame, sheet_name: str, header_row_idx: int
    ) -> Tuple[bool, str, List[Tuple]]:
        """Processes the Certificati DataFrame and returns formatted rows."""
        df.columns = [str(c).strip() for c in df.columns]

        # 1. Mapping e Validazione Colonne
        rename_map = cls._build_certificati_rename_map(df.columns.tolist())
        if not rename_map:
            found_cols = ", ".join(list(df.columns)[:5]) + "..."
            return (
                False,
                f"Nessuna colonna valida trovata. Sheet: {sheet_name}, Row: {header_row_idx}. Trovate: {found_cols}",
                [],
            )

        df.rename(columns=rename_map, inplace=True)

        # 2. Preparazione Schema
        df = cls._normalize_certificati_schema(df)

        # 3. Formattazione Dati (Date e Stati)
        df = cls._apply_certificati_formatting(df)

        # 4. Pulizia Finale
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        rows = list(df.itertuples(index=False, name=None))

        return True, f"Importate {len(rows)} righe in Certificati Campione.", rows

    @classmethod
    def _build_certificati_rename_map(cls, columns: List[str]) -> Dict[str, str]:
        """Costruisce la mappa di rinomina colonne basata sul mapping definito."""
        rename_map = {}
        for col in columns:
            col_clean = str(col).strip()
            # Cerca match esatto o parziale nel mapping
            for schema_col, db_col in cls.CERTIFICATI_CAMPIONE_MAPPING.items():
                if schema_col.lower() == col_clean.lower():
                    rename_map[col] = db_col
                    break
                # Fallback: se la colonna contiene la stringa schema (es. "Data\nScadenza")
                if schema_col in col_clean:
                    rename_map[col] = db_col
        return rename_map

    @classmethod
    def _normalize_certificati_schema(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Assicura l'ordine e l'esistenza delle colonne richieste."""
        target_cols = list(cls.CERTIFICATI_CAMPIONE_MAPPING.values())
        for c in target_cols:
            if c not in df.columns:
                df[c] = ""
        df = df[target_cols]
        df.dropna(how="all", inplace=True)
        return df

    @classmethod
    def _apply_certificati_formatting(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Applica formattazione date e calcolo giorni scadenza."""
        pd = cls._get_pd()

        def format_date_it(val):
            if pd.isna(val) or val == "":
                return ""
            try:
                dt = pd.to_datetime(val)
                return dt.strftime("%d/%m/%Y")
            except Exception:
                return str(val)

        def format_stato(val):
            if pd.isna(val) or val == "":
                return ""
            try:
                num = float(val)
                days = int(round(num))
                if days > 0:
                    return f"Scade tra {days} giorni"
                elif days < 0:
                    return f"Scaduto da {abs(days)} giorni"
                return "Scade oggi"
            except (ValueError, TypeError):
                return str(val)

        df["scadenza"] = df["scadenza"].apply(format_date_it)
        df["emissione"] = df["emissione"].apply(format_date_it)
        if "stato" in df.columns:
            df["stato"] = df["stato"].apply(format_stato)
        return df

    @classmethod
    def scan_workload(cls, file_path: str, giornaliere_path: str) -> Tuple[int, int]:
        """Scansiona rapidamente il carico di lavoro (fogli e file) per stima ETA."""
        sheets = cls._scan_excel_sheets(file_path)
        files = cls._scan_giornaliere_files(giornaliere_path)
        return sheets, files

    @classmethod
    def _scan_excel_sheets(cls, file_path: str) -> int:
        """Conta i fogli validi nell'Excel principale."""
        p_file = Path(file_path)
        if not file_path or not p_file.exists():
            return 0
        try:
            # Check if valid zip before opening
            if not zipfile.is_zipfile(p_file):
                # Maybe old xls format (OLE)
                return 1

            with zipfile.ZipFile(p_file, "r") as z:
                if "xl/workbook.xml" not in z.namelist():
                    return 1
                wb_xml = z.read("xl/workbook.xml").decode("utf-8")
                sheet_names = re.findall(r'name="([^"]+)"', wb_xml)
                return len([s for s in sheet_names if re.search(r"(\d{4})", s)])
        except Exception as e:
            logging.debug(f"Scan excel sheets error: {e}")
            return 1

    @classmethod
    def _scan_giornaliere_files(cls, giornaliere_path: str) -> int:
        """Conta i file validi nelle cartelle giornaliere."""
        p_giorn = Path(giornaliere_path)
        if not giornaliere_path or not p_giorn.exists():
            return 0

        count = 0
        current_year = datetime.now().year
        for folder in p_giorn.iterdir():
            if not folder.is_dir():
                continue
            match = re.match(r"Giornaliere\s+(\d{4})", folder.name, re.IGNORECASE)
            if not match:
                continue

            year = int(match.group(1))
            if year < current_year:
                continue

            for file_path in folder.glob("*.xls*"):
                if not file_path.name.startswith("~$"):
                    count += 1
        return count
