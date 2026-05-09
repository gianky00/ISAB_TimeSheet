import io
import json
import logging
import re
import warnings
import zipfile
from collections.abc import Callable
from contextlib import suppress
from pathlib import Path
from typing import Any, ClassVar

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

try:
    import openpyxl

    openpyxl_mod: Any = openpyxl
    HAS_OPENPYXL = True
except ImportError:
    openpyxl_mod = None
    HAS_OPENPYXL = False

from src.core.importers.base import BaseImporter

logger = logging.getLogger(__name__)


class ScaricoOreImporter(BaseImporter):
    """Importer per lo Scarico Ore Cantiere (con OpenPyXL per stili)."""

    SCARICO_ORE_COLS: ClassVar[list[str]] = [
        "data",
        "pers1",
        "pers2",
        "odc",
        "pos",
        "dalle",
        "alle",
        "totale_ore",
        "descrizione",
        "finito",
        "commessa",
        "styles",
    ]

    @classmethod
    def scan_scarico_ore_rows(cls, file_path: str) -> int:
        """
        Esegue una scansione rapida del file Excel per stimare il numero di righe totali.
        Legge direttamente la struttura XML del file .xlsx per massima velocità.

        Args:
            file_path: Percorso del file Excel.

        Returns:
            int: Numero stimato di righe.
        """
        path = Path(file_path)
        if not path.exists():
            return 0

        def _scan_zip(zip_file_obj: Any) -> int:
            try:
                cnt = 0
                with zipfile.ZipFile(zip_file_obj, "r") as z:
                    for name in z.namelist():
                        if name.startswith("xl/worksheets/sheet"):
                            with z.open(name) as f:
                                head = f.read(32768).decode("utf-8", errors="ignore")
                                match = re.search(r'<dimension ref="[A-Z]+[0-9]+:[A-Z]+(\d+)"', head)
                                if match:
                                    r = int(match.group(1))
                                    cnt = max(cnt, r)
                return cnt  # noqa: TRY300
            except Exception:
                return 0

        try:
            res = _scan_zip(path)
            if res > 0:
                return res
        except (zipfile.BadZipFile, Exception) as e:
            logger.debug(f"Scan excel rows error: {e}")

        from src.core.constants import Business  # noqa: PLC0415

        if msoffcrypto:
            with suppress(Exception):
                decrypted = io.BytesIO()
                with path.open("rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=Business.DEFAULT_EXCEL_PASSWORD)  # nosec: B106
                    office_file.decrypt(decrypted)
                decrypted.seek(0)
                return _scan_zip(decrypted)

        return 0

    @classmethod
    def import_scarico_ore(
        cls,
        file_path: str,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> tuple[bool, str, list[tuple[Any, ...]]]:
        """
        Importa i dati dal file Excel dello Scarico Ore Cantiere.
        Gestisce la decrittazione automatica se necessario e l'estrazione di dati e stili.

        Args:
            file_path: Percorso del file da importare.
            progress_callback: Funzione opzionale per monitorare l'avanzamento.

        Returns:
            tuple: (successo, messaggio, lista di righe processate).
        """
        path = Path(file_path)
        if not path.exists():
            return False, f"File Scarico Ore non trovato: {file_path}", []

        if not HAS_OPENPYXL:
            return False, "Modulo 'openpyxl' mancante.", []

        try:
            wb_data = cls._load_scarico_workbook(path)
            if "SCARICO ORE" not in wb_data.sheetnames:
                return False, "Foglio 'SCARICO ORÈ non trovato.", []

            ws_data = wb_data["SCARICO ORE"]
            rows = cls._process_all_scarico_rows(ws_data, progress_callback)

            return True, f"Importate {len(rows)} righe da Scarico Ore.", rows

        except Exception as e:
            return False, f"Errore importazione Scarico Ore: {e}", []

    @classmethod
    def _load_scarico_workbook(cls, path: Path) -> Any:
        from src.core.constants import Business  # noqa: PLC0415

        wb_file = io.BytesIO()
        is_encrypted = False

        if msoffcrypto:
            with suppress(Exception), path.open("rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=Business.DEFAULT_EXCEL_PASSWORD)  # nosec: B106
                office_file.decrypt(wb_file)
                is_encrypted = True

        if not is_encrypted:
            with path.open("rb") as f:
                wb_file.write(f.read())

        wb_file.seek(0)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            return openpyxl.load_workbook(
                wb_file,
                data_only=True,
                read_only=True,
                keep_vba=False,
                keep_links=False,
            )

    @classmethod
    def _process_all_scarico_rows(
        cls,
        ws: Any,
        progress_callback: Callable[[int, int], None] | None,
    ) -> list[tuple[Any, ...]]:
        start_row = 6
        col_keys = [
            "data",
            "pers1",
            "pers2",
            "odc",
            "pos",
            "dalle",
            "alle",
            "totale_ore",
            "descrizione",
            "finito",
            "commessa",
        ]
        total_rows: int = ws.max_row

        rows_to_insert: list[tuple[Any, ...]] = []
        rows_to_insert_append = rows_to_insert.append

        progress_interval = 5000
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, min_col=2, max_col=12, values_only=False),
            start=start_row,
        ):
            if progress_callback and row_idx % progress_interval == 0:
                progress_callback(row_idx, total_rows)

            db_row = cls._process_scarico_ore_row(row, col_keys)
            if db_row:
                rows_to_insert_append(db_row)

        return rows_to_insert

    @staticmethod
    def _fmt_excel_val(val: Any) -> str:
        """Formatta in modo intelligente i valori provenienti da openpyxl."""
        if val is None:
            return ""
        # Se  un numero intero rappresentato come float (comune in Excel), converti in int
        if isinstance(val, (float, int)) and float(val).is_integer():
            val = int(val)
        s = str(val)
        # Sostituisce newline e sequence di spazi con un singolo spazio
        return re.sub(r"\s+", " ", s).strip()

    @staticmethod
    def _fmt_excel_date(val: Any) -> str:
        """Tenta di convertire un valore cella in stringa data ISO."""
        if not val:
            return ""
        if hasattr(val, "strftime"):
            return str(val.strftime("%Y-%m-%d"))
        s = str(val).strip()
        return s.split(" ")[0] if " " in s else s

    @classmethod
    def _extract_row_values(cls, row: Any) -> list[str] | None:
        """Estrae e normalizza i valori di una riga di scarico ore."""
        (
            c_data,
            c_p1,
            c_p2,
            c_odc,
            c_pos,
            c_dalle,
            c_alle,
            c_tot,
            c_desc,
            c_fin,
            c_comm,
        ) = row[0:11]

        if c_odc.value is None and c_pos.value is None:
            return None

        # Estrazione Campi
        s_data = cls._fmt_excel_date(c_data.value)
        s_p1 = cls._fmt_excel_val(c_p1.value)
        s_p2 = cls._fmt_excel_val(c_p2.value)

        # Gestione OdC / Posizione (esclude zeri fittizi)
        def _clean_zero(v: Any) -> str:
            s = cls._fmt_excel_val(v)
            return "" if s in ("0", "0.0") else s

        s_odc = _clean_zero(c_odc.value)
        s_pos = _clean_zero(c_pos.value)

        # Altri Campi
        s_dalle = cls._fmt_excel_val(c_dalle.value)
        s_alle = cls._fmt_excel_val(c_alle.value)
        s_tot = cls._fmt_excel_val(c_tot.value)
        s_desc = cls._fmt_excel_val(c_desc.value)
        s_fin = cls._fmt_excel_val(c_fin.value)
        s_comm = _clean_zero(c_comm.value)

        return [
            s_data,
            s_p1,
            s_p2,
            s_odc,
            s_pos,
            s_dalle,
            s_alle,
            s_tot,
            s_desc,
            s_fin,
            s_comm,
        ]

    @classmethod
    def _process_scarico_ore_row(cls, row: Any, col_keys: list[str]) -> tuple[Any, ...] | None:
        vals = cls._extract_row_values(row)
        if not vals:
            return None

        if not cls._validate_scarico_row(vals):
            return None

        styles_json = cls._extract_row_styles(row, col_keys, vals)

        return (
            vals[0],
            vals[1],
            vals[2],
            vals[3],
            vals[4],
            vals[5],
            vals[6],
            vals[7],
            vals[8],
            vals[9],
            vals[10],
            styles_json,
        )

    @staticmethod
    def _validate_scarico_row(vals: list[str]) -> bool:
        if not vals[3] or not vals[4] or not vals[7]:
            return False
        return bool(vals[1] or vals[2])

    @staticmethod
    def _extract_row_styles(row: Any, col_keys: list[str], vals: list[str]) -> str:
        row_styles: dict[str, dict[str, str]] = {}
        for i, key in enumerate(col_keys):
            if vals[i] == "":
                continue

            cell = row[i]
            with suppress(AttributeError, TypeError):
                font = cell.font
                if font and font.color and font.color.type == "rgb":
                    rgb = str(font.color.rgb)
                    hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"  # noqa: PLR2004
                    if hex_code != "#000000":
                        row_styles.setdefault(key, {})["fg"] = hex_code

            with suppress(AttributeError, TypeError):
                fill = cell.fill
                if fill and fill.patternType == "solid":
                    start_color = fill.start_color
                    if start_color and start_color.type == "rgb":
                        rgb = str(start_color.rgb)
                        hex_code = f"#{rgb[2:]}" if len(rgb) > 6 else f"#{rgb}"  # noqa: PLR2004
                        if hex_code != "#000000" and hex_code != "#FFFFFF":  # noqa: PLR1714
                            row_styles.setdefault(key, {})["bg"] = hex_code

        return json.dumps(row_styles) if row_styles else ""
