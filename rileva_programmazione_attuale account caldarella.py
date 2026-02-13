# --- Importazioni Necessarie ---
import logging
import os
import time
from datetime import datetime, timedelta
from typing import Any

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

try:
    import win32com.client

    # Annotazione di tipo per l'oggetto COM
    win32com_client: Any | None = win32com.client
except ImportError:
    logging.error("Libreria 'pywin32' non trovata. Per favore, installala con: pip install pywin32")
    win32com_client = None

# --- CONFIGURAZIONE LOGGING E COSTANTI ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)-8s - %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)
URL = "https://safework.isab.com/"
EXCEL_FILE_PATH = r"\\192.168.11.251\Database_Tecnico_SMI\cartella strumentale condivisa\ALLEGRETTI\ATTIVITA_PROGRAMMATE.xlsm"
EXCEL_SHEET_CREDENTIALS = "Inserimento dati"
EXCEL_CELL_USERNAME = "P3"
EXCEL_CELL_PASSWORD = "Q3"
DOWNLOAD_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
# --- RIMOSSO: Il percorso del driver non è più necessario ---
# CHROMEDRIVER_PATH = r"C:\WebDriver\chromedriver.exe"
NOMI_RICHIEDENTI_DA_SELEZIONARE = [
    "Agusta Damiano",
    "Caldarella Ferdinando",
    "Messina Ivan",
    "Naselli Francesco",
    "Passanisi Domenico",
    "Barbagallo Giancarlo",
    "Prezzavento Manuel",
]
ID_PANNELLO_FILTRI = "pnlDataAttivita"
# --- MODIFICA QUI ---
FOGLI_DA_CONTROLLARE_PER_PDL = ["A1", "A2", "A3", "CTE", "BLENDING", "TAS", "IGCC"]
# --- FINE MODIFICA ---


# --- FUNZIONI HELPER ---


def leggi_valore_cella_excel(file_path, sheet_name, cell_address):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name]
        cell_value = sheet[cell_address].value
        return str(cell_value).strip() if cell_value is not None else ""
    except Exception as e:
        logger.error(f"ERRORE lettura cella {cell_address}: {e}")
        return None


def pulisci_e_inserisci_testo(driver, wait, by, value, testo_da_inserire, nome_campo_log):
    try:
        elemento_input = wait.until(EC.visibility_of_element_located((by, value)))
        elemento_input.click()
        time.sleep(0.3)
        elemento_input.send_keys(Keys.CONTROL + "a")
        time.sleep(0.3)
        elemento_input.send_keys(Keys.BACK_SPACE)
        if testo_da_inserire:
            time.sleep(0.2)
            elemento_input.send_keys(testo_da_inserire)
        logger.info(f"Testo '{testo_da_inserire}' inserito in '{nome_campo_log}'.")
    except Exception as e:
        logger.error(f"Impossibile inserire testo in '{nome_campo_log}': {e}")
        raise


def seleziona_richiedenti_con_ricerca(
    driver: webdriver.Chrome, wait: WebDriverWait, nomi_da_cercare: list[str]
) -> bool:
    try:
        dropdown_button = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//select[@id='fldIdRichiedente']/following-sibling::div[contains(@class,'ms-parent')]/button[contains(@class,'ms-choice')]",
                )
            )
        )
        dropdown_button.click()
        time.sleep(1)
        dropdown_container = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, "//div[contains(@class,'ms-drop') and contains(@style,'display: block')]")
            )
        )
        search_box = dropdown_container.find_element(By.XPATH, ".//input[@type='text']")
        nomi_selezionati_con_successo = []
        for nome in nomi_da_cercare:
            try:
                search_box.clear()
                search_box.send_keys(nome)
                time.sleep(0.8)
                opzione = dropdown_container.find_element(
                    By.XPATH, f".//label//span[normalize-space()='{nome}']"
                )
                driver.execute_script("arguments[0].click();", opzione)
                logger.info(f"  -> Selezionato: {nome}")
                nomi_selezionati_con_successo.append(nome)
                time.sleep(0.3)
            except (NoSuchElementException, TimeoutException):
                logger.error(f"  -> ATTENZIONE: Impossibile trovare o selezionare '{nome}' tramite ricerca.")
                dropdown_button.click()
                return False
        dropdown_button.click()
        time.sleep(0.5)
        return len(nomi_selezionati_con_successo) == len(nomi_da_cercare)
    except Exception as e:
        logger.error(f"Errore generale durante la selezione dei richiedenti: {e}", exc_info=True)
        try:
            if driver.find_element(
                By.XPATH, "//div[contains(@class,'ms-drop') and contains(@style,'display: block')]"
            ).is_displayed():
                driver.find_element(
                    By.XPATH,
                    "//select[@id='fldIdRichiedente']/following-sibling::div[contains(@class,'ms-parent')]/button[contains(@class,'ms-choice')]",
                ).click()
        except:
            pass
        return False


def attendi_scomparsa_overlay(driver, timeout_secondi=120):
    try:
        overlay_wait = WebDriverWait(driver, timeout_secondi)
        overlay_xpath = "//div[@id='GISWaitOverlay']"
        overlay_wait.until(EC.invisibility_of_element_located((By.XPATH, overlay_xpath)))
        logger.info(" -> Overlay 'GISWaitOverlay' scomparso.")
    except TimeoutException:
        logger.warning(
            f"Timeout ({timeout_secondi}s) attesa scomparsa di 'GISWaitOverlay'. Proseguo con cautela."
        )
    try:
        modale_attivo = WebDriverWait(driver, 3).until(
            EC.visibility_of_element_located(
                (By.XPATH, "//div[contains(@class, 'modal') and contains(@style, 'display: block')]")
            )
        )
        logger.info("Trovato un popup modale imprevisto. Tento di chiuderlo.")
        bottone_chiusura = modale_attivo.find_element(
            By.XPATH, ".//button[contains(text(), 'OK') or @data-dismiss='modal']"
        )
        bottone_chiusura.click()
        logger.info(" -> Popup modale chiuso.")
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element_located(
                (By.XPATH, "//div[contains(@class, 'modal') and contains(@style, 'display: block')]")
            )
        )
    except:
        pass
    time.sleep(1)


def attendi_caricamento_post_login(driver, wait_lungo):
    logger.info("Attesa di 5 secondi dopo il login per stabilizzazione pagina.")
    time.sleep(5)
    logger.info("Controllo per overlay e popup post-login.")
    attendi_scomparsa_overlay(driver, 120)
    time.sleep(2)


# --- NOTA: Questa funzione viene mantenuta per compatibilità ma non viene più utilizzata per Cerca/Esporta ---
def esegui_navigazione_con_tab(driver, numero_di_tab, nome_azione):
    logger.info(f"--- ESECUZIONE '{nome_azione}' CON NAVIGAZIONE TRAMITE TAB ---")
    try:
        driver.find_element(By.ID, ID_PANNELLO_FILTRI).click()
        time.sleep(1)
        elemento_attivo = driver.switch_to.active_element
        for i in range(numero_di_tab):
            elemento_attivo.send_keys(Keys.TAB)
            time.sleep(0.2)
        driver.switch_to.active_element.send_keys(Keys.ENTER)
        logger.info(f"Azione '{nome_azione}' eseguita.")
    except Exception as e:
        logger.error(f"ERRORE durante '{nome_azione}' con TAB: {e}")
        raise


def attendi_download_e_trova_file(cartella_download, nome_atteso, timeout_secondi):
    logger.info(f"Attesa del file '{nome_atteso}' in: {cartella_download}")
    percorso_file_atteso = os.path.join(cartella_download, nome_atteso)
    for _ in range(timeout_secondi):
        if not any(f.endswith(".crdownload") for f in os.listdir(cartella_download)):
            if os.path.exists(percorso_file_atteso):
                logger.info(f"File '{nome_atteso}' trovato.")
                return percorso_file_atteso
        time.sleep(1)
    logger.error(f"Timeout! Impossibile trovare '{nome_atteso}'.")
    return None


def calcola_settimana_lavorativa_corrente():
    oggi = datetime.now()
    lunedi = oggi - timedelta(days=oggi.weekday())
    return [lunedi + timedelta(days=i) for i in range(5)]


def pulisci_file_scaricati(dizionario_file):
    logger.info("\n--- Pulizia dei file temporanei di programmazione... ---")
    if not dizionario_file:
        logger.info("Nessun file da pulire.")
        return
    for percorso_file in dizionario_file.values():
        try:
            if os.path.exists(percorso_file):
                os.remove(percorso_file)
                logger.info(f"  -> File eliminato: {os.path.basename(percorso_file)}")
            else:
                logger.warning(f"  -> File non trovato per l'eliminazione: {os.path.basename(percorso_file)}")
        except Exception as e:
            logger.error(f"  -> Impossibile eliminare il file {os.path.basename(percorso_file)}: {e}")


# --- FASE 1: DOWNLOAD SINGOLO PER RANGE (CON LOGICA DI ATTESA ROBUSTA) ---


def fase_download_browser(data_inizio: datetime, data_fine: datetime) -> str | None:
    logger.info("====== INIZIO FASE GLOBALE DI DOWNLOAD BROWSER (RANGE UNICO) ======")
    max_tentativi = 10
    tentativi_falliti = 0
    file_scaricato_finale = None

    while not file_scaricato_finale and tentativi_falliti < max_tentativi:
        driver = None
        logger.info(f"\n--- TENTATIVO DI DOWNLOAD N.{tentativi_falliti + 1}/{max_tentativi} ---")
        try:
            # --- MODIFICA CHIAVE: Selenium ora gestisce il driver da solo ---
            options = Options()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--ignore-certificate-errors")
            options.add_argument("--disable-features=PasswordLeakDetection")

            prefs = {
                "download.default_directory": DOWNLOAD_DIR,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "credentials_enable_service": False,
                "profile.password_manager_enabled": False,
                "profile.password_manager_leak_detection": False,
            }
            options.add_experimental_option("prefs", prefs)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])

            logger.info("Avvio Chrome WebDriver (gestione automatica del driver)...")
            driver = webdriver.Chrome(options=options)

            wait_rapido = WebDriverWait(driver, 40)
            wait_medio = WebDriverWait(driver, 90)
            wait_lungo = WebDriverWait(driver, 300)
            username = leggi_valore_cella_excel(EXCEL_FILE_PATH, EXCEL_SHEET_CREDENTIALS, EXCEL_CELL_USERNAME)
            password = leggi_valore_cella_excel(EXCEL_FILE_PATH, EXCEL_SHEET_CREDENTIALS, EXCEL_CELL_PASSWORD)
            if not username or not password:
                raise ValueError("Username o password non trovati.")

            driver.get(URL)
            wait_medio.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='ms-choice']"))).click()
            wait_medio.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//div[contains(@class, 'ms-drop')]//span[normalize-space()='ISAB Sud']")
                )
            ).click()
            pulisci_e_inserisci_testo(driver, wait_medio, By.ID, "inpUtente", username, "Username")
            pulisci_e_inserisci_testo(driver, wait_medio, By.ID, "inpPassword", password, "Password")
            wait_medio.until(EC.element_to_be_clickable((By.ID, "btnLogin"))).click()
            logger.info("Login effettuato.")
            time.sleep(1)
            # --- FINE BLOCCO ---

            attendi_caricamento_post_login(driver, wait_lungo)

            attendi_scomparsa_overlay(driver, 60)
            wait_lungo.until(EC.element_to_be_clickable((By.ID, "topIcon-actHomePage"))).click()

            attendi_scomparsa_overlay(driver, 60)
            wait_lungo.until(EC.element_to_be_clickable((By.ID, "sideBar-actVisualizzaAttivita"))).click()

            logger.info("Attesa stabilizzazione pagina attività...")
            attendi_scomparsa_overlay(driver, 120)

            logger.info("--- IMPOSTAZIONE FILTRI STATICI (con logica di riprova ottimizzata) ---")

            # --- NUOVO: FILTRO DITTA ---
            try:
                logger.info("Impostazione filtro Ditta su CO.EMI SRL...")
                ditta_dropdown_button = wait_rapido.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//select[@id='fldIdDitta']/following-sibling::div/button")
                    )
                )
                ditta_dropdown_button.click()
                time.sleep(1)
                ditta_option = wait_rapido.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//div[contains(@class,'ms-drop')]//span[normalize-space()='CO.EMI SRL']")
                    )
                )
                ditta_option.click()
                logger.info(" -> Filtro Ditta impostato.")
            except Exception as e:
                logger.error(f"Impossibile impostare il filtro Ditta: {e}")
                raise
            # --- FINE FILTRO DITTA ---

            max_retry_selezioni = 3
            tutti_selezionati = False
            for tentativi_selezione in range(1, max_retry_selezioni + 1):
                logger.info(
                    f"Tentativo di selezione richiedenti N.{tentativi_selezione}/{max_retry_selezioni}..."
                )
                attendi_scomparsa_overlay(driver, 60)

                if seleziona_richiedenti_con_ricerca(driver, wait_rapido, NOMI_RICHIEDENTI_DA_SELEZIONARE):
                    tutti_selezionati = True
                    logger.info("Tutti i richiedenti sono stati selezionati con successo.")
                    break
                logger.warning(f"Tentativo N.{tentativi_selezione} fallito. Riprovo tra 3 secondi...")
                time.sleep(3)

            if not tutti_selezionati:
                raise Exception(
                    "CRITICO: Impossibile selezionare tutti i richiedenti necessari dopo vari tentativi."
                )

            try:
                icona_close_flag = driver.find_element(
                    By.XPATH,
                    "//select[@id='selFiltroFlag']/following-sibling::div[contains(@class,'ms-parent')]/button[contains(@class,'ms-choice')]//div[contains(@class, 'icon-close')]",
                )
                if icona_close_flag.is_displayed():
                    icona_close_flag.click()
                    time.sleep(1)
            except NoSuchElementException:
                pass

            # --- Download singolo per range ---
            data_formattata_dal = data_inizio.strftime("%d/%m/%Y")
            data_formattata_al = data_fine.strftime("%d/%m/%Y")
            logger.info(
                f"  -> Tentativo di download per il RANGE: {data_formattata_dal} - {data_formattata_al}"
            )

            attendi_scomparsa_overlay(driver, 60)
            pulisci_e_inserisci_testo(
                driver,
                wait_rapido,
                By.ID,
                "programmazioneDal",
                data_formattata_dal,
                f"Data Dal ({data_formattata_dal})",
            )
            pulisci_e_inserisci_testo(
                driver,
                wait_rapido,
                By.ID,
                "programmazioneAl",
                data_formattata_al,
                f"Data Al ({data_formattata_al})",
            )

            # --- MODIFICA: CLICK DIRETTO SU 'CERCA' (RIMOSSO TAB) ---
            logger.info("Esecuzione click diretto sul pulsante 'Cerca'...")
            try:
                btn_cerca = wait_rapido.until(EC.element_to_be_clickable((By.ID, "btnAvviaRicerca")))
                btn_cerca.click()
                logger.info("Click su 'Cerca' eseguito correttamente.")
            except Exception as e:
                logger.error(f"ERRORE cliccando su 'Cerca': {e}")
                raise

            logger.info("Attesa caricamento risultati della ricerca...")
            attendi_scomparsa_overlay(driver, 120)

            nome_file_programmazione = (
                f"Programmazione_{data_inizio.strftime('%Y%m%d')}_{data_fine.strftime('%Y%m%d')}.xlsx"
            )
            percorso_file_rinominato = os.path.join(DOWNLOAD_DIR, nome_file_programmazione)

            if os.path.exists(percorso_file_rinominato):
                try:
                    os.remove(percorso_file_rinominato)
                    logger.info(f"Vecchio file '{nome_file_programmazione}' eliminato.")
                except Exception as e_del:
                    logger.error(f"ATTENZIONE: Impossibile eliminare il vecchio file: {e_del}")

            attendi_scomparsa_overlay(driver, 60)

            # --- MODIFICA: CLICK DIRETTO SU 'ESPORTA' (RIMOSSO TAB) ---
            logger.info("Esecuzione click diretto sul pulsante 'Esporta'...")
            try:
                btn_esporta = wait_rapido.until(EC.element_to_be_clickable((By.ID, "btnEsporta")))
                btn_esporta.click()
                logger.info("Click su 'Esporta' eseguito correttamente.")
            except Exception as e:
                logger.error(f"ERRORE cliccando su 'Esporta': {e}")
                raise

            file_generico = attendi_download_e_trova_file(DOWNLOAD_DIR, "Programmazione.xlsx", 60)

            if file_generico:
                os.rename(file_generico, percorso_file_rinominato)
                file_scaricato_finale = percorso_file_rinominato
                logger.info(
                    f"    SUCCESSO! File per il range {data_formattata_dal}-{data_formattata_al} scaricato."
                )
            else:
                raise Exception(
                    f"Download fallito per il range {data_formattata_dal}-{data_formattata_al}. Riprovo l'intera sessione."
                )

        except Exception as e:
            logger.error(f"TENTATIVO N.{tentativi_falliti + 1} FALLITO. Causa: {e}", exc_info=False)
            tentativi_falliti += 1
            if tentativi_falliti < max_tentativi:
                logger.info("Attendo 5 secondi...")
                time.sleep(5)
        finally:
            if driver:
                driver.quit()
                logger.info("Browser chiuso per questo tentativo.")

    if not file_scaricato_finale:
        logger.critical(f"DOWNLOAD FALLITO. Impossibile scaricare il file dopo {max_tentativi} tentativi.")
        return None

    logger.info("====== DOWNLOAD COMPLETATO CON SUCCESSO PER IL RANGE ======")
    return file_scaricato_finale


# --- FASE 2 ed __main__ (modificate) ---
def get_excel_workbook(file_path):
    if not win32com_client:
        return None, None, False
    file_name = os.path.basename(file_path)
    try:
        excel_app = win32com_client.GetActiveObject("Excel.Application")
        for wb in excel_app.Workbooks:
            if wb.Name.lower() == file_name.lower():
                logger.info(
                    f"Trovato file '{file_name}' già aperto. Lo script si aggancerà a questa istanza."
                )
                return excel_app, wb, True
    except Exception:
        pass
    logger.info("Avvio di una nuova istanza di Excel in background.")
    excel_app = win32com_client.DispatchEx("Excel.Application")
    excel_app.Visible = False
    excel_app.DisplayAlerts = False
    workbook = excel_app.Workbooks.Open(file_path, UpdateLinks=0)
    return excel_app, workbook, False


def aggrega_e_applica_modifiche(file_scaricato: str):
    logger.info("\n====== INIZIO FASE GLOBALE DI ELABORAZIONE EXCEL (MODALITÀ OTTIMIZZATA) ======")
    if not file_scaricato:
        logger.warning("Nessun file da elaborare. La fase di download potrebbe essere fallita.")
        return

    excel_app = None
    wb_attivita = None
    era_gia_aperto = False
    original_calculation_mode = None
    xlCalculationManual = -4135
    xlCalculationAutomatic = -4105

    try:
        excel_app, wb_attivita, era_gia_aperto = get_excel_workbook(EXCEL_FILE_PATH)
        if not wb_attivita or not excel_app:
            raise Exception(f"Impossibile aprire il workbook '{EXCEL_FILE_PATH}'.")

        # --- INIZIO MODIFICA: Esegui la macro di pulizia ---
        logger.info("Esecuzione macro 'PulisciNomiDefiniti' per prevenire conflitti...")
        nome_file_excel = os.path.basename(EXCEL_FILE_PATH)
        excel_app.Run(f"'{nome_file_excel}'!PulisciNomiDefiniti")
        logger.info(" -> Macro 'PulisciNomiDefiniti' eseguita.")

        logger.info("Esecuzione macro 'RimuoviTuttiIFiltri'...")
        excel_app.Run(f"'{nome_file_excel}'!RimuoviTuttiIFiltri")
        logger.info(" -> Macro 'RimuoviTuttiIFiltri' eseguita.")

        excel_app.ScreenUpdating = False
        excel_app.EnableEvents = False
        original_calculation_mode = excel_app.Calculation
        excel_app.Calculation = xlCalculationManual
        # nome_file_excel = os.path.basename(EXCEL_FILE_PATH) # Già definito sopra

        logger.info("Esecuzione macro 'OrdinaEFormattaTabellaCorrente' prima della lettura...")
        excel_app.Run(f"'{nome_file_excel}'!OrdinaEFormattaTabellaCorrente")
        logger.info(" -> Macro 'OrdinaEFormattaTabellaCorrente' eseguita.")

        logger.info("Lettura stato iniziale direttamente dal file Excel aperto...")
        mappa_pdl_esistenti = {}
        for nome_foglio in FOGLI_DA_CONTROLLARE_PER_PDL:
            sheet = wb_attivita.Sheets(nome_foglio)
            last_row = sheet.Cells(sheet.Rows.Count, 5).End(-4162).Row  # xlUp
            if last_row >= 4:
                data = sheet.Range(sheet.Cells(4, 1), sheet.Cells(last_row, 13)).Value
                if data:
                    if not isinstance(data, tuple):
                        data = ((data,),)
                    for i, row in enumerate(data):
                        pdl_val = row[4]  # Colonna E
                        if pdl_val:
                            pdl_str = str(pdl_val).strip()
                            stato_val = row[12]  # Colonna M
                            mappa_pdl_esistenti[pdl_str] = {
                                "foglio": nome_foglio,
                                "riga": i + 4,
                                "stato": str(stato_val or "").strip().upper(),
                            }

        sheet_nuovi_pdl = wb_attivita.Sheets("nuovi PdL rilevati")
        last_row_nuovi_da_leggere = sheet_nuovi_pdl.Cells(sheet_nuovi_pdl.Rows.Count, 1).End(-4162).Row
        pdl_gia_rilevati = set()
        if last_row_nuovi_da_leggere >= 3:
            pdl_data = sheet_nuovi_pdl.Range(
                sheet_nuovi_pdl.Cells(3, 1), sheet_nuovi_pdl.Cells(last_row_nuovi_da_leggere, 1)
            ).Value
            if pdl_data:
                if not isinstance(pdl_data, tuple):
                    pdl_data = ((pdl_data,),)
                pdl_gia_rilevati = {str(row[0]).strip() for row in pdl_data if row[0]}

        logger.info(
            f"Stato letto. Trovati {len(mappa_pdl_esistenti)} PdL esistenti e {len(pdl_gia_rilevati)} già rilevati."
        )

        logger.info("Inizio aggregazione modifiche dal file scaricato...")
        nuovi_pdl_da_scrivere = {}
        modifiche_marcatura = {}
        modifiche_stato = {}

        # Mappatura indici colonne file sorgente (0-based)
        mappa_col_sorgente_X = {
            8: 2,  # Lunedì (H) -> C (2)
            9: 4,  # Martedì (I) -> E (4)
            10: 6,  # Mercoledì (J) -> G (6)
            11: 8,  # Giovedì (K) -> I (8)
            12: 10,  # Venerdì (L) -> K (10)
        }

        wb_prog = openpyxl.load_workbook(file_scaricato, read_only=True, data_only=True)
        sheet_prog = wb_prog.active

        for row in sheet_prog.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            pdl_str = str(row[0] or "").strip()
            if not pdl_str:
                continue

            pdl_info = mappa_pdl_esistenti.get(pdl_str)
            is_pdl_esistente = pdl_info is not None

            if not is_pdl_esistente:
                if pdl_str not in pdl_gia_rilevati and pdl_str not in nuovi_pdl_da_scrivere:
                    nuovi_pdl_da_scrivere[pdl_str] = [
                        row[0],
                        row[1],
                        row[14],
                        row[16],
                        row[18],
                        row[19],
                        row[20],
                        row[13],
                    ]

            if is_pdl_esistente:
                for col_dest, idx_sorg in mappa_col_sorgente_X.items():
                    val_sorg = str(row[idx_sorg] or "").strip().lower()
                    if val_sorg == "si":
                        modifiche_marcatura.setdefault(pdl_str, {})[col_dest] = "X"

            if is_pdl_esistente:
                col_O_val = str(row[14] or "").strip()
                is_richiesto = col_O_val in ["Richiesto", "Richiesto (Ese ok)"]
                if is_richiesto and pdl_info["stato"] != "RICHIESTO":
                    modifiche_stato[pdl_str] = "RICHIESTO"
                elif not is_richiesto and pdl_info["stato"] == "RICHIESTO":
                    modifiche_stato[pdl_str] = "EMESSO"

        wb_prog.close()

        logger.info("--- Riepilogo modifiche calcolate: ---")
        logger.info(f"    - Nuovi PdL da scrivere: {len(nuovi_pdl_da_scrivere)}")
        logger.info(f"    - PdL con marcature 'X' da aggiornare: {len(modifiche_marcatura)}")
        logger.info(f"    - PdL con stato da aggiornare: {len(modifiche_stato)}")

        logger.info("Esecuzione macro 'reset_programmazione' prima di applicare le modifiche...")
        excel_app.Run(f"'{nome_file_excel}'!reset_programmazione")
        logger.info(" -> Macro 'reset_programmazione' eseguita.")

        if not (nuovi_pdl_da_scrivere or modifiche_marcatura or modifiche_stato):
            logger.info("Nessuna nuova modifica da applicare. Il file è stato solo resettato.")
        else:
            logger.info("Applicazione delle modifiche calcolate sul file Excel...")
            if nuovi_pdl_da_scrivere:
                logger.info(
                    "Ricerca della prima riga vuota nel range 3-23 del foglio 'nuovi PdL rilevati'..."
                )
                riga_corrente = 24
                check_range = sheet_nuovi_pdl.Range("A3:A23").Value
                if check_range:
                    if not isinstance(check_range, tuple):
                        check_range = ((check_range,),)
                    for i, row in enumerate(check_range):
                        val = row[0]
                        if val is None or str(val).strip() == "":
                            riga_corrente = 3 + i
                            break

                logger.info(f" -> Scrittura a partire dalla riga: {riga_corrente}")
                rows_to_write = list(nuovi_pdl_da_scrivere.values())
                target_range = sheet_nuovi_pdl.Range(
                    sheet_nuovi_pdl.Cells(riga_corrente, 1),
                    sheet_nuovi_pdl.Cells(riga_corrente + len(rows_to_write) - 1, 8),
                )
                target_range.Value = rows_to_write
                logger.info(f" -> Inseriti {len(rows_to_write)} nuovi PdL.")
                sheet_nuovi_pdl.Columns.AutoFit()

            if modifiche_marcatura:
                for pdl, modifiche_giorno in modifiche_marcatura.items():
                    if pdl in mappa_pdl_esistenti:
                        loc = mappa_pdl_esistenti[pdl]
                        for col, val in modifiche_giorno.items():
                            wb_attivita.Sheets(loc["foglio"]).Cells(loc["riga"], col).Value = val

            if modifiche_stato:
                for pdl, stato in modifiche_stato.items():
                    if pdl in mappa_pdl_esistenti:
                        loc = mappa_pdl_esistenti[pdl]
                        wb_attivita.Sheets(loc["foglio"]).Cells(loc["riga"], 13).Value = stato
            logger.info("Tutte le modifiche sono state applicate.")

        logger.info("Salvataggio finale delle modifiche...")
        wb_attivita.Save()

    except Exception as e_excel:
        logger.critical(f"ERRORE CRITICO NELLA FASE DI ELABORAZIONE EXCEL: {e_excel}", exc_info=True)
    finally:
        if excel_app:
            logger.info("Ripristino impostazioni Excel e chiusura.")
            excel_app.Calculation = (
                original_calculation_mode if original_calculation_mode is not None else xlCalculationAutomatic
            )
            excel_app.EnableEvents = True
            excel_app.ScreenUpdating = True
            if wb_attivita and not era_gia_aperto:
                wb_attivita.Close(SaveChanges=False)
            if not era_gia_aperto:
                excel_app.Quit()
        logger.info("====== FINE FASE GLOBALE DI ELABORAZIONE EXCEL ======")


if __name__ == "__main__":
    date_da_elaborare = calcola_settimana_lavorativa_corrente()

    # --- MODIFICA: Prendi data inizio e fine ---
    data_inizio = date_da_elaborare[0]
    data_fine = date_da_elaborare[-1]
    logger.info(
        f"Verrà elaborato il range unico: {data_inizio.strftime('%d/%m/%Y')} - {data_fine.strftime('%d/%m/%Y')}"
    )

    file_scaricato_unico = fase_download_browser(data_inizio, data_fine)

    if file_scaricato_unico:
        aggrega_e_applica_modifiche(file_scaricato_unico)
        # Passa un dizionario per compatibilità con la funzione di pulizia
        pulisci_file_scaricati({"file_range": file_scaricato_unico})
        logger.info("\nElaborazione completata. Il file delle attività è stato aggiornato.")
    else:
        logger.error("La fase di download non è stata completata. L'elaborazione Excel non può iniziare.")
    # --- FINE MODIFICA ---

    # input("\n--- PROCESSO SCRIPT COMPLETATO --- \nPremi INVIO per chiudere questa finestra.")
    logger.info("\n--- PROCESSO SCRIPT COMPLETATO --- \nLa finestra si chiuderà automaticamente.")
    time.sleep(3)  # Pausa opzionale di 3 secondi per leggere l'ultimo messaggio
