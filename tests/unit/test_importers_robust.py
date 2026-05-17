import json
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest
from openpyxl.styles import Font

from src.core.importers.scarico_ore import ScaricoOreImporter


class TestImportersRobust:
    # --- Giornaliere Tests (Pandas based) ---

    @pytest.fixture
    def mock_giornaliera_df(self):
        """DataFrame simulato per Giornaliere già normalizzato."""
        data = {
            "data": ["2024-01-01", "2024-01-02"],
            "personale": ["Rossi", "Bianchi"],
            "descrizione": ["Lavoro", "Ferie"],
            "tcl": ["T1", "T2"],
            "odc": ["540012345", ""],
            "pdl": ["123", "456"],
            "inizio": ["08:00", "09:00"],
            "fine": ["17:00", "18:00"],
            "ore": [8, 8],
            "n_prev": ["5400123", "5400456"],
        }
        return pd.DataFrame(data)

    def test_process_single_giornaliera_success(self, mock_giornaliera_df, tmp_path):
        """Test processamento completo singola giornaliera usando la pipeline."""
        from src.core.processing.base import Pipeline
        from src.core.processing.giornaliere.steps import (
            EnrichGiornalieraStep,
            ReadGiornalieraStep,
        )

        path = tmp_path / "Giornaliera_2024.xlsx"

        # Setup context
        context = {
            "file_path": path,
            "file_obj": path,
            "year": 2024,
            "lookup_map": {"5400456": "540099999"},
            "success": True,
        }

        pipeline = Pipeline()

        # Mock ReadGiornalieraStep per iniettare il dataframe
        class MockReadStep(ReadGiornalieraStep):
            def execute(self, ctx):
                ctx["df"] = mock_giornaliera_df
                ctx["success"] = True

        pipeline.add_step(MockReadStep())
        pipeline.add_step(EnrichGiornalieraStep())

        result = pipeline.run(context)

        assert result.get("success") is True
        rows = result.get("rows", [])
        assert len(rows) == 2

        # Verifica mapping ODC (colonna 5, indice 5 nel record)
        # Record: year, data, personale, descrizione, tcl, odc, pdl, inizio, fine, ore, n_prev, nome_file
        # Indice odc è 5
        assert rows[0][5] == "540012345"
        assert rows[1][5] == "540099999"

    def test_giornaliera_normalize_columns(self):
        """Test normalizzazione nomi colonne tramite step dedicato."""
        from src.core.processing.giornaliere.steps import NormalizeGiornalieraStep

        df = pd.DataFrame({"  Data  ": [], "Personale": [], "Unknown": []})
        context = {"df": df, "success": True}

        NormalizeGiornalieraStep().execute(context)
        norm_df = context["df"]

        assert "data" in norm_df.columns
        assert "personale" in norm_df.columns
        assert "Unknown" in norm_df.columns

    def test_giornaliera_clean_data(self, mock_giornaliera_df):
        """Test pulizia dati (rimozione totali) tramite step dedicato."""
        from src.core.processing.giornaliere.steps import NormalizeGiornalieraStep

        # Aggiungiamo una riga "Totale" esplicita al DF normalizzato per verificare la rimozione
        df_con_totali = pd.concat(
            [mock_giornaliera_df, pd.DataFrame([{"data": "Totale", "ore": 16}])], ignore_index=True
        )

        context = {"df": df_con_totali, "success": True}
        NormalizeGiornalieraStep().execute(context)

        clean_df = context["df"]

        # Dovrebbero rimanere solo le 2 righe originali
        assert len(clean_df) == 2
        assert not clean_df["data"].str.contains("Totale", na=False).any()

    # --- Scarico Ore Tests (OpenPyXL based) ---

    @pytest.fixture
    def mock_scarico_wb(self):
        """Crea un workbook Excel reale in memoria."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SCARICO ORE"

        # Header (righe 1-5 ignorate)
        for i in range(1, 6):
            ws.cell(row=i, column=1, value="Header")

        # Riga 6: Dati validi
        # Col 2-12: Data, P1, P2, ODC, POS, Dalle, Alle, Tot, Desc, Fin, Comm
        row_valid = [
            None,
            "2024-01-01",
            "P1",
            "P2",
            "ODC1",
            "10",
            "8",
            "17",
            "8",
            "Desc",
            "S",
            "C",
        ]
        for col, val in enumerate(row_valid, start=1):
            if val:
                c = ws.cell(row=6, column=col, value=val)
                # Aggiungi stile
                if col == 5:  # ODC
                    c.font = Font(color="FF0000")  # Red

        # Riga 7: Dati invalidi (manca ODC e POS)
        row_invalid = [
            None,
            "2024-01-02",
            "P1",
            "",
            "",
            "",
            "8",
            "17",
            "8",
            "Desc",
            "S",
            "C",
        ]
        for col, val in enumerate(row_invalid, start=1):
            if val:
                ws.cell(row=7, column=col, value=val)

        return wb

    def test_scarico_ore_parsing_logic(self, mock_scarico_wb):
        """Test logica parsing Scarico Ore con workbook reale."""
        ws = mock_scarico_wb["SCARICO ORE"]

        rows = ScaricoOreImporter._process_all_scarico_rows(ws, None)

        assert len(rows) == 1  # Solo riga 6 valida
        valid_row = rows[0]

        # Verifica dati
        assert valid_row[0] == "2024-01-01"  # Data
        assert valid_row[3] == "ODC1"  # ODC

        # Verifica stili JSON
        styles = json.loads(valid_row[11])
        assert "odc" in styles
        assert styles["odc"]["fg"] == "#FF0000"

    def test_scarico_ore_import_file_not_found(self):
        """Test file non trovato."""
        success, msg, _ = ScaricoOreImporter.import_scarico_ore("non_existent.xlsx")
        assert success is False
        assert "non trovato" in msg

    @patch("src.core.importers.scarico_ore.msoffcrypto")
    def test_scan_encrypted_file(self, mock_crypto, tmp_path):
        """Test scansione file cifrato."""
        from src.core.constants import Business

        path = tmp_path / "protected.xlsx"
        path.touch()

        # Simuliamo decifratura
        mock_file = MagicMock()
        mock_crypto.OfficeFile.return_value = mock_file

        # Se msoffcrypto c'è, deve tentare di aprire
        ScaricoOreImporter.scan_scarico_ore_rows(str(path))
        mock_file.load_key.assert_called_with(password=Business.DEFAULT_EXCEL_PASSWORD)
        mock_file.decrypt.assert_called()
