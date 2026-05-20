"""
Tests for TimesheetProcessor.process_and_move.
Aims for 100% coverage and parity before refactoring.
"""

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from src.core.timesheet_processor import TimesheetProcessor


@pytest.fixture
def sample_timesheet(tmp_path):
    """Crea un file Excel timesheet di test."""
    path = tmp_path / "original_ts.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Imposta ODC nella cella A2
    ws["A2"] = "ODC123"

    # B column: POS (Same value to ensure len(pos_values) == 1)
    ws["B1"] = "OLD_POS"
    ws["B2"] = "10.0"
    ws["B3"] = "10.0"

    # Header row
    for col in range(1, 30):
        ws.cell(row=1, column=col, value=f"Header{col}")

    wb.save(path)
    return path


def test_process_and_move_file_not_found():
    success, msg = TimesheetProcessor.process_and_move(Path("non_existent.xlsx"), Path("dest"))
    assert success is False
    assert "non trovato" in msg


def test_process_and_move_missing_sheet(tmp_path):
    path = tmp_path / "wrong_sheet.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    success, msg = TimesheetProcessor.process_and_move(path, tmp_path / "dest")
    assert success is False
    assert "non trovato" in msg


def test_process_and_move_missing_odc(tmp_path):
    path = tmp_path / "testsource.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws["A2"] = ""
    wb.save(path)
    success, msg = TimesheetProcessor.process_and_move(path, tmp_path / "dest")
    assert success is False
    assert "mancante" in msg

def test_process_and_move_success_single_pos(sample_timesheet, tmp_path):
    dest_dir = tmp_path / "dest"
    success, _msg = TimesheetProcessor.process_and_move(sample_timesheet, dest_dir)

    assert success is True
    # ODC123 and POS 10 -> ODC123_10_TS.xlsx
    expected_path = dest_dir / "ODC123_10_TS.xlsx"
    assert expected_path.exists()
    assert not sample_timesheet.exists()


def test_process_and_move_success_multiple_pos(tmp_path):
    path = tmp_path / "multi_pos.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws["A2"] = "ODCMULTI"
    ws["B2"] = "10"
    ws["B3"] = "20"
    wb.save(path)

    dest_dir = tmp_path / "dest"
    success, _msg = TimesheetProcessor.process_and_move(path, dest_dir)

    assert success is True
    # Multiple POS -> ODCMULTI_TS.xlsx
    assert (dest_dir / "ODCMULTI_TS.xlsx").exists()


def test_process_and_move_conflict_handling(sample_timesheet, tmp_path):
    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()
    # Create conflict for ODC123_10_TS.xlsx
    conflict_path = dest_dir / "ODC123_10_TS.xlsx"
    conflict_path.write_text("dummy")

    success, _msg = TimesheetProcessor.process_and_move(sample_timesheet, dest_dir)
    assert success is True

    # It should have created a second file with timestamp
    files = list(dest_dir.glob("ODC123_10_TS*.xlsx"))
    assert len(files) >= 2


def test_process_and_move_mkdir_error(tmp_path):
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws["A2"] = "5400123"  # ODC
    ws["B2"] = "10"  # POS
    wb.save(src)
    with patch("src.core.processing.timesheet.steps.Path.mkdir", side_effect=PermissionError("Perm error")):
        success, msg = TimesheetProcessor.process_and_move(src, tmp_path / "new_dir")
        assert success is False
        assert "Impossibile creare dest_dir" in msg


def test_process_and_move_critical_exception(sample_timesheet, tmp_path):
    with patch("openpyxl.load_workbook", side_effect=Exception("Memory Error")):
        success, msg = TimesheetProcessor.process_and_move(sample_timesheet, tmp_path / "dest")
        assert success is False
        assert "Memory Error" in msg
