from pathlib import Path

import pandas as pd
import pytest

from src.core.timesheet_processor import TimesheetProcessor


class TestTimesheetProcessorCoverage:
    """Test suite for src/core/timesheet_processor.py"""

    @pytest.fixture
    def sample_df(self):
        """Dataframe sample for testing."""
        return pd.DataFrame({"POS": ["POS1", "POS2", "POS1"], "Ore": [8, 4, 4]})

    def test_process_file_not_found(self, tmp_path):
        """Test elaborazione file non esistente."""
        path = Path("non_existent_file.xlsx")
        success, msg = TimesheetProcessor.process_and_move(path, dest_dir=tmp_path)
        assert success is False
        assert "non trovato" in msg

    def test_process_file_no_sheet(self, tmp_path):
        """Test file senza foglio Timesheet."""
        file_path = tmp_path / "test_no_sheet.xlsx"
        # Crea un excel con un foglio diverso
        df = pd.DataFrame({"A": [1]})
        with pd.ExcelWriter(file_path) as writer:
            df.to_excel(writer, sheet_name="Altro", index=False)

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is False
        assert "Foglio 'Timesheet' non trovato" in msg

    def test_process_file_with_pos(self, tmp_path, sample_df):
        """Test elaborazione corretta con colonna POS."""
        file_path = tmp_path / "test_ok.xlsx"

        # Use openpyxl directly to create the file with the required structure
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC_TEST"
        # Write POS header and data
        ws["B1"] = "POS"
        ws["B2"] = "POS1"
        ws["B3"] = "POS2"
        wb.save(file_path)

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is True
        assert "Salvato in" in msg

    def test_process_file_no_pos_column(self, tmp_path):
        """Test file con Timesheet ma senza colonna POS."""
        file_path = tmp_path / "test_no_pos.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC_TEST"
        ws["C1"] = "Data"
        wb.save(file_path)

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is True
        assert "Salvato in" in msg

    def test_process_file_exception(self, tmp_path):
        """Test gestione eccezioni (es. file corrotto o non leggibile)."""
        file_path = tmp_path / "corrupted.xlsx"
        file_path.write_text("Not an excel file")  # Scrive testo, openpyxl fallirà

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is False
        # msg conterrà l'errore di openpyxl
        assert msg is not None
