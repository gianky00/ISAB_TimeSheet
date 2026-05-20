import openpyxl
import pytest

from src.core.processing.timesheet.steps import ExtractMetadataStep, SaveWorkbookStep
from src.core.timesheet_processor import TimesheetProcessor


class TestTimesheetProcessor:
    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        """Crea un file Excel di test con la struttura attesa."""
        path = tmp_path / "source.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"

        # Metadata
        ws["A2"] = "5400123"  # ODC

        # Headers
        ws["B1"] = "OLD_B"
        ws["C1"] = "OLD_C"

        # Dati POS
        ws["B2"] = "10.0"
        ws["B3"] = "10.0"

        wb.save(path)
        return path

    def test_process_and_move_success(self, sample_xlsx, tmp_path):
        """Verifica il ciclo completo di elaborazione e rinomina."""
        dest_dir = tmp_path / "processed"

        success, _msg = TimesheetProcessor.process_and_move(sample_xlsx, dest_dir)

        assert success is True
        # Nome file atteso: ODC_POS_TS.xlsx (perché c'è solo un POS '10')
        expected_name = "5400123_10_TS.xlsx"
        dest_path = dest_dir / expected_name
        assert dest_path.exists()
        assert not sample_xlsx.exists(), "Il sorgente deve essere rimosso"

        # Verifica trasformazioni
        wb = openpyxl.load_workbook(dest_path)
        ws = wb["Timesheet"]
        # Dopo eliminazione colonne:
        # Orig B (POS) -> A
        # Orig C (Data) -> B
        assert ws["B1"].value == "Data"
        assert ws["A1"].value == "POS"

    def test_analyze_pos_column_logic(self):
        """Verifica il conteggio dei POS univoci tramite ExtractMetadataStep."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "10"
        ws["B3"] = "20"
        ws["B4"] = "10.0"

        step = ExtractMetadataStep()
        context = {"worksheet": ws}
        # Aggiungiamo ODC per non far fallire lo step
        ws["A2"] = "123"

        step.execute(context)
        metadata = context["metadata"]

        assert len(metadata.pos_values) == 3
        assert metadata.first_pos_cleaned == "10"

    def test_clean_pos_value(self):
        step = ExtractMetadataStep()
        assert step._clean_pos_value("10.0") == "10"
        assert step._clean_pos_value("5") == "5"
        assert step._clean_pos_value("abc") == "abc"

    def test_save_workbook_logic_conflict(self, tmp_path):
        """Verifica la gestione del conflitto tramite SaveWorkbookStep."""
        from src.models.timesheet import TimesheetMetadata

        dest_dir = tmp_path
        odc = "5400123"
        (dest_dir / f"{odc}_TS.xlsx").touch()

        wb = openpyxl.Workbook()
        metadata = TimesheetMetadata(odc=odc, pos_values={"10", "20"}, first_pos_cleaned="10")

        step = SaveWorkbookStep()
        context = {"workbook": wb, "dest_dir": dest_dir, "metadata": metadata}

        step.execute(context)
        path = context["dest_path"]

        assert path.name.startswith(f"{odc}_TS_")
        assert path.suffix == ".xlsx"
        assert path.exists()
