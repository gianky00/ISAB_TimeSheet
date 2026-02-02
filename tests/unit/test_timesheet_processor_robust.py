from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from src.core.timesheet_processor import TimesheetProcessor


class TestTimesheetProcessorRobust:
    @pytest.fixture
    def sample_workbook(self, tmp_path):
        """Crea un file Excel di test valido."""
        file_path = tmp_path / "test_source.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"

        # Setup dati minimi
        ws["A2"] = "ODC123"  # ODC
        ws["B1"] = "OldHeader"
        ws["B2"] = "10.0"  # POS da pulire
        ws["B3"] = "10.0"  # Stesso valore, per non attivare multi-pos

        # Colonne da eliminare per verificare la logica (es. colonna A, D, E...)
        # Popoliamo un po' la riga 1
        for i in range(1, 30):
            ws.cell(row=1, column=i, value=f"Col{i}")

        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.fixture
    def dest_dir(self, tmp_path):
        d = tmp_path / "processed"
        d.mkdir()
        return d

    def test_process_and_move_success_single_pos(self, sample_workbook, dest_dir):
        """Test happy path con un singolo POS."""
        success, msg = TimesheetProcessor.process_and_move(sample_workbook, dest_dir)

        print(f"DEBUG: Success={success}, Msg={msg}")  # DEBUG
        assert success is True
        assert "Salvato in" in msg

        # Verifica nome file atteso: ODC123_10_TS.xlsx
        expected_file = dest_dir / "ODC123_10_TS.xlsx"
        # Se fallisce, elenchiamo i file nella dir per capire cosa ha creato
        if not expected_file.exists():
            print(f"FILES IN DEST: {list(dest_dir.glob('*'))}")

        assert expected_file.exists()

        # Verifica che il sorgente sia stato rimosso
        assert not sample_workbook.exists()

    def test_process_and_move_multiple_pos(self, tmp_path, dest_dir):
        """Test happy path con POS multipli -> Nome file generico."""
        file_path = tmp_path / "multi_pos.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC999"
        ws["B2"] = "10"
        ws["B3"] = "20"  # Secondo POS diverso
        wb.save(file_path)
        wb.close()

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir)

        assert success is True
        expected_file = dest_dir / "ODC999_TS.xlsx"
        assert expected_file.exists()

    def test_file_not_found(self, dest_dir):
        """Test file inesistente."""
        success, msg = TimesheetProcessor.process_and_move(
            Path("non_existent.xlsx"), dest_dir
        )
        assert success is False
        assert "non trovato" in msg

    def test_missing_sheet(self, tmp_path, dest_dir):
        """Test foglio 'Timesheet' mancante."""
        file_path = tmp_path / "wrong_sheet.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WrongName"
        wb.save(file_path)
        wb.close()

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir)
        assert success is False
        assert "Foglio 'Timesheet' non trovato" in msg

    def test_missing_odc(self, tmp_path, dest_dir):
        """Test ODC mancante in cella A2."""
        file_path = tmp_path / "no_odc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = None  # ODC vuoto
        wb.save(file_path)
        wb.close()

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir)
        assert success is False
        assert "Valore ODC" in msg

    def test_transformations_applied(self, sample_workbook, dest_dir):
        """Verifica che le trasformazioni (header, delete cols) siano avvenute."""
        TimesheetProcessor.process_and_move(sample_workbook, dest_dir)

        dest_file = dest_dir / "ODC123_10_TS.xlsx"
        wb = openpyxl.load_workbook(dest_file)
        ws = wb["Timesheet"]

        # Verifica Header Rinominate e Spostate
        # B1 -> "POS", poi cancello A -> A1 diventa "POS"
        assert ws["A1"].value == "POS"

        # C1 -> "Data", poi cancello A -> B1 diventa "Data"
        assert ws["B1"].value == "Data"

        # N1 -> "Ing". Colonne eliminate prima di N: A, D, E, F, G, H, I, L (8 colonne)
        # N(14) - 8 = 6 -> F
        assert ws["F1"].value == "Ing"

        # Verifica Pulizia colonna B originale (ora A)
        # B2 -> 10.0 -> 10. Cancello A -> A2
        assert str(ws["A2"].value) == "10"

        wb.close()

    def test_filename_collision(self, sample_workbook, dest_dir):
        """Test gestione collisione nomi file (timestamp)."""
        # Crea un file esistente che confligge
        (dest_dir / "ODC123_10_TS.xlsx").touch()

        # Esegue processo (dovrebbe creare file con timestamp)
        with patch("time.strftime", return_value="20230101-120000"):
            success, msg = TimesheetProcessor.process_and_move(
                sample_workbook, dest_dir
            )

        assert success is True
        expected_file = dest_dir / "ODC123_10_TS_20230101-120000.xlsx"
        assert expected_file.exists()

    def test_permission_error_dest_dir(self, sample_workbook):
        """Test errore permessi creazione directory."""
        # Usiamo un path che sicuramente fallisce o mockiamo mkdir
        with patch("pathlib.Path.mkdir", side_effect=PermissionError("Access Denied")):
            success, msg = TimesheetProcessor.process_and_move(
                sample_workbook, Path("/root/protected")
            )
            assert success is False
            assert "Impossibile creare dest_dir" in msg

    def test_clean_pos_value(self):
        """Unit test per metodo helper _clean_pos_value."""
        assert TimesheetProcessor._clean_pos_value("10.0") == "10"
        assert TimesheetProcessor._clean_pos_value("10") == "10"
        assert TimesheetProcessor._clean_pos_value("abc") == "abc"
        assert (
            TimesheetProcessor._clean_pos_value("10.5") == "10"
        )  # int conversion truncate? Yes: int(float("10.5")) -> 10
        assert TimesheetProcessor._clean_pos_value("") == ""
