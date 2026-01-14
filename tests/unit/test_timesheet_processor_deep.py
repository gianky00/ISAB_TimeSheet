
from pathlib import Path

import pandas as pd
import pytest

from src.core.timesheet_processor import TimesheetProcessor


class TestTimesheetProcessorDeep:
    @pytest.fixture
    def mock_excel(self, tmp_path):
        """Crea un file Excel finto per i test."""
        file_path = tmp_path / "timesheet.xlsx"
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC_TEST"
        # Scrivi header POS in B1 e dati
        ws["B1"] = "POS"
        ws["B2"] = "10"
        ws["B3"] = "20"
        ws["B4"] = "10"
        ws["B5"] = "30"
        wb.save(file_path)
        return file_path

    def test_process_file_success(self, mock_excel, tmp_path):
        """Verifica elaborazione corretta con conteggio POS."""
        success, msg = TimesheetProcessor.process_and_move(mock_excel, dest_dir=tmp_path)
        assert success is True
        assert "Salvato in" in msg

    def test_process_file_missing_sheet(self, tmp_path):
        """Verifica errore se il foglio 'Timesheet' manca."""
        file_path = tmp_path / "wrong.xlsx"
        df = pd.DataFrame({"A": [1]})
        df.to_excel(file_path, sheet_name="Sheet1")

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is False
        assert "non trovato" in msg

    def test_process_file_no_pos_column(self, tmp_path):
        """Verifica comportamento se la colonna POS manca."""
        file_path = tmp_path / "no_pos.xlsx"
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC_TEST"
        ws["C1"] = "Data"
        wb.save(file_path)

        success, msg = TimesheetProcessor.process_and_move(file_path, dest_dir=tmp_path)
        assert success is True
        assert "Salvato in" in msg

    def test_process_file_not_found(self, tmp_path):
        """Verifica errore se il file non esiste."""
        success, msg = TimesheetProcessor.process_and_move(Path("ghost.xlsx"), dest_dir=tmp_path)
        assert success is False
        assert "non trovato" in msg
