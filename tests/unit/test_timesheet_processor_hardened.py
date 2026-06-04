"""Hardened tests for TimesheetProcessor.
Verifies Excel transformation logic and robustness.
"""

import openpyxl
import pytest

from src.application.services.timesheet_processor import TimesheetProcessor


class TestTimesheetProcessorHardened:
    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        """Crea un file Excel minimale compatibile con SafeWork."""
        path = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"

        # Metadata
        ws["A2"] = "ODC12345"
        ws["B2"] = "10.0"

        # Headers originali (alcuni)
        ws["B1"] = "Original POS"
        ws["C1"] = "Original Date"

        # Data rows
        ws["B3"] = "10.0"
        ws["C3"] = "2026-01-01"

        wb.save(path)
        return path

    def test_process_and_move_success(self, sample_xlsx, tmp_path):
        """Verifica il flusso completo di trasformazione e spostamento."""
        dest_dir = tmp_path / "output"

        success, msg = TimesheetProcessor.process_and_move(sample_xlsx, dest_dir)

        assert success is True
        assert "ODC12345_10_TS.xlsx" in msg

        # Verifica file creato
        dest_file = dest_dir / "ODC12345_10_TS.xlsx"
        assert dest_file.exists()

        # Verifica trasformazioni (Headers)
        wb = openpyxl.load_workbook(dest_file)
        ws = wb.active
        # Notare che le colonne vengono eliminate, quindi le lettere cambiano
        # TimesheetProcessor elimina: AC, Z, X, L, I, H, G, F, E, D, A
        # B1 (POS) era la 2, dopo eliminazione di A(1) diventa la 1? No, delete_cols lavora su indici.
        # Proviamo a verificare se POS è presente come header in una delle prime celle
        headers = [ws.cell(row=1, column=i).value for i in range(1, 5)]
        assert "POS" in headers
        assert "Data" in headers

        # Verifica pulizia POS (10.0 -> 10)
        # La cella B2 (row 2, col 2) originale conteneva 10.0
        # Dopo eliminazione col A, dovrebbe essere in colonna 1
        pos_val = ws.cell(row=2, column=1).value
        assert pos_val == 10
        assert not isinstance(pos_val, float)

    def test_missing_sheet_failure(self, tmp_path):
        """Verifica fallimento se il foglio 'Timesheet' manca."""
        path = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        wb.save(path)

        success, msg = TimesheetProcessor.process_and_move(path, tmp_path / "out")
        assert not success
        assert "non trovato" in msg

    def test_missing_odc_failure(self, tmp_path):
        """Verifica fallimento se ODC (A2) è vuoto."""
        path = tmp_path / "no_odc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = None  # ODC mancante
        wb.save(path)

        success, msg = TimesheetProcessor.process_and_move(path, tmp_path / "out")
        assert not success
        assert "ODC" in msg

    def test_multiple_pos_filename(self, tmp_path):
        """Verifica che con più POS il nome file non contenga il numero POS."""
        path = tmp_path / "multi_pos.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws["A2"] = "ODC999"
        ws["B2"] = "10"
        ws["B3"] = "20"  # Secondo POS
        wb.save(path)

        dest_dir = tmp_path / "out"
        success, msg = TimesheetProcessor.process_and_move(path, dest_dir)

        assert success
        assert "ODC999_TS.xlsx" in msg  # Niente "_10_" o "_20_"
